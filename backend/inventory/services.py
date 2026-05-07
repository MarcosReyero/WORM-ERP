import json
import logging
import re
import unicodedata
import urllib.error
import urllib.request
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from uuid import uuid4

from django.conf import settings
from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.core.mail import EmailMessage, send_mail
from django.core.validators import validate_email
from django.db import transaction
from django.db.models import Count, Q, Sum
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.utils.dateparse import parse_time
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel

from accounts.models import UserProfile
from communications.services import (
    active_message_contacts,
    serialize_contact,
    list_inventory_alarms,
)

from .automation import (
    TASK_KEY_FULL_STOCK_REPORT,
    TASK_KEY_MINIMUM_STOCK_DIGEST,
    TASK_KEY_MINIMUM_STOCK_RECONCILE,
    TASK_KEY_SCHEDULER,
    ensure_automation_task_states,
    get_automation_task_state,
    get_full_stock_report_due_context,
    get_minimum_stock_digest_due_context,
    mark_full_stock_report_result,
    mark_minimum_stock_digest_result,
    serialize_automation_task_state,
)
from .models import (
    Article,
    ArticleCategory,
    AssetCheckout,
    FullStockReportConfig,
    InternalRequest,
    InternalRequestLine,
    InventoryBalance,
    InventoryBatch,
    Location,
    MinimumStockDigestConfig,
    MinimumStockAlarmConfig,
    MinimumStockAlarmState,
    Pallet,
    PalletEvent,
    PersonalDailyReport,
    Person,
    PhysicalCountSession,
    SafetyStockAlertRule,
    Sector,
    StockDiscrepancy,
    StockMovement,
    StatusCatalog,
    Supplier,
    TrackedUnit,
    UnitOfMeasure,
)


DIGEST_AUTOMATION_LOGGER = logging.getLogger("inventory.automation.digest")
FULL_STOCK_REPORT_AUTOMATION_LOGGER = logging.getLogger("inventory.automation.full_stock_report")


class InventoryApiError(Exception):
    def __init__(self, detail, status=400):
        """Inicializa la instancia."""
        super().__init__(detail)
        self.detail = detail
        self.status = status


MASTER_ROLES = {
    UserProfile.Role.ADMINISTRATOR,
    UserProfile.Role.STOREKEEPER,
    UserProfile.Role.MAINTENANCE,
}
MOVEMENT_ROLES = {
    UserProfile.Role.ADMINISTRATOR,
    UserProfile.Role.STOREKEEPER,
    UserProfile.Role.MAINTENANCE,
}
CHECKOUT_ROLES = {
    UserProfile.Role.ADMINISTRATOR,
    UserProfile.Role.STOREKEEPER,
    UserProfile.Role.MAINTENANCE,
}
COUNT_ROLES = {
    UserProfile.Role.ADMINISTRATOR,
    UserProfile.Role.STOREKEEPER,
    UserProfile.Role.SUPERVISOR,
    UserProfile.Role.MAINTENANCE,
}
APPROVER_ROLES = {
    UserProfile.Role.ADMINISTRATOR,
    UserProfile.Role.SUPERVISOR,
}
ALARM_ROLES = {
    UserProfile.Role.ADMINISTRATOR,
    UserProfile.Role.STOREKEEPER,
    UserProfile.Role.SUPERVISOR,
    UserProfile.Role.MAINTENANCE,
    UserProfile.Role.PURCHASING,
}

ARTICLE_CODE_PREFIXES = {
    Article.ArticleType.CONSUMABLE: "CON",
    Article.ArticleType.INPUT: "INS",
    Article.ArticleType.TOOL: "HER",
    Article.ArticleType.SPARE_PART: "REP",
    Article.ArticleType.PPE: "EPP",
}

ARTICLE_TYPE_ALIASES = {
    "consumable": Article.ArticleType.CONSUMABLE,
    "consumible": Article.ArticleType.CONSUMABLE,
    "input": Article.ArticleType.INPUT,
    "insumo": Article.ArticleType.INPUT,
    "insumo productivo": Article.ArticleType.INPUT,
    "tool": Article.ArticleType.TOOL,
    "herramienta": Article.ArticleType.TOOL,
    "spare_part": Article.ArticleType.SPARE_PART,
    "spare part": Article.ArticleType.SPARE_PART,
    "repuesto": Article.ArticleType.SPARE_PART,
    "ppe": Article.ArticleType.PPE,
    "epp": Article.ArticleType.PPE,
    "elemento de proteccion personal": Article.ArticleType.PPE,
}


def parse_json(request):
    """Parsea json."""
    if not request.body:
        return {}

    try:
        return json.loads(request.body)
    except json.JSONDecodeError as exc:
        raise InventoryApiError("Invalid JSON payload") from exc


def serialize_datetime(value):
    """Maneja serialize datetime."""
    return value.isoformat() if value else None


def serialize_date(value):
    """Maneja serialize date."""
    return value.isoformat() if value else None


def serialize_decimal(value):
    """Maneja serialize decimal."""
    if value is None:
        return None
    return float(value)


def serialize_internal_request_line(line):
    """Maneja serialize internal request line."""
    return {
        "id": line.id,
        "article_id": line.article_id,
        "article": line.article.name,
        "internal_code": line.article.internal_code,
        "quantity_requested": serialize_decimal(line.quantity_requested),
        "quantity_delivered": serialize_decimal(line.quantity_delivered),
        "notes": line.notes,
    }


def serialize_internal_request(request_item):
    """Maneja serialize internal request."""
    lines = list(request_item.lines.select_related("article").all())
    return {
        "id": request_item.id,
        "request_number": request_item.request_number,
        "requested_at": serialize_datetime(request_item.requested_at),
        "requester_id": request_item.requester_id,
        "requester": request_item.requester.full_name,
        "requesting_sector_id": request_item.requesting_sector_id,
        "requesting_sector": request_item.requesting_sector.name,
        "status": request_item.status,
        "status_label": request_item.get_status_display(),
        "notes": request_item.notes,
        "closed_at": serialize_datetime(request_item.closed_at),
        "reject_reason": request_item.reject_reason,
        "line_count": len(lines),
        "quantity_requested_total": serialize_decimal(
            sum((line.quantity_requested or 0) for line in lines)
        ),
        "quantity_delivered_total": serialize_decimal(
            sum((line.quantity_delivered or 0) for line in lines)
        ),
        "lines": [serialize_internal_request_line(line) for line in lines],
    }


def list_internal_requests(filters=None):
    """Lista internal requests."""
    filters = filters or {}
    query = clean_casefold(filters.get("q") or "")
    status_filter = clean_casefold(filters.get("status") or "all")

    queryset = InternalRequest.objects.select_related(
        "requester",
        "requesting_sector",
    ).prefetch_related("lines", "lines__article")

    if status_filter and status_filter != "all":
        queryset = queryset.filter(status=status_filter)

    if query:
        queryset = queryset.filter(
            Q(request_number__icontains=query)
            | Q(requester__full_name__icontains=query)
            | Q(requesting_sector__name__icontains=query)
        )

    items = list(queryset.order_by("-requested_at", "-id")[:250])
    return [serialize_internal_request(item) for item in items]


def create_internal_request(payload):
    """Crea internal request."""
    if not isinstance(payload, dict):
        raise InventoryApiError("Invalid payload")

    requester_id = payload.get("requester_id")
    sector_id = payload.get("requesting_sector_id")
    notes = clean_string(payload.get("notes"))
    lines_payload = payload.get("lines") or []

    if not requester_id or not sector_id:
        raise InventoryApiError("requester_id and requesting_sector_id are required")
    if not isinstance(lines_payload, list) or not lines_payload:
        raise InventoryApiError("lines must be a non-empty list")

    requester = get_object_or_404(Person, pk=requester_id)
    requesting_sector = get_object_or_404(Sector, pk=sector_id)

    with transaction.atomic():
        placeholder = f"TEMP-{uuid4()}"
        request_item = InternalRequest.objects.create(
            request_number=placeholder,
            requester=requester,
            requesting_sector=requesting_sector,
            status=InternalRequest.RequestStatus.PENDING,
            notes=notes,
        )
        request_item.request_number = f"REQ-{timezone.localdate().strftime('%Y%m%d')}-{request_item.id:06d}"
        request_item.save(update_fields=["request_number"])

        for entry in lines_payload:
            if not isinstance(entry, dict):
                continue
            article_id = entry.get("article_id")
            quantity_requested = entry.get("quantity_requested")
            if not article_id or quantity_requested in (None, ""):
                continue
            article = get_object_or_404(Article, pk=article_id)
            try:
                qty = Decimal(str(quantity_requested))
            except (InvalidOperation, TypeError) as exc:
                raise InventoryApiError("Invalid quantity_requested") from exc
            if qty <= 0:
                raise InventoryApiError("quantity_requested must be greater than zero")

            InternalRequestLine.objects.create(
                request=request_item,
                article=article,
                quantity_requested=qty,
                notes=clean_string(entry.get("notes")),
            )

    return serialize_internal_request(
        InternalRequest.objects.select_related("requester", "requesting_sector")
        .prefetch_related("lines", "lines__article")
        .get(pk=request_item.pk)
    )


def get_profile(user):
    """Devuelve profile."""
    defaults = {
        "role": UserProfile.Role.ADMINISTRATOR
        if user.is_superuser
        else UserProfile.Role.OPERATOR,
    }
    return UserProfile.objects.get_or_create(user=user, defaults=defaults)[0]


def require_role(user, allowed_roles):
    """Maneja require role."""
    profile = get_profile(user)
    if profile.status != UserProfile.Status.ACTIVE:
        raise InventoryApiError("User profile is inactive", status=403)
    if user.is_superuser or profile.role in allowed_roles:
        return profile
    raise InventoryApiError("You do not have permission for this action", status=403)


def parse_decimal(value, field_name):
    """Parsea decimal."""
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError):
        raise InventoryApiError(f"Invalid decimal for {field_name}")


def parse_optional_decimal(value):
    """Parsea optional decimal."""
    if value in (None, "", []):
        return None
    return parse_decimal(value, "decimal")


def parse_optional_int(value, field_name):
    """Parsea optional int."""
    if value in (None, "", []):
        return None

    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise InventoryApiError(f"Invalid integer for {field_name}") from exc


def parse_boolean(value):
    """Parsea boolean."""
    if isinstance(value, bool):
        return value
    if value in (None, ""):
        return False
    if isinstance(value, str):
        return value.strip().lower() in {"true", "1", "yes", "on"}
    return bool(value)


def resolve_instance(model, value, label, required=True):
    """Maneja resolve instance."""
    if value in (None, "", []):
        if required:
            raise InventoryApiError(f"{label} is required")
        return None
    return get_object_or_404(model, pk=value)


def update_audit(instance, user, is_new=False):
    """Actualiza audit."""
    if hasattr(instance, "updated_by"):
        instance.updated_by = user
    if is_new and hasattr(instance, "created_by") and not instance.created_by_id:
        instance.created_by = user


def save_validated(instance):
    """Guarda validated."""
    try:
        instance.full_clean()
    except ValidationError as exc:
        raise InventoryApiError(exc.message_dict or exc.messages)
    instance.save()
    return instance


def get_default_location():
    """Devuelve default location."""
    return Location.objects.filter(code="DEP-PRINCIPAL").first() or Location.objects.order_by("id").first()


def choose_tracking_mode(article_type, payload_tracking_mode):
    """Maneja choose tracking mode."""
    if payload_tracking_mode:
        return payload_tracking_mode
    if article_type == Article.ArticleType.TOOL:
        return Article.TrackingMode.UNIT
    return Article.TrackingMode.QUANTITY


def should_require_minimum(article_type, is_critical):
    """Maneja should require minimum."""
    return article_type in {
        Article.ArticleType.CONSUMABLE,
        Article.ArticleType.INPUT,
    } or (article_type == Article.ArticleType.SPARE_PART and is_critical)


def clean_string(value):
    """Maneja clean string."""
    return (value or "").strip()


def clean_casefold(value):
    """Maneja clean casefold."""
    return clean_string(value).casefold()


def normalize_search_text(value):
    """Maneja normalize search text."""
    normalized = unicodedata.normalize("NFD", str(value or ""))
    ascii_only = normalized.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", ascii_only).casefold().strip()


def build_search_target(values):
    """Construye search target."""
    return normalize_search_text(" ".join(str(value) for value in values if value))


def matches_normalized_query(target, query):
    """Maneja matches normalized query."""
    normalized_query = normalize_search_text(query)
    if not normalized_query:
        return True
    return all(term in target for term in normalized_query.split())


def parse_email_list(value):
    """Parsea email list."""
    raw_value = str(value or "")
    if not raw_value.strip():
        return []

    emails = []
    seen = set()
    for item in re.split(r"[,\n;]+", raw_value):
        email = clean_string(item).lower()
        if not email:
            continue
        try:
            validate_email(email)
        except ValidationError as exc:
            raise InventoryApiError(f"Invalid email address: {email}") from exc
        if email in seen:
            continue
        seen.add(email)
        emails.append(email)
    return emails


def split_email_list(value):
    """Maneja split email list."""
    raw_value = str(value or "")
    if not raw_value.strip():
        return [], []

    valid_emails = []
    invalid_emails = []
    seen = set()
    for item in re.split(r"[,\n;]+", raw_value):
        email = clean_string(item).lower()
        if not email:
            continue
        try:
            validate_email(email)
        except ValidationError:
            invalid_emails.append(email)
            continue
        if email in seen:
            continue
        seen.add(email)
        valid_emails.append(email)
    return valid_emails, invalid_emails


def parse_time_or_error(value, field_name, default=None):
    """Parsea time or error."""
    raw_value = clean_string(value)
    if not raw_value:
        return default
    parsed = parse_time(raw_value)
    if not parsed:
        raise InventoryApiError(f"El valor de {field_name} no es una hora valida.")
    return parsed.replace(second=0, microsecond=0)


def parse_weekday_or_error(value, field_name="run_weekday", default=0):
    """Parsea weekday or error."""
    if value in (None, "", []):
        return default
    weekday = parse_optional_int(value, field_name)
    if weekday is None or weekday < 0 or weekday > 6:
        raise InventoryApiError("El dia de envio debe estar entre 0 y 6.")
    return weekday


def article_code_prefix(article_type):
    """Maneja article code prefix."""
    return ARTICLE_CODE_PREFIXES.get(article_type, "ART")


def generate_article_code(article_type):
    """Maneja generate article code."""
    prefix = article_code_prefix(article_type)
    pattern = re.compile(rf"^{re.escape(prefix)}-(\d+)$")
    max_number = 0

    for code in Article.objects.filter(internal_code__startswith=f"{prefix}-").values_list(
        "internal_code", flat=True
    ):
        match = pattern.match(code)
        if match:
            max_number = max(max_number, int(match.group(1)))

    next_number = max_number + 1
    while True:
        candidate = f"{prefix}-{next_number:06d}"
        if not Article.objects.filter(internal_code=candidate).exists():
            return candidate
        next_number += 1


def resolve_article_type(value):
    """Maneja resolve article type."""
    normalized = clean_casefold(value)
    if not normalized:
        raise InventoryApiError("article_type is required")

    if normalized in ARTICLE_TYPE_ALIASES:
        return ARTICLE_TYPE_ALIASES[normalized]

    for candidate, label in Article.ArticleType.choices:
        if normalized in {candidate.casefold(), label.casefold()}:
            return candidate

    raise InventoryApiError(f"Unsupported article type: {value}")


def resolve_choice_value(value, choices, label, default=None):
    """Maneja resolve choice value."""
    raw_value = clean_string(value)
    if not raw_value:
        return default

    lowered = raw_value.casefold()
    for candidate, choice_label in choices:
        if lowered in {candidate.casefold(), choice_label.casefold()}:
            return candidate

    raise InventoryApiError(f"Unsupported {label}: {value}")


def resolve_catalog_by_name_or_code(queryset, value, label):
    """Maneja resolve catalog by name or code."""
    raw_value = clean_string(value)
    if not raw_value:
        raise InventoryApiError(f"{label} is required")

    lowered = raw_value.casefold()
    for item in queryset:
        code = clean_string(getattr(item, "code", "")).casefold()
        name = clean_string(getattr(item, "name", "")).casefold()
        if lowered in {code, name}:
            return item

    raise InventoryApiError(f"{label} '{raw_value}' does not exist")


def resolve_optional_catalog_by_name_or_code(queryset, value):
    """Maneja resolve optional catalog by name or code."""
    raw_value = clean_string(value)
    if not raw_value:
        return None

    lowered = raw_value.casefold()
    for item in queryset:
        code = clean_string(getattr(item, "code", "")).casefold()
        name = clean_string(getattr(item, "name", "")).casefold()
        if lowered in {code, name}:
            return item

    return None


def get_or_create_category_by_name(name, user, parent=None):
    """Devuelve or create category by name."""
    raw_name = clean_string(name)
    if not raw_name:
        return None

    category = ArticleCategory.objects.filter(name__iexact=raw_name, parent=parent).first()
    if category:
        return category

    category = ArticleCategory(name=raw_name, parent=parent, status=StatusCatalog.ACTIVE)
    update_audit(category, user, is_new=True)
    save_validated(category)
    return category


def article_payload_value(payload, key, default=None):
    """Maneja article payload value."""
    if hasattr(payload, "get"):
        value = payload.get(key, default)
        if value is None:
            return default
        return value
    return default


def normalize_excel_header(value):
    """Maneja normalize excel header."""
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    ascii_only = normalized.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", "_", ascii_only.strip().lower()).strip("_")


def get_or_create_balance(article, location, user, batch=None):
    """Devuelve or create balance."""
    balance, created = InventoryBalance.objects.get_or_create(
        article=article,
        location=location,
        batch=batch,
        defaults={
            "created_by": user,
            "updated_by": user,
        },
    )
    if not created:
        update_audit(balance, user)
    return balance


def apply_balance_delta(article, location, quantity_delta, user, batch=None):
    """Maneja apply balance delta."""
    balance = get_or_create_balance(article, location, user, batch=batch)
    new_value = balance.on_hand + quantity_delta
    if new_value < 0:
        raise InventoryApiError(
            f"Not enough stock at {location.name} for {article.name}",
            status=400,
        )
    balance.on_hand = new_value
    save_validated(balance)
    return balance


def next_unit_tag(article, index):
    """Maneja next unit tag."""
    return f"{article.internal_code}-{index:03d}"


def create_tracked_units(article, quantity, user, location=None, sector=None, notes="", status=None):
    """Crea tracked units."""
    if quantity <= 0:
        return []

    current_total = TrackedUnit.objects.filter(article=article).count()
    created_units = []
    for offset in range(1, quantity + 1):
        unit = TrackedUnit(
            article=article,
            internal_tag=next_unit_tag(article, current_total + offset),
            status=status or TrackedUnit.UnitStatus.AVAILABLE,
            current_location=location,
            current_sector=sector or article.sector_responsible,
            notes=notes,
        )
        update_audit(unit, user, is_new=True)
        save_validated(unit)
        created_units.append(unit)
    return created_units


def current_stock_maps():
    """Maneja current stock maps."""
    quantity_map = {
        row["article"]: row["total"] or Decimal("0")
        for row in InventoryBalance.objects.values("article").annotate(total=Sum("on_hand"))
    }
    available_quantity_map = {
        row["article"]: (row["on_hand"] or Decimal("0")) - (row["reserved"] or Decimal("0"))
        for row in InventoryBalance.objects.values("article").annotate(
            on_hand=Sum("on_hand"),
            reserved=Sum("reserved"),
        )
    }
    unit_total_map = {
        row["article"]: row["total"]
        for row in TrackedUnit.objects.exclude(
            status__in=[TrackedUnit.UnitStatus.RETIRED, TrackedUnit.UnitStatus.LOST]
        )
        .values("article")
        .annotate(total=Count("id"))
    }
    unit_available_map = {
        row["article"]: row["total"]
        for row in TrackedUnit.objects.filter(status=TrackedUnit.UnitStatus.AVAILABLE)
        .values("article")
        .annotate(total=Count("id"))
    }
    return quantity_map, available_quantity_map, unit_total_map, unit_available_map


def article_current_stock(article, quantity_map, unit_total_map):
    """Maneja article current stock."""
    if article.tracking_mode == Article.TrackingMode.UNIT:
        return Decimal(unit_total_map.get(article.id, 0))
    return quantity_map.get(article.id, Decimal("0"))


def article_available_stock(article, available_quantity_map, unit_available_map):
    """Maneja article available stock."""
    if article.tracking_mode == Article.TrackingMode.UNIT:
        return Decimal(unit_available_map.get(article.id, 0))
    return available_quantity_map.get(article.id, Decimal("0"))


def serialize_article(article, quantity_map, available_quantity_map, unit_total_map, unit_available_map):
    """Maneja serialize article."""
    current_stock = article_current_stock(article, quantity_map, unit_total_map)
    available_stock = article_available_stock(article, available_quantity_map, unit_available_map)
    low_stock = article.minimum_stock is not None and current_stock <= article.minimum_stock
    return {
        "id": article.id,
        "internal_code": article.internal_code,
        "name": article.name,
        "image_url": article.image.url if article.image else None,
        "article_type": article.article_type,
        "article_type_label": article.get_article_type_display(),
        "tracking_mode": article.tracking_mode,
        "tracking_mode_label": article.get_tracking_mode_display(),
        "status": article.status,
        "status_label": article.get_status_display(),
        "description": article.description,
        "sector_responsible_id": article.sector_responsible_id,
        "sector_responsible": article.sector_responsible.name,
        "primary_location_id": article.primary_location_id,
        "primary_location": article.primary_location.name if article.primary_location else None,
        "category_id": article.category_id,
        "category": article.category.name if article.category else None,
        "subcategory_id": article.subcategory_id,
        "subcategory": article.subcategory.name if article.subcategory else None,
        "minimum_stock": serialize_decimal(article.minimum_stock),
        "safety_stock": serialize_decimal(article.safety_stock),
        "reorder_point": serialize_decimal(article.reorder_point),
        "max_stock": serialize_decimal(article.max_stock),
        "suggested_purchase_qty": serialize_decimal(article.suggested_purchase_qty),
        "current_stock": serialize_decimal(current_stock),
        "available_stock": serialize_decimal(available_stock),
        "low_stock": low_stock,
        "loanable": article.loanable,
        "requires_lot": article.requires_lot,
        "requires_expiry": article.requires_expiry,
        "requires_serial": article.requires_serial,
        "requires_size": article.requires_size,
        "requires_quality": article.requires_quality,
        "requires_assignee": article.requires_assignee,
        "is_critical": article.is_critical,
        "supplier_id": article.supplier_id,
        "supplier": article.supplier.name if article.supplier else None,
        "supplier_availability_days": article.supplier.availability_days if article.supplier else None,
        "reference_price": serialize_decimal(article.reference_price),
        "lead_time_days": article.lead_time_days,
        "availability_days": (
            article.supplier.availability_days
            if article.supplier and article.supplier.availability_days is not None
            else article.lead_time_days
        ),
        "last_purchase": serialize_date(article.last_purchase),
        "unit_of_measure": {
            "id": article.unit_of_measure_id,
            "code": article.unit_of_measure.code,
            "name": article.unit_of_measure.name,
        },
        "observations": article.observations,
    }


def active_alarm_recipients(user):
    """Maneja active alarm recipients."""
    current_user = user if user and user.is_authenticated else None
    if not current_user:
        return []
    return active_message_contacts(current_user, include_current=True)


def current_stock_for_article(article):
    """Maneja current stock for article."""
    quantity_map, _available_quantity_map, unit_total_map, _unit_available_map = current_stock_maps()
    return article_current_stock(article, quantity_map, unit_total_map)


def _system_person():
    """
    Usuario solicitante para eventos automaticos.

    No usamos FK directa a User en InternalRequest; en modo automatico necesitamos
    garantizar que exista una Persona disponible.
    """
    system_code = "SYSTEM-AUTO"
    person = Person.objects.filter(employee_code=system_code).first()
    if person:
        return person
    return Person.objects.create(
        full_name="Sistema",
        employee_code=system_code,
        status=StatusCatalog.ACTIVE,
    )


def _purchase_quantity_for_minimum(article, current_stock):
    """Maneja purchase quantity for minimum."""
    candidates = [
        article.suggested_purchase_qty,
        (article.max_stock - current_stock) if article.max_stock is not None else None,
        (article.reorder_point - current_stock) if article.reorder_point is not None else None,
    ]
    for candidate in candidates:
        if candidate is None:
            continue
        try:
            value = Decimal(str(candidate))
        except (InvalidOperation, TypeError):
            continue
        if value > 0:
            return value
    return Decimal("1")


def maybe_create_purchase_request_for_minimum_stock(
    article,
    *,
    previous_stock=None,
    current_stock=None,
    triggered_by_user=None,
    requester_person=None,
    requester_sector=None,
    source_label="stock_minimum",
):
    """
    Crea una solicitud de compra (InternalRequest) cuando el stock cruza por
    debajo (o igual) del stock minimo.

    Garantias:
    - Idempotente: si ya existe una solicitud abierta para el articulo, no crea otra
    - Seguro ante concurrencia: bloquea el articulo con SELECT FOR UPDATE
    """
    if article.minimum_stock is None:
        return None

    minimum_stock = article.minimum_stock
    if current_stock is None:
        current_stock = current_stock_for_article(article)

    if current_stock != minimum_stock:
        return None
    if previous_stock is not None and previous_stock <= minimum_stock:
        return None

    requesting_sector = requester_sector or article.sector_responsible
    if not requesting_sector:
        return None

    open_statuses = {
        InternalRequest.RequestStatus.DRAFT,
        InternalRequest.RequestStatus.PENDING,
        InternalRequest.RequestStatus.APPROVED,
        InternalRequest.RequestStatus.PARTIAL,
    }

    with transaction.atomic():
        locked_article = (
            Article.objects.select_for_update()
            .select_related("sector_responsible")
            .get(pk=article.pk)
        )
        if locked_article.minimum_stock is None:
            return None
        minimum_stock = locked_article.minimum_stock
        current_stock = current_stock_for_article(locked_article)
        if current_stock != minimum_stock:
            return None

        already_requested = InternalRequestLine.objects.filter(
            article=locked_article,
            request__status__in=open_statuses,
        ).exists()
        if already_requested:
            return None

        resolved_requester = requester_person
        if not resolved_requester:
            resolved_requester = (
                Person.objects.filter(
                    status=StatusCatalog.ACTIVE,
                    sector=requesting_sector,
                )
                .order_by("id")
                .first()
            )
        if not resolved_requester:
            resolved_requester = _system_person()

        qty = _purchase_quantity_for_minimum(locked_article, current_stock)
        notes_parts = [
            "Generada automaticamente por stock minimo.",
            f"Articulo: {locked_article.internal_code}",
            f"Stock actual: {serialize_decimal(current_stock)}",
            f"Stock minimo: {serialize_decimal(minimum_stock)}",
            f"Origen: {source_label}",
        ]
        if triggered_by_user is not None:
            notes_parts.append(f"Registrado por: {getattr(triggered_by_user, 'username', '')}")

        placeholder = f"TEMP-{uuid4()}"
        request_item = InternalRequest.objects.create(
            request_number=placeholder,
            requester=resolved_requester,
            requesting_sector=requesting_sector,
            status=InternalRequest.RequestStatus.PENDING,
            notes="\n".join([part for part in notes_parts if part]),
            created_by=triggered_by_user if triggered_by_user else None,
            updated_by=triggered_by_user if triggered_by_user else None,
        )
        request_item.request_number = (
            f"REQ-{timezone.localdate().strftime('%Y%m%d')}-{request_item.id:06d}"
        )
        request_item.save(update_fields=["request_number"])

        line = InternalRequestLine(
            request=request_item,
            article=locked_article,
            quantity_requested=qty,
            notes="",
        )
        update_audit(line, triggered_by_user, is_new=True)
        save_validated(line)

    return request_item

def safety_alert_email_addresses(alert):
    """Maneja safety alert email addresses."""
    recipient_emails = []
    seen = set()
    recipient_users = list(alert.recipients.select_related("profile__sector_default").all())

    for user in recipient_users:
        profile = get_profile(user)
        email = clean_string(user.email).lower()
        if profile.status != UserProfile.Status.ACTIVE or not email:
            continue
        if email in seen:
            continue
        seen.add(email)
        recipient_emails.append(email)

    valid_extra_emails, _invalid_extra_emails = split_email_list(alert.additional_emails)
    for email in valid_extra_emails:
        if email in seen:
            continue
        seen.add(email)
        recipient_emails.append(email)

    return recipient_emails


def serialize_safety_alert_rule(alert, quantity_map=None, unit_total_map=None):
    """Maneja serialize safety alert rule."""
    recipients = list(alert.recipients.select_related("profile__sector_default").order_by("first_name", "last_name", "username"))
    if quantity_map is None or unit_total_map is None:
        quantity_map, _available_quantity_map, unit_total_map, _unit_available_map = current_stock_maps()
    current_stock = article_current_stock(alert.article, quantity_map, unit_total_map)
    return {
        "id": alert.id,
        "article_id": alert.article_id,
        "article_name": alert.article.name,
        "article_code": alert.article.internal_code,
        "article_type": alert.article.article_type,
        "article_type_label": alert.article.get_article_type_display(),
        "is_enabled": alert.is_enabled,
        "status": alert.status,
        "status_label": alert.get_status_display(),
        "current_stock": serialize_decimal(current_stock),
        "minimum_stock": serialize_decimal(alert.article.minimum_stock),
        "safety_stock": serialize_decimal(alert.article.minimum_stock),
        "triggered": alert.status == SafetyStockAlertRule.AlertStatus.TRIGGERED,
        "recipients": [serialize_contact(user) for user in recipients],
        "additional_emails": alert.additional_emails,
        "additional_email_list": split_email_list(alert.additional_emails)[0],
        "notify_email": alert.notify_email,
        "notify_telegram": alert.notify_telegram,
        "notes": alert.notes,
        "last_stock_value": serialize_decimal(alert.last_stock_value),
        "triggered_at": serialize_datetime(alert.triggered_at),
        "resolved_at": serialize_datetime(alert.resolved_at),
        "last_notified_at": serialize_datetime(alert.last_notified_at),
        "last_email_error": alert.last_email_error,
        "last_telegram_error": alert.last_telegram_error,
    }


def serialize_minimum_stock_alarm_config(config):
    """Maneja serialize minimum stock alarm config."""
    recipients = list(
        config.recipients.select_related("profile__sector_default").order_by(
            "first_name",
            "last_name",
            "username",
        )
    )
    return {
        "id": config.id,
        "key": config.key,
        "is_enabled": config.is_enabled,
        "notify_email": config.notify_email,
        "notify_telegram": config.notify_telegram,
        "recipients": [serialize_contact(user) for user in recipients],
        "additional_emails": config.additional_emails,
        "additional_email_list": split_email_list(config.additional_emails)[0],
        "notes": config.notes,
        "last_notified_at": serialize_datetime(config.last_notified_at),
        "last_email_error": config.last_email_error,
        "last_telegram_error": config.last_telegram_error,
    }


def get_purchasing_minimum_stock_alarm_config(user):
    """Devuelve purchasing minimum stock alarm config."""
    require_role(user, ALARM_ROLES)
    config, _created = MinimumStockAlarmConfig.objects.get_or_create(key="purchasing_default")
    return serialize_minimum_stock_alarm_config(config)


def save_purchasing_minimum_stock_alarm_config(user, payload):
    """Guarda purchasing minimum stock alarm config."""
    require_role(user, ALARM_ROLES)
    if not isinstance(payload, dict):
        raise InventoryApiError("Invalid payload")

    recipient_ids = payload.get("recipient_user_ids") or []
    if isinstance(recipient_ids, str):
        recipient_ids = [item for item in re.split(r"[,\s]+", recipient_ids) if item]

    is_enabled = parse_boolean(payload.get("is_enabled"))
    notify_email = parse_boolean(payload.get("notify_email")) if "notify_email" in payload else True
    notify_telegram = parse_boolean(payload.get("notify_telegram")) if "notify_telegram" in payload else False

    if is_enabled and not (notify_email or notify_telegram):
        raise InventoryApiError("Selecciona al menos un canal de notificacion (email o Telegram).")

    recipients = validate_alarm_rule_recipients(
        recipient_ids,
        require_email=notify_email,
        require_telegram=notify_telegram,
    )

    additional_emails = payload.get("additional_emails") or ""
    validated_extra_emails = parse_email_list(additional_emails) if notify_email else []

    if is_enabled and notify_email and not (recipients or validated_extra_emails):
        raise InventoryApiError("Selecciona al menos un destinatario o agrega un email adicional.")
    if is_enabled and notify_telegram and not recipients:
        raise InventoryApiError("Selecciona al menos un destinatario con Telegram configurado.")

    with transaction.atomic():
        config, created = MinimumStockAlarmConfig.objects.get_or_create(
            key="purchasing_default",
            defaults={
                "created_by": user,
                "updated_by": user,
            },
        )
        config.is_enabled = is_enabled
        config.notify_email = notify_email
        config.notify_telegram = notify_telegram
        config.additional_emails = "\n".join(validated_extra_emails) if notify_email else ""
        config.notes = payload.get("notes") or ""
        update_audit(config, user, is_new=created)
        save_validated(config)
        config.recipients.set(recipients)

    refreshed = (
        MinimumStockAlarmConfig.objects.prefetch_related("recipients__profile__sector_default")
        .get(pk=config.pk)
    )
    return serialize_minimum_stock_alarm_config(refreshed)


def low_stock_articles_snapshot(serialized_articles=None):
    """Maneja low stock articles snapshot."""
    articles = serialized_articles if serialized_articles is not None else list_articles()
    return [article for article in articles if article.get("low_stock")]


WEEKDAY_LABELS = {
    0: "Lunes",
    1: "Martes",
    2: "Miercoles",
    3: "Jueves",
    4: "Viernes",
    5: "Sabado",
    6: "Domingo",
}


def serialize_digest_run_time(value):
    """Maneja serialize digest run time."""
    if not value:
        return "08:00"
    return value.strftime("%H:%M")


def resolve_digest_delivery_tone(status):
    """Maneja resolve digest delivery tone."""
    if status == MinimumStockDigestConfig.DeliveryStatus.SUCCESS:
        return "ok"
    if status == MinimumStockDigestConfig.DeliveryStatus.WARNING:
        return "low"
    if status == MinimumStockDigestConfig.DeliveryStatus.ERROR:
        return "out"
    if status == MinimumStockDigestConfig.DeliveryStatus.SKIPPED:
        return "low"
    return "out"


def serialize_inventory_automation_status():
    """Maneja serialize inventory automation status."""
    ensure_automation_task_states()
    return {
        "scheduler": serialize_automation_task_state(get_automation_task_state(TASK_KEY_SCHEDULER)),
        "minimum_stock_reconcile": serialize_automation_task_state(
            get_automation_task_state(TASK_KEY_MINIMUM_STOCK_RECONCILE)
        ),
        "minimum_stock_digest": serialize_automation_task_state(
            get_automation_task_state(TASK_KEY_MINIMUM_STOCK_DIGEST)
        ),
        "full_stock_report": serialize_automation_task_state(
            get_automation_task_state(TASK_KEY_FULL_STOCK_REPORT)
        ),
    }


def serialize_minimum_stock_digest_config(config=None, serialized_articles=None, save_warning=""):
    """Maneja serialize minimum stock digest config."""
    low_stock_articles = low_stock_articles_snapshot(serialized_articles)
    recipients = []
    next_run_at = None
    run_weekday_label = WEEKDAY_LABELS.get(0)
    last_delivery_status = MinimumStockDigestConfig.DeliveryStatus.NEVER
    if config is not None:
        recipients = list(
            config.recipients.select_related("profile__sector_default").order_by(
                "first_name",
                "last_name",
                "username",
            )
        )
        next_run_at = get_minimum_stock_digest_due_context(config).get("next_run_at")
        run_weekday_label = WEEKDAY_LABELS.get(config.run_weekday, WEEKDAY_LABELS[0])
        last_delivery_status = config.last_delivery_status

    return {
        "id": config.id if config else None,
        "is_enabled": config.is_enabled if config else False,
        "frequency": config.frequency if config else MinimumStockDigestConfig.Frequency.DAILY,
        "frequency_label": (
            config.get_frequency_display()
            if config
            else MinimumStockDigestConfig.Frequency.DAILY.label
        ),
        "recipients": [serialize_contact(user) for user in recipients],
        "additional_emails": config.additional_emails if config else "",
        "additional_email_list": split_email_list(config.additional_emails)[0] if config else [],
        "notes": config.notes if config else "",
        "run_at": serialize_digest_run_time(config.run_at) if config else "08:00",
        "run_weekday": config.run_weekday if config else 0,
        "run_weekday_label": run_weekday_label,
        "next_run_at": serialize_datetime(next_run_at),
        "last_notified_at": serialize_datetime(config.last_notified_at) if config else None,
        "last_email_error": config.last_email_error if config else "",
        "last_period_key": config.last_period_key if config else "",
        "inflight_period_key": config.inflight_period_key if config else "",
        "inflight_started_at": serialize_datetime(config.inflight_started_at) if config else None,
        "last_delivery_status": last_delivery_status,
        "last_delivery_status_label": (
            config.get_last_delivery_status_display()
            if config
            else MinimumStockDigestConfig.DeliveryStatus.NEVER.label
        ),
        "last_delivery_tone": resolve_digest_delivery_tone(last_delivery_status),
        "last_recipient_warning": config.last_recipient_warning if config else "",
        "last_summary_count": config.last_summary_count if config else None,
        "save_warning": save_warning,
        "low_stock_count": len(low_stock_articles),
        "preview_articles": [
            {
                "id": article["id"],
                "name": article["name"],
                "internal_code": article["internal_code"],
                "current_stock": article["current_stock"],
                "minimum_stock": article["minimum_stock"],
            }
            for article in low_stock_articles[:6]
        ],
    }


def full_stock_articles_snapshot(serialized_articles=None):
    """Maneja full stock articles snapshot."""
    return serialized_articles if serialized_articles is not None else list_articles()


def serialize_full_stock_report_config(config=None, serialized_articles=None, save_warning=""):
    """Maneja serialize full stock report config."""
    articles = full_stock_articles_snapshot(serialized_articles)
    recipients = []
    next_run_at = None
    run_weekday_label = WEEKDAY_LABELS.get(0)
    last_delivery_status = FullStockReportConfig.DeliveryStatus.NEVER
    if config is not None:
        recipients = list(
            config.recipients.select_related("profile__sector_default").order_by(
                "first_name",
                "last_name",
                "username",
            )
        )
        next_run_at = get_full_stock_report_due_context(config).get("next_run_at")
        run_weekday_label = WEEKDAY_LABELS.get(config.run_weekday, WEEKDAY_LABELS[0])
        last_delivery_status = config.last_delivery_status

    return {
        "id": config.id if config else None,
        "is_enabled": config.is_enabled if config else False,
        "frequency": config.frequency if config else FullStockReportConfig.Frequency.DAILY,
        "frequency_label": (
            config.get_frequency_display()
            if config
            else FullStockReportConfig.Frequency.DAILY.label
        ),
        "recipients": [serialize_contact(user) for user in recipients],
        "additional_emails": config.additional_emails if config else "",
        "additional_email_list": split_email_list(config.additional_emails)[0] if config else [],
        "notes": config.notes if config else "",
        "run_at": serialize_digest_run_time(config.run_at) if config else "08:00",
        "run_weekday": config.run_weekday if config else 0,
        "run_weekday_label": run_weekday_label,
        "next_run_at": serialize_datetime(next_run_at),
        "last_notified_at": serialize_datetime(config.last_notified_at) if config else None,
        "last_email_error": config.last_email_error if config else "",
        "last_period_key": config.last_period_key if config else "",
        "inflight_period_key": config.inflight_period_key if config else "",
        "inflight_started_at": serialize_datetime(config.inflight_started_at) if config else None,
        "last_delivery_status": last_delivery_status,
        "last_delivery_status_label": (
            config.get_last_delivery_status_display()
            if config
            else FullStockReportConfig.DeliveryStatus.NEVER.label
        ),
        "last_delivery_tone": resolve_digest_delivery_tone(last_delivery_status),
        "last_recipient_warning": config.last_recipient_warning if config else "",
        "last_summary_count": config.last_summary_count if config else None,
        "save_warning": save_warning,
        "article_count": len(articles),
        "preview_articles": [
            {
                "id": article["id"],
                "name": article["name"],
                "internal_code": article["internal_code"],
                "current_stock": article["current_stock"],
                "available_stock": article["available_stock"],
                "minimum_stock": article["minimum_stock"],
            }
            for article in articles[:6]
        ],
    }


def resolve_digest_frequency(value):
    """Maneja resolve digest frequency."""
    raw_value = clean_casefold(value or MinimumStockDigestConfig.Frequency.DAILY)
    for option, _label in MinimumStockDigestConfig.Frequency.choices:
        if clean_casefold(option) == raw_value:
            return option
    raise InventoryApiError("La frecuencia del resumen periodico no es valida.")


def resolve_digest_recipients(config):
    """Maneja resolve digest recipients."""
    recipient_emails = []
    discarded = []
    seen = set()

    recipient_users = list(
        config.recipients.select_related("profile__sector_default").order_by(
            "first_name",
            "last_name",
            "username",
        )
    )
    for user in recipient_users:
        profile = get_profile(user)
        email = clean_string(user.email).lower()
        label = user.get_full_name() or user.username
        if profile.status != UserProfile.Status.ACTIVE:
            discarded.append(f"{label}: usuario inactivo")
            continue
        if not email:
            discarded.append(f"{label}: sin email")
            continue
        try:
            validate_email(email)
        except ValidationError:
            discarded.append(f"{label}: email invalido")
            continue
        if email in seen:
            continue
        seen.add(email)
        recipient_emails.append(email)

    valid_extra_emails, invalid_extra_emails = split_email_list(config.additional_emails)
    for invalid_email in invalid_extra_emails:
        discarded.append(f"{invalid_email}: email invalido")
    for email in valid_extra_emails:
        if email in seen:
            continue
        seen.add(email)
        recipient_emails.append(email)

    warning_message = ""
    if discarded:
        warning_message = "Se ignoraron destinatarios: " + ", ".join(discarded)

    return {
        "emails": recipient_emails,
        "discarded": discarded,
        "warning_message": warning_message,
    }


def build_minimum_stock_digest_message(config, low_stock_articles):
    """Construye minimum stock digest message."""
    subject = f"[Inventario] Resumen de stock minimo ({len(low_stock_articles)} articulos)"
    lines = [
        "Resumen automatico de articulos en o por debajo del stock minimo.",
        "",
        f"Total afectados: {len(low_stock_articles)}",
        "",
    ]

    for article in low_stock_articles:
        lines.extend(
            [
                f"- {article['name']} ({article['internal_code']})",
                f"  Stock actual: {article['current_stock']}",
                f"  Stock minimo: {article['minimum_stock']}",
                f"  Sector responsable: {article['sector_responsible'] or 'Sin sector'}",
                f"  Ubicacion principal: {article['primary_location'] or 'Sin ubicacion'}",
                "",
            ]
        )

    if clean_string(config.notes):
        lines.extend(
            [
                "Notas:",
                config.notes.strip(),
            ]
        )

    return subject, "\n".join(lines).strip()


def send_minimum_stock_digest_email(config, recipient_emails, low_stock_articles):
    """Env?a minimum stock digest email."""
    if not getattr(settings, "INVENTORY_ALARM_EMAILS_ENABLED", True):
        return False, "El envio por mail esta desactivado en la configuracion."

    subject, body = build_minimum_stock_digest_message(config, low_stock_articles)
    try:
        sent = send_mail(
            subject=subject,
            message=body,
            from_email=getattr(settings, "DEFAULT_FROM_EMAIL", "inventario@erp.local"),
            recipient_list=recipient_emails,
            fail_silently=False,
        )
    except Exception as exc:  # noqa: BLE001
        return False, str(exc)

    if not sent:
        return False, "No se pudo confirmar el envio del mail."
    return True, ""


def serialize_movement(movement):
    """Maneja serialize movement."""
    return {
        "id": movement.id,
        "timestamp": serialize_datetime(movement.timestamp),
        "movement_type": movement.movement_type,
        "movement_type_label": movement.get_movement_type_display(),
        "article": movement.article.name,
        "article_id": movement.article_id,
        "quantity": serialize_decimal(movement.quantity),
        "recorded_by": movement.recorded_by.username,
        "tracked_unit": movement.tracked_unit.internal_tag if movement.tracked_unit else None,
        "source_location": movement.source_location.name if movement.source_location else None,
        "target_location": movement.target_location.name if movement.target_location else None,
        "person": movement.person.full_name if movement.person else None,
        "sector": movement.sector.name if movement.sector else None,
        "reason_text": movement.reason_text,
        "document_ref": movement.document_ref,
        "authorized_by": movement.authorized_by.username if movement.authorized_by else None,
        "notes": movement.notes,
    }


def serialize_checkout(checkout):
    """Maneja serialize checkout."""
    return {
        "id": checkout.id,
        "tracked_unit": checkout.tracked_unit.internal_tag,
        "article": checkout.tracked_unit.article.name,
        "checkout_kind": checkout.checkout_kind,
        "checkout_kind_label": checkout.get_checkout_kind_display(),
        "status": checkout.status,
        "status_label": checkout.get_status_display(),
        "receiver_person": checkout.receiver_person.full_name if checkout.receiver_person else None,
        "receiver_sector": checkout.receiver_sector.name if checkout.receiver_sector else None,
        "checked_out_at": serialize_datetime(checkout.checked_out_at),
        "expected_return_at": serialize_datetime(checkout.expected_return_at),
        "returned_at": serialize_datetime(checkout.returned_at),
        "recorded_by": checkout.recorded_by.username,
        "notes": checkout.notes,
    }


def serialize_count_session(session):
    """Maneja serialize count session."""
    return {
        "id": session.id,
        "count_type": session.count_type,
        "count_type_label": session.get_count_type_display(),
        "scope": session.scope,
        "scheduled_for": serialize_datetime(session.scheduled_for),
        "status": session.status,
        "status_label": session.get_status_display(),
        "created_by": session.created_by.username,
        "notes": session.notes,
        "line_count": session.lines.count(),
    }


def serialize_discrepancy(discrepancy):
    """Maneja serialize discrepancy."""
    return {
        "id": discrepancy.id,
        "article": discrepancy.article.name,
        "article_id": discrepancy.article_id,
        "location": discrepancy.location.name if discrepancy.location else None,
        "difference_qty": serialize_decimal(discrepancy.difference_qty),
        "difference_type": discrepancy.difference_type,
        "difference_type_label": discrepancy.get_difference_type_display(),
        "status": discrepancy.status,
        "status_label": discrepancy.get_status_display(),
        "detected_by": discrepancy.detected_by.username,
        "detected_at": serialize_datetime(discrepancy.detected_at),
        "possible_cause": discrepancy.possible_cause,
        "action_taken": discrepancy.action_taken,
        "comment": discrepancy.comment,
        "movement_id": discrepancy.movement_id,
    }


def serialize_tracked_unit(unit):
    """Maneja serialize tracked unit."""
    return {
        "id": unit.id,
        "internal_tag": unit.internal_tag,
        "article": unit.article.name,
        "article_id": unit.article_id,
        "status": unit.status,
        "status_label": unit.get_status_display(),
        "current_location": unit.current_location.name if unit.current_location else None,
        "current_sector": unit.current_sector.name if unit.current_sector else None,
        "current_holder_person": unit.current_holder_person.full_name if unit.current_holder_person else None,
        "serial_number": unit.serial_number,
        "brand": unit.brand,
        "model": unit.model,
        "size": unit.size,
        "last_revision_at": serialize_date(unit.last_revision_at),
    }


def serialize_balance(balance):
    """Maneja serialize balance."""
    return {
        "id": balance.id,
        "article": balance.article.name,
        "article_id": balance.article_id,
        "location": balance.location.name,
        "location_id": balance.location_id,
        "batch": balance.batch.lot_code if balance.batch else None,
        "on_hand": serialize_decimal(balance.on_hand),
        "reserved": serialize_decimal(balance.reserved),
        "available": serialize_decimal(balance.available),
    }


def serialize_batch(batch):
    """Maneja serialize batch."""
    return {
        "id": batch.id,
        "article": batch.article.name,
        "article_id": batch.article_id,
        "lot_code": batch.lot_code,
        "expiry_date": serialize_date(batch.expiry_date),
        "received_at": serialize_date(batch.received_at),
        "supplier": batch.supplier.name if batch.supplier else None,
        "document_ref": batch.document_ref,
        "quality_status": batch.quality_status,
        "quality_status_label": batch.get_quality_status_display() if batch.quality_status else None,
        "notes": batch.notes,
    }


def serialize_person(person):
    """Maneja serialize person."""
    return {
        "id": person.id,
        "full_name": person.full_name,
        "employee_code": person.employee_code,
        "sector": person.sector.name if person.sector else None,
        "status": person.status,
    }


def serialize_catalogs():
    """Maneja serialize catalogs."""
    return {
        "article_types": [{"value": value, "label": label} for value, label in Article.ArticleType.choices],
        "tracking_modes": [{"value": value, "label": label} for value, label in Article.TrackingMode.choices],
        "movement_types": [{"value": value, "label": label} for value, label in StockMovement.MovementType.choices],
        "unit_statuses": [{"value": value, "label": label} for value, label in TrackedUnit.UnitStatus.choices],
        "checkout_statuses": [{"value": value, "label": label} for value, label in AssetCheckout.CheckoutStatus.choices],
        "count_statuses": [{"value": value, "label": label} for value, label in PhysicalCountSession.CountStatus.choices],
        "roles": [{"value": value, "label": label} for value, label in UserProfile.Role.choices],
        "categories": [{"id": category.id, "name": category.name, "parent_id": category.parent_id} for category in ArticleCategory.objects.order_by("name")],
        "units": [{"id": unit.id, "code": unit.code, "name": unit.name} for unit in UnitOfMeasure.objects.order_by("name")],
        "sectors": [{"id": sector.id, "name": sector.name, "code": sector.code} for sector in Sector.objects.order_by("name")],
        "locations": [{"id": location.id, "name": location.name, "code": location.code, "sector_id": location.sector_id} for location in Location.objects.order_by("name")],
        "people": [serialize_person(person) for person in Person.objects.order_by("full_name")],
        "suppliers": [{"id": supplier.id, "name": supplier.name} for supplier in Supplier.objects.order_by("name")],
    }


def validate_alarm_rule_recipients(user_ids, *, require_email=True, require_telegram=False):
    """Valida alarm rule recipients."""
    if not user_ids:
        return []

    normalized_ids = []
    for value in user_ids:
        normalized_id = parse_optional_int(value, "recipient_user_ids")
        if normalized_id is None:
            continue
        normalized_ids.append(normalized_id)

    normalized_ids = list(dict.fromkeys(normalized_ids))
    users = list(
        get_user_model()
        .objects.select_related("profile__sector_default")
        .filter(pk__in=normalized_ids)
        .order_by("first_name", "last_name", "username")
    )
    if len(users) != len(normalized_ids):
        raise InventoryApiError("Some selected recipients do not exist")

    invalid_recipients = []
    for recipient in users:
        profile = get_profile(recipient)
        if profile.status != UserProfile.Status.ACTIVE:
            invalid_recipients.append(f"{recipient.get_full_name() or recipient.username} esta inactivo")
        elif require_email and not clean_string(recipient.email):
            invalid_recipients.append(f"{recipient.get_full_name() or recipient.username} no tiene email")
        elif require_telegram and not clean_string(profile.telegram_chat_id):
            invalid_recipients.append(
                f"{recipient.get_full_name() or recipient.username} no tiene Telegram configurado"
            )

    if invalid_recipients:
        raise InventoryApiError(", ".join(invalid_recipients))

    return users


def send_safety_stock_alert_email(alert):
    """Env?a safety stock alert email."""
    recipients = safety_alert_email_addresses(alert)
    if not recipients:
        alert.last_email_error = "No hay destinatarios con email valido para esta regla."
        save_validated(alert)
        return False

    if not getattr(settings, "INVENTORY_ALARM_EMAILS_ENABLED", True):
        alert.last_email_error = "El envio por mail esta desactivado en la configuracion."
        save_validated(alert)
        return False

    article = alert.article
    current_stock = alert.last_stock_value if alert.last_stock_value is not None else current_stock_for_article(article)
    location_name = article.primary_location.name if article.primary_location else "Sin ubicacion principal"
    subject = f"[Alarma] Stock minimo alcanzado: {article.name} ({article.internal_code})"
    body = "\n".join(
        [
            "Se activo una alarma automatica por stock minimo.",
            "",
            f"Articulo: {article.name}",
            f"Codigo interno: {article.internal_code}",
            f"Tipo: {article.get_article_type_display()}",
            f"Stock actual: {serialize_decimal(current_stock)}",
            f"Stock minimo: {serialize_decimal(article.minimum_stock)}",
            f"Sector responsable: {article.sector_responsible.name}",
            f"Ubicacion principal: {location_name}",
            "",
            "Revisa el modulo correspondiente para tomar accion.",
            alert.notes.strip() if alert.notes.strip() else "",
        ]
    ).strip()

    try:
        sent = send_mail(
            subject=subject,
            message=body,
            from_email=getattr(settings, "DEFAULT_FROM_EMAIL", "inventario@erp.local"),
            recipient_list=recipients,
            fail_silently=False,
        )
    except Exception as exc:  # noqa: BLE001
        alert.last_email_error = str(exc)
        save_validated(alert)
        return False

    if sent:
        alert.last_notified_at = timezone.now()
        alert.last_email_error = ""
        save_validated(alert)
        return True

    alert.last_email_error = "No se pudo confirmar el envio del mail."
    save_validated(alert)
    return False


def send_safety_stock_alert_telegram(alert):
    """Env?a safety stock alert telegram."""
    if not getattr(settings, "INVENTORY_ALARM_TELEGRAM_ENABLED", True):
        alert.last_telegram_error = "El envio por Telegram esta desactivado en la configuracion."
        save_validated(alert)
        return False

    bot_token = clean_string(getattr(settings, "TELEGRAM_BOT_TOKEN", ""))
    if not bot_token:
        alert.last_telegram_error = "Telegram no esta configurado (TELEGRAM_BOT_TOKEN)."
        save_validated(alert)
        return False

    recipients = list(alert.recipients.select_related("profile").all())
    chat_ids = []
    seen = set()
    for recipient in recipients:
        chat_id = clean_string(getattr(get_profile(recipient), "telegram_chat_id", ""))
        if not chat_id or chat_id in seen:
            continue
        seen.add(chat_id)
        chat_ids.append(chat_id)

    if not chat_ids:
        alert.last_telegram_error = "No hay destinatarios con Telegram configurado para esta regla."
        save_validated(alert)
        return False

    article = alert.article
    current_stock = (
        alert.last_stock_value
        if alert.last_stock_value is not None
        else current_stock_for_article(article)
    )
    location_name = article.primary_location.name if article.primary_location else "Sin ubicacion principal"
    message = "\n".join(
        [
            "Alarma por stock minimo alcanzado.",
            f"Articulo: {article.name} ({article.internal_code})",
            f"Stock actual: {serialize_decimal(current_stock)}",
            f"Stock minimo: {serialize_decimal(article.minimum_stock)}",
            f"Sector: {article.sector_responsible.name}",
            f"Ubicacion: {location_name}",
        ]
        + ([alert.notes.strip()] if alert.notes.strip() else [])
    ).strip()

    errors = []
    sent_any = False
    url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
    timeout = getattr(settings, "TELEGRAM_SEND_TIMEOUT", 8)
    for chat_id in chat_ids:
        payload = json.dumps(
            {
                "chat_id": chat_id,
                "text": message,
                "disable_web_page_preview": True,
            }
        ).encode("utf-8")
        request = urllib.request.Request(
            url,
            data=payload,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        try:
            with urllib.request.urlopen(request, timeout=timeout) as response:  # noqa: S310
                status = int(getattr(response, "status", 200) or 200)
                if 200 <= status < 300:
                    sent_any = True
                else:
                    errors.append(f"{chat_id}: status {status}")
        except urllib.error.HTTPError as exc:
            errors.append(f"{chat_id}: {exc.code}")
        except Exception as exc:  # noqa: BLE001
            errors.append(f"{chat_id}: {exc}")

    if sent_any and not errors:
        alert.last_notified_at = timezone.now()
        alert.last_telegram_error = ""
        save_validated(alert)
        return True

    if errors:
        alert.last_telegram_error = "; ".join(errors)[:900]
    else:
        alert.last_telegram_error = "No se pudo confirmar el envio por Telegram."
    if sent_any:
        alert.last_notified_at = timezone.now()
    save_validated(alert)
    return sent_any


def minimum_stock_alarm_email_addresses(config):
    """Maneja minimum stock alarm email addresses."""
    recipients = list(config.recipients.select_related("profile").all())
    recipient_emails = []
    seen = set()
    for recipient in recipients:
        email = clean_string(recipient.email).lower()
        if not email or email in seen:
            continue
        seen.add(email)
        recipient_emails.append(email)

    valid_extra_emails, _invalid_extra_emails = split_email_list(config.additional_emails)
    for email in valid_extra_emails:
        if email in seen:
            continue
        seen.add(email)
        recipient_emails.append(email)

    return recipient_emails


def send_minimum_stock_alarm_email(config, article, current_stock=None):
    """Env?a minimum stock alarm email."""
    recipients = minimum_stock_alarm_email_addresses(config)
    if not recipients:
        config.last_email_error = "No hay destinatarios con email valido para esta regla."
        save_validated(config)
        return False

    if not getattr(settings, "INVENTORY_ALARM_EMAILS_ENABLED", True):
        config.last_email_error = "El envio por mail esta desactivado en la configuracion."
        save_validated(config)
        return False

    current_stock_value = current_stock if current_stock is not None else current_stock_for_article(article)
    location_name = article.primary_location.name if article.primary_location else "Sin ubicacion principal"
    subject = f"[Compras] Stock minimo alcanzado: {article.name} ({article.internal_code})"
    body = "\n".join(
        [
            "Se activo una alarma automatica por stock minimo.",
            "",
            f"Articulo: {article.name}",
            f"Codigo interno: {article.internal_code}",
            f"Tipo: {article.get_article_type_display()}",
            f"Stock actual: {serialize_decimal(current_stock_value)}",
            f"Stock minimo: {serialize_decimal(article.minimum_stock)}",
            f"Sector responsable: {article.sector_responsible.name}",
            f"Ubicacion principal: {location_name}",
            "",
            "Revisa el modulo de Compras para tomar accion.",
            config.notes.strip() if config.notes.strip() else "",
        ]
    ).strip()

    try:
        sent = send_mail(
            subject=subject,
            message=body,
            from_email=getattr(settings, "DEFAULT_FROM_EMAIL", "inventario@erp.local"),
            recipient_list=recipients,
            fail_silently=False,
        )
    except Exception as exc:  # noqa: BLE001
        config.last_email_error = str(exc)
        save_validated(config)
        return False

    if sent:
        config.last_notified_at = timezone.now()
        config.last_email_error = ""
        save_validated(config)
        return True

    config.last_email_error = "No se pudo confirmar el envio del mail."
    save_validated(config)
    return False


def send_minimum_stock_alarm_telegram(config, article, current_stock=None):
    """Env?a minimum stock alarm telegram."""
    if not getattr(settings, "INVENTORY_ALARM_TELEGRAM_ENABLED", True):
        config.last_telegram_error = "El envio por Telegram esta desactivado en la configuracion."
        save_validated(config)
        return False

    bot_token = clean_string(getattr(settings, "TELEGRAM_BOT_TOKEN", ""))
    if not bot_token:
        config.last_telegram_error = "Telegram no esta configurado (TELEGRAM_BOT_TOKEN)."
        save_validated(config)
        return False

    recipients = list(config.recipients.select_related("profile").all())
    chat_ids = []
    seen = set()
    for recipient in recipients:
        chat_id = clean_string(getattr(get_profile(recipient), "telegram_chat_id", ""))
        if not chat_id or chat_id in seen:
            continue
        seen.add(chat_id)
        chat_ids.append(chat_id)

    if not chat_ids:
        config.last_telegram_error = "No hay destinatarios con Telegram configurado para esta regla."
        save_validated(config)
        return False

    current_stock_value = current_stock if current_stock is not None else current_stock_for_article(article)
    location_name = article.primary_location.name if article.primary_location else "Sin ubicacion principal"
    message = "\n".join(
        [
            "Alarma por stock minimo alcanzado.",
            f"Articulo: {article.name} ({article.internal_code})",
            f"Stock actual: {serialize_decimal(current_stock_value)}",
            f"Stock minimo: {serialize_decimal(article.minimum_stock)}",
            f"Sector: {article.sector_responsible.name}",
            f"Ubicacion: {location_name}",
        ]
        + ([config.notes.strip()] if config.notes.strip() else [])
    ).strip()

    errors = []
    sent_any = False
    url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
    timeout = getattr(settings, "TELEGRAM_SEND_TIMEOUT", 8)
    for chat_id in chat_ids:
        payload = json.dumps(
            {
                "chat_id": chat_id,
                "text": message,
                "disable_web_page_preview": True,
            }
        ).encode("utf-8")
        request = urllib.request.Request(
            url,
            data=payload,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        try:
            with urllib.request.urlopen(request, timeout=timeout) as response:  # noqa: S310
                status = int(getattr(response, "status", 200) or 200)
                if 200 <= status < 300:
                    sent_any = True
                else:
                    errors.append(f"{chat_id}: status {status}")
        except urllib.error.HTTPError as exc:
            errors.append(f"{chat_id}: {exc.code}")
        except Exception as exc:  # noqa: BLE001
            errors.append(f"{chat_id}: {exc}")

    if sent_any and not errors:
        config.last_notified_at = timezone.now()
        config.last_telegram_error = ""
        save_validated(config)
        return True

    if errors:
        config.last_telegram_error = "; ".join(errors)[:900]
    else:
        config.last_telegram_error = "No se pudo confirmar el envio por Telegram."
    if sent_any:
        config.last_notified_at = timezone.now()
    save_validated(config)
    return sent_any


def evaluate_purchasing_minimum_stock_alarm(article):
    """Maneja evaluate purchasing minimum stock alarm."""
    try:
        with transaction.atomic():
            config = (
                MinimumStockAlarmConfig.objects.select_for_update()
                .prefetch_related("recipients__profile")
                .filter(key="purchasing_default", is_enabled=True)
                .first()
            )
            if not config:
                return None
            if not (config.notify_email or config.notify_telegram):
                return None
            if SafetyStockAlertRule.objects.filter(article=article, is_enabled=True).exists():
                # Si hay regla individual, la global no aplica (override)
                return None

            if article.minimum_stock is None:
                return None

            current_stock = current_stock_for_article(article)
            is_under_minimum = current_stock <= article.minimum_stock

            state, created = MinimumStockAlarmState.objects.select_for_update().get_or_create(
                article=article,
                defaults={
                    "created_by": None,
                    "updated_by": None,
                },
            )
            old_status = state.status
            new_status = (
                MinimumStockAlarmState.AlarmStatus.TRIGGERED
                if is_under_minimum
                else MinimumStockAlarmState.AlarmStatus.MONITORING
            )
            should_send = (
                old_status == MinimumStockAlarmState.AlarmStatus.MONITORING
                and new_status == MinimumStockAlarmState.AlarmStatus.TRIGGERED
            )

            state.status = new_status
            state.last_stock_value = current_stock
            if should_send:
                state.triggered_at = timezone.now()
                state.resolved_at = None
                state.last_notified_at = timezone.now()
                state.last_email_error = ""
                state.last_telegram_error = ""
            elif (
                old_status == MinimumStockAlarmState.AlarmStatus.TRIGGERED
                and new_status == MinimumStockAlarmState.AlarmStatus.MONITORING
            ):
                state.resolved_at = timezone.now()
                state.last_email_error = ""
                state.last_telegram_error = ""

            update_audit(state, None, is_new=created)
            save_validated(state)

            if should_send:
                try:
                    if config.notify_email:
                        send_minimum_stock_alarm_email(config, article, current_stock=current_stock)
                    if config.notify_telegram:
                        send_minimum_stock_alarm_telegram(config, article, current_stock=current_stock)
                except Exception as exc:  # noqa: BLE001
                    logging.getLogger("inventory.automation.alert").error(
                        "purchasing_alarm_send_failed",
                        extra={"article_id": article.id, "error": str(exc)},
                        exc_info=True,
                    )

            return state
    except MinimumStockAlarmConfig.DoesNotExist:
        return None


def evaluate_safety_stock_alert(article):
    """
    Evalúa el alert de seguridad de stock de un artículo de forma idempotente.
    
    Garantías:
    - La evaluación es transaccional (SELECT FOR UPDATE)
    - El email se envía SOLO en la transición MONITORING -> TRIGGERED
    - Ante concurrencia, el último cambio gana y el estado es consistente
    - Errores de email no rompen la transacción de cambio de estado
    """
    try:
        with transaction.atomic():
            # SELECT FOR UPDATE garantiza exclusividad en la transacción
            alert = (
                SafetyStockAlertRule.objects.select_for_update()
                .select_related(
                    "article__sector_responsible",
                    "article__primary_location",
                )
                .prefetch_related("recipients__profile__sector_default")
                .filter(article=article, is_enabled=True)
                .first()
            )
            
            if not alert:
                # No hay regla habilitada
                return None
            
            current_stock = current_stock_for_article(article)
            is_under_minimum = article.minimum_stock is not None and current_stock <= article.minimum_stock
            
            # Lógica de transición de estado
            old_status = alert.status
            new_status = SafetyStockAlertRule.AlertStatus.TRIGGERED if is_under_minimum else SafetyStockAlertRule.AlertStatus.MONITORING
            
            # *** CLAVE: Email solo en transición MONITORING -> TRIGGERED ***
            should_send_notifications = (
                old_status == SafetyStockAlertRule.AlertStatus.MONITORING and 
                new_status == SafetyStockAlertRule.AlertStatus.TRIGGERED
            )
            
            # Actualiza campos de estado
            alert.status = new_status
            alert.last_stock_value = current_stock
            
            if should_send_notifications:
                alert.triggered_at = timezone.now()
                alert.resolved_at = None
                alert.last_notified_at = timezone.now()
                alert.last_email_error = ""
                alert.last_telegram_error = ""
            elif old_status == SafetyStockAlertRule.AlertStatus.TRIGGERED and new_status == SafetyStockAlertRule.AlertStatus.MONITORING:
                # Transición inversa: volvió arriba del mínimo
                alert.resolved_at = timezone.now()
                alert.last_email_error = ""
                alert.last_telegram_error = ""
            
            # Guardaguarda el estado ANTES de intentar enviar email
            save_validated(alert)
            
            # Despues de guardar estado, intenta enviar si aplica
            if should_send_notifications:
                try:
                    if alert.notify_email:
                        send_safety_stock_alert_email(alert)
                    if alert.notify_telegram:
                        send_safety_stock_alert_telegram(alert)
                except Exception as exc:  # noqa: BLE001
                    # Error de email: registralo pero no rompas la transacción
                    logging.getLogger("inventory.automation.alert").error(
                        "alert_email_send_failed",
                        extra={
                            "article_id": article.id,
                            "rule_id": alert.id,
                            "error": str(exc),
                        },
                        exc_info=True
                    )
                    # NOTA: last_email_error debe ser actualizado en send_safety_stock_alert_email
                    # Si falló en send_mail, ya está poblado allá
            
            return alert
    
    except SafetyStockAlertRule.DoesNotExist:
        # No hay regla o no está habilitada; no hacer nada
        logging.getLogger("inventory.automation.alert").debug(
            f"alert_rule_not_found_or_disabled article_id={article.id}"
        )
        return None
    except Exception as exc:  # noqa: BLE001
        logging.getLogger("inventory.automation.alert").error(
            "alert_evaluation_error",
            extra={
                "article_id": article.id,
                "error": str(exc),
            },
            exc_info=True
        )
        # No relanzar: errores en evaluación no deben romper cambios de stock
        return None


def list_safety_stock_alerts(user):
    """Lista safety stock alerts."""
    require_role(user, ALARM_ROLES)
    quantity_map, _available_quantity_map, unit_total_map, _unit_available_map = current_stock_maps()
    alerts = list(
        SafetyStockAlertRule.objects.select_related(
            "article__primary_location",
            "article__sector_responsible",
        )
        .prefetch_related("recipients__profile__sector_default")
        .order_by("article__name")
    )
    alerts.sort(
        key=lambda alert: (
            0 if alert.status == SafetyStockAlertRule.AlertStatus.TRIGGERED else 1,
            alert.article.name.casefold(),
        )
    )
    return [serialize_safety_alert_rule(alert, quantity_map, unit_total_map) for alert in alerts]


def save_safety_stock_alert_rule(user, payload):
    """Guarda safety stock alert rule."""
    require_role(user, ALARM_ROLES)
    article = resolve_instance(Article, payload.get("article_id"), "article")
    if article.minimum_stock is None:
        raise InventoryApiError("El articulo debe tener stock minimo para activar una alarma automatica.")

    recipient_ids = payload.get("recipient_user_ids") or []
    if isinstance(recipient_ids, str):
        recipient_ids = [item for item in re.split(r"[,\s]+", recipient_ids) if item]
    additional_emails = payload.get("additional_emails") or ""
    is_enabled = parse_boolean(payload.get("is_enabled"))
    notify_email = parse_boolean(payload.get("notify_email")) if "notify_email" in payload else True
    notify_telegram = parse_boolean(payload.get("notify_telegram")) if "notify_telegram" in payload else False

    if is_enabled and not (notify_email or notify_telegram):
        raise InventoryApiError("Selecciona al menos un canal de notificacion (email o Telegram).")

    recipients = validate_alarm_rule_recipients(
        recipient_ids,
        require_email=notify_email,
        require_telegram=notify_telegram,
    )
    validated_extra_emails = parse_email_list(additional_emails) if notify_email else []

    if is_enabled and notify_email and not (recipients or validated_extra_emails):
        raise InventoryApiError("Selecciona al menos un destinatario o agrega un email adicional.")
    if is_enabled and notify_telegram and not recipients:
        raise InventoryApiError("Selecciona al menos un destinatario con Telegram configurado.")

    with transaction.atomic():
        alert, created = SafetyStockAlertRule.objects.get_or_create(
            article=article,
            defaults={
                "created_by": user,
                "updated_by": user,
            },
        )
        alert.is_enabled = is_enabled
        alert.notify_email = notify_email
        alert.notify_telegram = notify_telegram
        alert.additional_emails = "\n".join(validated_extra_emails) if notify_email else ""
        alert.notes = payload.get("notes") or ""
        update_audit(alert, user, is_new=created)
        save_validated(alert)
        alert.recipients.set(recipients)
        alert = evaluate_safety_stock_alert(article) or alert

    refreshed_alert = (
        SafetyStockAlertRule.objects.select_related(
            "article__primary_location",
            "article__sector_responsible",
        )
        .prefetch_related("recipients__profile__sector_default")
        .get(pk=alert.id)
    )
    return serialize_safety_alert_rule(refreshed_alert)


def get_minimum_stock_digest_config(user, serialized_articles=None):
    """Devuelve minimum stock digest config."""
    require_role(user, ALARM_ROLES)
    config = (
        MinimumStockDigestConfig.objects.prefetch_related("recipients__profile__sector_default")
        .filter(key="default")
        .first()
    )
    return serialize_minimum_stock_digest_config(config, serialized_articles=serialized_articles)


def get_full_stock_report_config(user, serialized_articles=None):
    """Devuelve full stock report config."""
    require_role(user, ALARM_ROLES)
    config = (
        FullStockReportConfig.objects.prefetch_related("recipients__profile__sector_default")
        .filter(key="default")
        .first()
    )
    return serialize_full_stock_report_config(config, serialized_articles=serialized_articles)


def save_minimum_stock_digest_config(user, payload):
    """Guarda minimum stock digest config."""
    require_role(user, ALARM_ROLES)

    recipient_ids = payload.get("recipient_user_ids") or []
    if isinstance(recipient_ids, str):
        recipient_ids = [item for item in re.split(r"[,\s]+", recipient_ids) if item]

    recipients = validate_alarm_rule_recipients(recipient_ids, require_email=True)
    additional_emails = payload.get("additional_emails") or ""
    validated_extra_emails, invalid_extra_emails = split_email_list(additional_emails)
    is_enabled = parse_boolean(payload.get("is_enabled"))
    frequency = resolve_digest_frequency(payload.get("frequency"))
    run_at = parse_time_or_error(payload.get("run_at"), "run_at", default=parse_time("08:00"))
    run_weekday = parse_weekday_or_error(payload.get("run_weekday"), default=0)

    if is_enabled and not (recipients or validated_extra_emails):
        raise InventoryApiError("Selecciona al menos un destinatario o agrega un email adicional.")

    save_warning = ""
    if invalid_extra_emails:
        save_warning = (
            "Se ignoraron emails invalidos al guardar: " + ", ".join(invalid_extra_emails)
        )

    with transaction.atomic():
        config, created = MinimumStockDigestConfig.objects.get_or_create(
            key="default",
            defaults={
                "created_by": user,
                "updated_by": user,
            },
        )
        config.is_enabled = is_enabled
        config.frequency = frequency
        config.run_at = run_at
        config.run_weekday = run_weekday
        config.additional_emails = "\n".join(validated_extra_emails)
        config.notes = payload.get("notes") or ""
        update_audit(config, user, is_new=created)
        save_validated(config)
        config.recipients.set(recipients)

    refreshed_config = (
        MinimumStockDigestConfig.objects.prefetch_related("recipients__profile__sector_default")
        .filter(pk=config.pk)
        .first()
    )
    return serialize_minimum_stock_digest_config(refreshed_config, save_warning=save_warning)


def save_full_stock_report_config(user, payload):
    """Guarda full stock report config."""
    require_role(user, ALARM_ROLES)

    recipient_ids = payload.get("recipient_user_ids") or []
    if isinstance(recipient_ids, str):
        recipient_ids = [item for item in re.split(r"[,\s]+", recipient_ids) if item]

    recipients = validate_alarm_rule_recipients(recipient_ids, require_email=True)
    additional_emails = payload.get("additional_emails") or ""
    validated_extra_emails, invalid_extra_emails = split_email_list(additional_emails)
    is_enabled = parse_boolean(payload.get("is_enabled"))
    frequency = resolve_digest_frequency(payload.get("frequency"))
    run_at = parse_time_or_error(payload.get("run_at"), "run_at", default=parse_time("08:00"))
    run_weekday = parse_weekday_or_error(payload.get("run_weekday"), default=0)

    if is_enabled and not (recipients or validated_extra_emails):
        raise InventoryApiError("Selecciona al menos un destinatario o agrega un email adicional.")

    save_warning = ""
    if invalid_extra_emails:
        save_warning = (
            "Se ignoraron emails invalidos al guardar: " + ", ".join(invalid_extra_emails)
        )

    with transaction.atomic():
        config, created = FullStockReportConfig.objects.get_or_create(
            key="default",
            defaults={
                "created_by": user,
                "updated_by": user,
            },
        )
        config.is_enabled = is_enabled
        config.frequency = frequency
        config.run_at = run_at
        config.run_weekday = run_weekday
        config.additional_emails = "\n".join(validated_extra_emails)
        config.notes = payload.get("notes") or ""
        update_audit(config, user, is_new=created)
        save_validated(config)
        config.recipients.set(recipients)

    refreshed_config = (
        FullStockReportConfig.objects.prefetch_related("recipients__profile__sector_default")
        .filter(pk=config.pk)
        .first()
    )
    return serialize_full_stock_report_config(refreshed_config, save_warning=save_warning)


def dispatch_minimum_stock_digest(config_id, due_key):
    """Maneja dispatch minimum stock digest."""
    config = (
        MinimumStockDigestConfig.objects.prefetch_related("recipients__profile__sector_default")
        .filter(pk=config_id)
        .first()
    )
    if not config or not config.is_enabled:
        mark_minimum_stock_digest_result(
            config_id,
            due_key,
            MinimumStockDigestConfig.DeliveryStatus.SKIPPED,
            now=timezone.now(),
            summary_count=0,
            consume_period=True,
        )
        DIGEST_AUTOMATION_LOGGER.info(
            "digest_skipped_disabled",
            extra={"period_key": due_key},
        )
        return {
            "delivery_status": MinimumStockDigestConfig.DeliveryStatus.SKIPPED,
            "summary_count": 0,
            "recipient_warning": "",
            "email_error": "",
        }

    # Marca el periodo como inflight para reflejar estado incluso si se llama fuera del runner.
    started_at = timezone.now()
    MinimumStockDigestConfig.objects.filter(pk=config_id).update(
        inflight_period_key=due_key,
        inflight_started_at=started_at,
    )
    config.inflight_period_key = due_key
    config.inflight_started_at = started_at

    low_stock_articles = low_stock_articles_snapshot()
    if not low_stock_articles:
        mark_minimum_stock_digest_result(
            config_id,
            due_key,
            MinimumStockDigestConfig.DeliveryStatus.SKIPPED,
            now=timezone.now(),
            summary_count=0,
            consume_period=True,
            email_error="",
            recipient_warning="",
        )
        DIGEST_AUTOMATION_LOGGER.info(
            "digest_skipped_no_items",
            extra={"period_key": due_key},
        )
        return {
            "delivery_status": MinimumStockDigestConfig.DeliveryStatus.SKIPPED,
            "summary_count": 0,
            "recipient_warning": "",
            "email_error": "",
        }

    recipient_resolution = resolve_digest_recipients(config)
    recipient_emails = recipient_resolution["emails"]
    recipient_warning = recipient_resolution["warning_message"]
    if recipient_warning:
        DIGEST_AUTOMATION_LOGGER.warning(
            "digest_recipients_warning",
            extra={
                "period_key": due_key,
                "recipient_count": len(recipient_emails),
                "discarded_recipient_count": len(recipient_resolution["discarded"]),
            },
        )

    if not recipient_emails:
        recipient_warning = recipient_warning or "No hay destinatarios validos para el resumen."
        mark_minimum_stock_digest_result(
            config_id,
            due_key,
            MinimumStockDigestConfig.DeliveryStatus.WARNING,
            now=timezone.now(),
            summary_count=len(low_stock_articles),
            recipient_warning=recipient_warning,
            email_error="",
            consume_period=True,
        )
        return {
            "delivery_status": MinimumStockDigestConfig.DeliveryStatus.WARNING,
            "summary_count": len(low_stock_articles),
            "recipient_warning": recipient_warning,
            "email_error": "",
        }

    DIGEST_AUTOMATION_LOGGER.info(
        "digest_send_start",
        extra={
            "period_key": due_key,
            "recipient_count": len(recipient_emails),
            "discarded_recipient_count": len(recipient_resolution["discarded"]),
            "summary_count": len(low_stock_articles),
        },
    )
    sent, email_error = send_minimum_stock_digest_email(
        config,
        recipient_emails,
        low_stock_articles,
    )
    if not sent:
        mark_minimum_stock_digest_result(
            config_id,
            due_key,
            MinimumStockDigestConfig.DeliveryStatus.ERROR,
            now=timezone.now(),
            summary_count=len(low_stock_articles),
            recipient_warning=recipient_warning,
            email_error=email_error,
            consume_period=False,
        )
        DIGEST_AUTOMATION_LOGGER.error(
            "digest_send_error",
            extra={
                "period_key": due_key,
                "recipient_count": len(recipient_emails),
                "summary_count": len(low_stock_articles),
            },
        )
        return {
            "delivery_status": MinimumStockDigestConfig.DeliveryStatus.ERROR,
            "summary_count": len(low_stock_articles),
            "recipient_warning": recipient_warning,
            "email_error": email_error,
        }

    delivery_status = (
        MinimumStockDigestConfig.DeliveryStatus.WARNING
        if recipient_warning
        else MinimumStockDigestConfig.DeliveryStatus.SUCCESS
    )
    mark_minimum_stock_digest_result(
        config_id,
        due_key,
        delivery_status,
        now=timezone.now(),
        summary_count=len(low_stock_articles),
        recipient_warning=recipient_warning,
        email_error="",
        consume_period=True,
    )
    DIGEST_AUTOMATION_LOGGER.info(
        "digest_send_success",
        extra={
            "period_key": due_key,
            "recipient_count": len(recipient_emails),
            "discarded_recipient_count": len(recipient_resolution["discarded"]),
            "summary_count": len(low_stock_articles),
        },
    )
    return {
        "delivery_status": delivery_status,
        "summary_count": len(low_stock_articles),
        "recipient_warning": recipient_warning,
        "email_error": "",
    }


def build_full_stock_report_message(config, article_count, report_filename, due_key=""):
    """Construye full stock report message."""
    report_label = ""
    if due_key and ":" in str(due_key):
        report_label = due_key.split(":", 1)[1]

    subject_suffix = report_label or timezone.localdate().isoformat()
    subject = f"[Inventario] Reporte de stock completo ({subject_suffix})"
    lines = [
        "Reporte automatico del stock completo (archivo Excel adjunto).",
        "",
        f"Total articulos: {article_count}",
        f"Archivo: {report_filename}",
        "",
    ]

    if clean_string(config.notes):
        lines.extend(
            [
                "Notas:",
                config.notes.strip(),
            ]
        )

    return subject, "\n".join(lines).strip()


def send_full_stock_report_email(config, recipient_emails, report_filename, report_payload, article_count, due_key=""):
    """Env?a full stock report email."""
    if not getattr(settings, "INVENTORY_ALARM_EMAILS_ENABLED", True):
        return False, "El envio por mail esta desactivado en la configuracion."

    subject, body = build_full_stock_report_message(
        config,
        article_count,
        report_filename,
        due_key=due_key,
    )
    try:
        message = EmailMessage(
            subject=subject,
            body=body,
            from_email=getattr(settings, "DEFAULT_FROM_EMAIL", "inventario@erp.local"),
            to=recipient_emails,
        )
        message.attach(
            report_filename,
            report_payload,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        sent = message.send(fail_silently=False)
    except Exception as exc:  # noqa: BLE001
        return False, str(exc)

    if not sent:
        return False, "No se pudo confirmar el envio del mail."
    return True, ""


def dispatch_full_stock_report(config_id, due_key):
    """Maneja dispatch full stock report."""
    config = (
        FullStockReportConfig.objects.prefetch_related("recipients__profile__sector_default")
        .filter(pk=config_id)
        .first()
    )
    if not config or not config.is_enabled:
        mark_full_stock_report_result(
            config_id,
            due_key,
            FullStockReportConfig.DeliveryStatus.SKIPPED,
            now=timezone.now(),
            summary_count=0,
            consume_period=True,
        )
        FULL_STOCK_REPORT_AUTOMATION_LOGGER.info(
            "stock_report_skipped_disabled",
            extra={"period_key": due_key},
        )
        return {
            "delivery_status": FullStockReportConfig.DeliveryStatus.SKIPPED,
            "summary_count": 0,
            "recipient_warning": "",
            "email_error": "",
        }

    # Marca el periodo como inflight para reflejar estado incluso si se llama fuera del runner.
    started_at = timezone.now()
    FullStockReportConfig.objects.filter(pk=config_id).update(
        inflight_period_key=due_key,
        inflight_started_at=started_at,
    )
    config.inflight_period_key = due_key
    config.inflight_started_at = started_at

    articles = filter_articles_for_stock_view(list_articles())
    report_filename, report_payload = build_stock_export_excel_from_articles(articles)

    recipient_resolution = resolve_digest_recipients(config)
    recipient_emails = recipient_resolution["emails"]
    recipient_warning = recipient_resolution["warning_message"]
    if recipient_warning:
        FULL_STOCK_REPORT_AUTOMATION_LOGGER.warning(
            "stock_report_recipients_warning",
            extra={
                "period_key": due_key,
                "recipient_count": len(recipient_emails),
                "discarded_recipient_count": len(recipient_resolution["discarded"]),
            },
        )

    if not recipient_emails:
        recipient_warning = recipient_warning or "No hay destinatarios validos para el reporte."
        mark_full_stock_report_result(
            config_id,
            due_key,
            FullStockReportConfig.DeliveryStatus.WARNING,
            now=timezone.now(),
            summary_count=len(articles),
            recipient_warning=recipient_warning,
            email_error="",
            consume_period=True,
        )
        return {
            "delivery_status": FullStockReportConfig.DeliveryStatus.WARNING,
            "summary_count": len(articles),
            "recipient_warning": recipient_warning,
            "email_error": "",
        }

    FULL_STOCK_REPORT_AUTOMATION_LOGGER.info(
        "stock_report_send_start",
        extra={
            "period_key": due_key,
            "recipient_count": len(recipient_emails),
            "discarded_recipient_count": len(recipient_resolution["discarded"]),
            "summary_count": len(articles),
        },
    )
    sent, email_error = send_full_stock_report_email(
        config,
        recipient_emails,
        report_filename,
        report_payload,
        len(articles),
        due_key=due_key,
    )
    if not sent:
        mark_full_stock_report_result(
            config_id,
            due_key,
            FullStockReportConfig.DeliveryStatus.ERROR,
            now=timezone.now(),
            summary_count=len(articles),
            recipient_warning=recipient_warning,
            email_error=email_error,
            consume_period=False,
        )
        FULL_STOCK_REPORT_AUTOMATION_LOGGER.error(
            "stock_report_send_error",
            extra={
                "period_key": due_key,
                "recipient_count": len(recipient_emails),
                "summary_count": len(articles),
            },
        )
        return {
            "delivery_status": FullStockReportConfig.DeliveryStatus.ERROR,
            "summary_count": len(articles),
            "recipient_warning": recipient_warning,
            "email_error": email_error,
        }

    delivery_status = (
        FullStockReportConfig.DeliveryStatus.WARNING
        if recipient_warning
        else FullStockReportConfig.DeliveryStatus.SUCCESS
    )
    mark_full_stock_report_result(
        config_id,
        due_key,
        delivery_status,
        now=timezone.now(),
        summary_count=len(articles),
        recipient_warning=recipient_warning,
        email_error="",
        consume_period=True,
    )
    FULL_STOCK_REPORT_AUTOMATION_LOGGER.info(
        "stock_report_send_success",
        extra={
            "period_key": due_key,
            "recipient_count": len(recipient_emails),
            "discarded_recipient_count": len(recipient_resolution["discarded"]),
            "summary_count": len(articles),
        },
    )
    return {
        "delivery_status": delivery_status,
        "summary_count": len(articles),
        "recipient_warning": recipient_warning,
        "email_error": "",
    }


def build_dashboard(user=None):
    """Construye dashboard."""
    from .deposits import resolve_deposit_permissions
    from accounts.permissions import has_module_permission
    from accounts.services import ensure_permission_catalog

    quantity_map, available_quantity_map, unit_total_map, unit_available_map = current_stock_maps()
    articles = Article.objects.select_related("sector_responsible").all()
    low_stock_count = 0

    for article in articles:
        if article.minimum_stock is None:
            continue
        current_stock = article_current_stock(article, quantity_map, unit_total_map)
        if current_stock <= article.minimum_stock:
            low_stock_count += 1

    open_checkouts = AssetCheckout.objects.filter(status=AssetCheckout.CheckoutStatus.OPEN).count()
    open_counts = PhysicalCountSession.objects.exclude(
        status=PhysicalCountSession.CountStatus.CLOSED
    ).count()
    open_discrepancies = StockDiscrepancy.objects.filter(
        status=StockDiscrepancy.DiscrepancyStatus.OPEN
    ).count()
    if user:
        ensure_permission_catalog()
    deposit_permissions = resolve_deposit_permissions(user) if user else {"can_view_module": False}
    active_pallets = Pallet.objects.exclude(status=Pallet.PalletStatus.ARCHIVED).count()
    deposit_events_today = PalletEvent.objects.filter(created_at__date=timezone.localdate()).count()

    modules = []

    if user and has_module_permission(user, "inventory_overview", "view"):
        modules.append(
            {
                "slug": "inventario",
                "name": "Inventario",
                "description": "Stock, prestamos, conteos y diferencias",
                "color": "#38a6ff",
                "badge": str(low_stock_count or 0),
                "status": "active",
            }
        )

    if deposit_permissions["can_view_module"]:
        modules.append(
            {
                "slug": "depositos",
                "name": "Depósitos",
                "description": "Pallets, plano visual y escaneo QR",
                "color": "#ff8b3d",
                "badge": str(active_pallets or deposit_events_today or 0),
                "status": "active",
            }
        )

    modules.extend(
        [
            *(
                [
                    {
                        "slug": "tia",
                        "name": "TIA",
                        "description": "Integracion Siemens S7-300 y monitoreo industrial",
                        "color": "#14b8a6",
                        "badge": "0",
                        "status": "active",
                    }
                ]
                if user and has_module_permission(user, "tia", "view")
                else []
            ),
            *(
                [
                    {
                        "slug": "personal",
                        "name": "Personal",
                        "description": "Informes y espacio personal del usuario",
                        "color": "#8b5cf6",
                        "badge": "0",
                        "status": "active",
                    }
                ]
                if user and has_module_permission(user, "personal", "view")
                else []
            ),
            *(
                [
                    {
                        "slug": "administracion",
                        "name": "Administración",
                        "description": "Usuarios y permisos",
                        "color": "#35596f",
                        "badge": "0",
                        "status": "active",
                    }
                ]
                if user and has_module_permission(user, "admin_users", "view")
                else []
            ),
            {
                "slug": "ventas",
                "name": "Ventas",
                "description": "Pendiente de fase posterior",
                "color": "#ea4cb8",
                "badge": "3",
                "status": "planned",
            },
            *(
                [
                    {
                        "slug": "compras",
                        "name": "Compras",
                        "description": "Solicitudes de compra y abastecimiento",
                        "color": "#d0472f",
                        "badge": "0",
                        "status": "active",
                    }
                ]
                if user and has_module_permission(user, "purchasing", "view")
                else []
            ),
            {
                "slug": "clientes",
                "name": "Clientes",
                "description": "Fuera del alcance de esta version",
                "color": "#25c2a0",
                "badge": "20+",
                "status": "planned",
            },
        ]
    )

    return {
        "welcome": {
            "title": "ERP modular",
            "subtitle": "Inventario operativo para planta con foco en velocidad y control suficiente.",
        },
        "modules": modules,
        "kpis": [
            {"label": "Articulos activos", "value": articles.count()},
            {"label": "Bajo stock", "value": low_stock_count},
            {"label": "Prestamos abiertos", "value": open_checkouts},
            {"label": "Conteos y diferencias", "value": open_counts + open_discrepancies},
        ],
    }


def build_inventory_overview(user):
    """Construye inventory overview."""
    profile = get_profile(user)
    from accounts.permissions import has_module_permission
    from accounts.services import ensure_permission_catalog

    ensure_permission_catalog()

    can_manage_master = (
        user.is_superuser
        or has_module_permission(user, "stock_management", "create")
        or has_module_permission(user, "stock_management", "change")
        or has_module_permission(user, "stock_management", "delete")
    )
    can_record_movement = user.is_superuser or has_module_permission(user, "movements", "create")
    can_checkout = user.is_superuser or has_module_permission(user, "checkouts", "create")
    can_count = user.is_superuser or has_module_permission(user, "counts", "create")
    can_approve = user.is_superuser or has_module_permission(user, "movements", "approve")
    can_manage_alarms = (
        user.is_superuser
        or has_module_permission(user, "alarms", "create")
        or has_module_permission(user, "alarms", "change")
    )
    quantity_map, available_quantity_map, unit_total_map, unit_available_map = current_stock_maps()
    articles = Article.objects.select_related(
        "unit_of_measure",
        "sector_responsible",
        "primary_location",
        "category",
        "subcategory",
        "supplier",
    ).order_by("name")
    serialized_articles = [
        serialize_article(
            article,
            quantity_map,
            available_quantity_map,
            unit_total_map,
            unit_available_map,
        )
        for article in articles
    ]
    low_stock_articles = [article for article in serialized_articles if article["low_stock"]][:8]

    return {
        "header": {
            "title": "Inventario",
            "subtitle": "",
        },
        "stats": [
            {"label": "Articulos", "value": len(serialized_articles), "hint": "Maestro activo"},
            {"label": "Bajo stock", "value": len(low_stock_articles), "hint": "Controlados por minimo"},
            {
                "label": "Prestamos abiertos",
                "value": AssetCheckout.objects.filter(status=AssetCheckout.CheckoutStatus.OPEN).count(),
                "hint": "Herramientas y EPP asignado",
            },
            {
                "label": "Conteos / diferencias",
                "value": PhysicalCountSession.objects.exclude(
                    status=PhysicalCountSession.CountStatus.CLOSED
                ).count()
                + StockDiscrepancy.objects.filter(
                    status=StockDiscrepancy.DiscrepancyStatus.OPEN
                ).count(),
                "hint": "Pendientes de cierre",
            },
        ],
        "permissions": {
            "can_manage_master": can_manage_master,
            "can_record_movement": can_record_movement,
            "can_checkout": can_checkout,
            "can_count": can_count,
            "can_approve": can_approve,
            "can_manage_alarms": can_manage_alarms,
        },
        "catalogs": {
            **serialize_catalogs(),
            "users": active_message_contacts(user),
            "alarm_recipients": active_alarm_recipients(user),
        },
        "articles": serialized_articles,
        "low_stock": low_stock_articles,
        "balances": [
            serialize_balance(balance)
            for balance in InventoryBalance.objects.select_related("article", "location", "batch")
            .order_by("article__name", "location__name")[:20]
        ],
        "batches": [
            serialize_batch(batch)
            for batch in InventoryBatch.objects.select_related("article", "supplier")
            .order_by("article__name", "lot_code")[:20]
        ],
        "tracked_units": [
            serialize_tracked_unit(unit)
            for unit in TrackedUnit.objects.select_related(
                "article",
                "current_location",
                "current_sector",
                "current_holder_person",
            ).order_by("internal_tag")[:20]
        ],
        "movements": [
            serialize_movement(movement)
            for movement in StockMovement.objects.select_related(
                "article",
                "recorded_by",
                "tracked_unit",
                "source_location",
                "target_location",
                "person",
                "sector",
                "authorized_by",
            ).order_by("-timestamp", "-id")[:20]
        ],
        "checkouts": [
            serialize_checkout(checkout)
            for checkout in AssetCheckout.objects.select_related(
                "tracked_unit__article",
                "receiver_person",
                "receiver_sector",
                "recorded_by",
            ).order_by("-checked_out_at")[:20]
        ],
        "count_sessions": [
            serialize_count_session(session)
            for session in PhysicalCountSession.objects.prefetch_related("lines")
            .order_by("-scheduled_for")[:12]
        ],
        "discrepancies": [
            serialize_discrepancy(discrepancy)
            for discrepancy in StockDiscrepancy.objects.select_related(
                "article",
                "location",
                "detected_by",
                "movement",
            ).order_by("-detected_at")[:20]
        ],
        "alarms": list_inventory_alarms(user),
        "safety_alerts": list_safety_stock_alerts(user) if can_manage_alarms else [],
        "minimum_stock_digest": (
            get_minimum_stock_digest_config(user, serialized_articles=serialized_articles)
            if can_manage_alarms
            else None
        ),
        "full_stock_report": (
            get_full_stock_report_config(user, serialized_articles=serialized_articles)
            if can_manage_alarms
            else None
        ),
        "automation_status": serialize_inventory_automation_status() if can_manage_alarms else None,
    }


def get_article_detail(article_id):
    """Devuelve article detail."""
    quantity_map, available_quantity_map, unit_total_map, unit_available_map = current_stock_maps()
    article = get_object_or_404(
        Article.objects.select_related(
            "unit_of_measure",
            "sector_responsible",
            "primary_location",
            "category",
            "subcategory",
            "supplier",
        ),
        pk=article_id,
    )

    return {
        "article": serialize_article(
            article,
            quantity_map,
            available_quantity_map,
            unit_total_map,
            unit_available_map,
        ),
        "balances": [
            serialize_balance(balance)
            for balance in InventoryBalance.objects.select_related("article", "location", "batch")
            .filter(article=article)
            .order_by("location__name", "batch__lot_code")[:20]
        ],
        "movements": [
            serialize_movement(movement)
            for movement in StockMovement.objects.select_related(
                "article",
                "recorded_by",
                "tracked_unit",
                "source_location",
                "target_location",
                "person",
                "sector",
                "authorized_by",
            )
            .filter(article=article)
            .order_by("-timestamp", "-id")[:15]
        ],
        "tracked_units": [
            serialize_tracked_unit(unit)
            for unit in TrackedUnit.objects.select_related(
                "article",
                "current_location",
                "current_sector",
                "current_holder_person",
            )
            .filter(article=article)
            .order_by("internal_tag")[:20]
        ],
    }


def create_or_update_batch(article, payload, user):
    """Crea or update batch."""
    batch_id = payload.get("batch_id")
    lot_code = (payload.get("lot_code") or "").strip()

    if batch_id:
        batch = get_object_or_404(InventoryBatch, pk=batch_id)
        if batch.article_id != article.id:
            raise InventoryApiError("The selected batch does not belong to the article")
        return batch

    if article.requires_lot and not lot_code:
        raise InventoryApiError("lot_code is required for this article")

    if not lot_code:
        return None

    batch, created = InventoryBatch.objects.get_or_create(
        article=article,
        lot_code=lot_code,
        defaults={
            "created_by": user,
            "updated_by": user,
            "expiry_date": payload.get("expiry_date") or None,
            "received_at": payload.get("received_at") or None,
            "supplier": resolve_instance(
                Supplier,
                payload.get("supplier_id"),
                "supplier",
                required=False,
            ),
            "document_ref": payload.get("document_ref") or "",
            "quality_status": payload.get("quality_status") or "",
            "notes": payload.get("notes") or "",
        },
    )
    if not created:
        update_audit(batch, user)
        if payload.get("expiry_date") is not None:
            batch.expiry_date = payload.get("expiry_date") or None
        if payload.get("received_at") is not None:
            batch.received_at = payload.get("received_at") or None
        if payload.get("supplier_id") is not None:
            batch.supplier = resolve_instance(
                Supplier,
                payload.get("supplier_id"),
                "supplier",
                required=False,
            )
        if payload.get("document_ref") is not None:
            batch.document_ref = payload.get("document_ref") or ""
        if payload.get("quality_status") is not None:
            batch.quality_status = payload.get("quality_status") or ""
        if payload.get("notes") is not None:
            batch.notes = payload.get("notes") or ""
        save_validated(batch)
    else:
        save_validated(batch)
    return batch


def _assign_article_fields(article, payload, user, files=None, is_new=False):
    """Maneja assign article fields."""
    article_type = resolve_article_type(payload.get("article_type") or article.article_type)
    is_critical = parse_boolean(payload.get("is_critical")) if "is_critical" in payload else article.is_critical
    if "tracking_mode" in payload and clean_string(payload.get("tracking_mode")):
        tracking_mode = payload.get("tracking_mode")
    elif article.pk:
        tracking_mode = article.tracking_mode
    else:
        tracking_mode = choose_tracking_mode(article_type, None)
    minimum_stock = (
        parse_optional_decimal(payload.get("minimum_stock"))
        if "minimum_stock" in payload
        else article.minimum_stock
    )

    if minimum_stock is None and should_require_minimum(article_type, is_critical):
        raise InventoryApiError("minimum_stock is required for this article")

    internal_code = clean_string(payload.get("internal_code")) if "internal_code" in payload else article.internal_code
    if not internal_code:
        internal_code = generate_article_code(article_type)

    article.internal_code = internal_code
    article.name = clean_string(payload.get("name")) if "name" in payload else article.name
    article.article_type = article_type
    article.tracking_mode = tracking_mode
    article.status = payload.get("status") or article.status or Article.ArticleStatus.ACTIVE
    article.unit_of_measure = resolve_instance(
        UnitOfMeasure,
        payload.get("unit_of_measure_id") if "unit_of_measure_id" in payload else article.unit_of_measure_id,
        "unit_of_measure",
    )
    article.sector_responsible = resolve_instance(
        Sector,
        payload.get("sector_responsible_id")
        if "sector_responsible_id" in payload
        else article.sector_responsible_id,
        "sector_responsible",
    )
    article.description = payload.get("description") if "description" in payload else article.description
    article.category = (
        resolve_instance(ArticleCategory, payload.get("category_id"), "category", required=False)
        if "category_id" in payload
        else article.category
    )
    article.subcategory = (
        resolve_instance(ArticleCategory, payload.get("subcategory_id"), "subcategory", required=False)
        if "subcategory_id" in payload
        else article.subcategory
    )
    article.primary_location = (
        resolve_instance(Location, payload.get("primary_location_id"), "primary_location", required=False)
        if "primary_location_id" in payload
        else article.primary_location
    )
    article.observations = payload.get("observations") if "observations" in payload else article.observations
    article.minimum_stock = minimum_stock
    article.safety_stock = (
        parse_optional_decimal(payload.get("safety_stock"))
        if "safety_stock" in payload
        else article.safety_stock
    )
    article.reorder_point = (
        parse_optional_decimal(payload.get("reorder_point"))
        if "reorder_point" in payload
        else article.reorder_point
    )
    article.max_stock = (
        parse_optional_decimal(payload.get("max_stock")) if "max_stock" in payload else article.max_stock
    )
    article.suggested_purchase_qty = (
        parse_optional_decimal(payload.get("suggested_purchase_qty"))
        if "suggested_purchase_qty" in payload
        else article.suggested_purchase_qty
    )
    article.supplier = (
        resolve_instance(Supplier, payload.get("supplier_id"), "supplier", required=False)
        if "supplier_id" in payload
        else article.supplier
    )
    article.reference_price = (
        parse_optional_decimal(payload.get("reference_price"))
        if "reference_price" in payload
        else article.reference_price
    )
    article.lead_time_days = (
        parse_optional_int(payload.get("lead_time_days"), "lead_time_days")
        if "lead_time_days" in payload
        else article.lead_time_days
    )
    article.last_purchase = (
        payload.get("last_purchase") or None
        if "last_purchase" in payload
        else article.last_purchase
    )
    article.requires_lot = parse_boolean(payload.get("requires_lot")) if "requires_lot" in payload else article.requires_lot
    article.requires_expiry = (
        parse_boolean(payload.get("requires_expiry"))
        if "requires_expiry" in payload
        else article.requires_expiry
    )
    article.requires_serial = (
        parse_boolean(payload.get("requires_serial"))
        if "requires_serial" in payload
        else article.requires_serial
    )
    article.requires_size = (
        parse_boolean(payload.get("requires_size"))
        if "requires_size" in payload
        else article.requires_size
    )
    article.requires_quality = (
        parse_boolean(payload.get("requires_quality"))
        if "requires_quality" in payload
        else article.requires_quality
    )
    article.requires_assignee = (
        parse_boolean(payload.get("requires_assignee"))
        if "requires_assignee" in payload
        else article.requires_assignee
    )
    article.is_critical = is_critical
    article.loanable = (
        parse_boolean(payload.get("loanable")) if "loanable" in payload else article.loanable
    ) or article_type == Article.ArticleType.TOOL

    if files and files.get("image"):
        article.image = files["image"]
    elif parse_boolean(payload.get("clear_image")):
        article.image.delete(save=False)
        article.image = None

    update_audit(article, user, is_new=is_new)
    save_validated(article)
    return article


def create_article(user, payload, files=None):
    """Crea article."""
    require_role(user, MASTER_ROLES)

    with transaction.atomic():
        article = _assign_article_fields(Article(), payload, user, files=files, is_new=True)

        initial_quantity = parse_optional_decimal(payload.get("initial_quantity")) or Decimal("0")
        if initial_quantity > 0:
            initial_location = resolve_instance(
                Location,
                payload.get("initial_location_id"),
                "initial_location",
                required=False,
            ) or article.primary_location or get_default_location()
            if not initial_location:
                raise InventoryApiError("A default location is required to load initial stock")

            movement_payload = {
                "article_id": article.id,
                "movement_type": StockMovement.MovementType.ADJUSTMENT_IN,
                "quantity": initial_quantity,
                "target_location_id": initial_location.id,
                "reason_text": "Carga inicial",
                "notes": "Carga inicial del articulo",
            }
            if article.tracking_mode == Article.TrackingMode.UNIT:
                movement_payload["movement_type"] = StockMovement.MovementType.PURCHASE_IN
            if payload.get("lot_code"):
                movement_payload["lot_code"] = payload.get("lot_code")
                movement_payload["expiry_date"] = payload.get("expiry_date")
            create_movement(user, movement_payload, allow_initial_load=True)

        return article


def update_article(user, article_id, payload, files=None):
    """Actualiza article."""
    require_role(user, MASTER_ROLES)

    with transaction.atomic():
        article = get_object_or_404(Article, pk=article_id)
        article = _assign_article_fields(article, payload, user, files=files, is_new=False)
        evaluate_safety_stock_alert(article)
        return article


def _resolve_quantity_article_movement(article, movement, payload, user):
    """Maneja resolve quantity article movement."""
    batch = create_or_update_batch(article, payload, user)
    source_location = movement.source_location or article.primary_location or get_default_location()
    target_location = movement.target_location or article.primary_location or get_default_location()

    if movement.movement_type in {
        StockMovement.MovementType.PURCHASE_IN,
        StockMovement.MovementType.RETURN_IN,
        StockMovement.MovementType.ADJUSTMENT_IN,
        StockMovement.MovementType.COUNT_ADJUST,
    }:
        if not target_location:
            raise InventoryApiError("target_location is required")
        movement.target_location = target_location
        apply_balance_delta(article, target_location, movement.quantity, user, batch=batch)
    elif movement.movement_type in {
        StockMovement.MovementType.CONSUMPTION_OUT,
        StockMovement.MovementType.PRODUCTION_OUT,
        StockMovement.MovementType.LOAN_OUT,
        StockMovement.MovementType.DAMAGE_OUT,
        StockMovement.MovementType.EXPIRED_OUT,
        StockMovement.MovementType.DISPOSAL_OUT,
    }:
        if not source_location:
            raise InventoryApiError("source_location is required")
        movement.source_location = source_location
        apply_balance_delta(article, source_location, movement.quantity * Decimal("-1"), user, batch=batch)
    elif movement.movement_type == StockMovement.MovementType.TRANSFER:
        if not source_location or not target_location:
            raise InventoryApiError("source_location and target_location are required")
        movement.source_location = source_location
        movement.target_location = target_location
        apply_balance_delta(article, source_location, movement.quantity * Decimal("-1"), user, batch=batch)
        apply_balance_delta(article, target_location, movement.quantity, user, batch=batch)
    else:
        raise InventoryApiError("Unsupported movement type")

    movement.batch = batch


def _resolve_unit_article_movement(article, movement, payload, user):
    """Maneja resolve unit article movement."""
    unit_id = payload.get("tracked_unit_id")
    source_location = movement.source_location or article.primary_location or get_default_location()
    target_location = movement.target_location or article.primary_location or get_default_location()

    if movement.movement_type in {
        StockMovement.MovementType.PURCHASE_IN,
        StockMovement.MovementType.ADJUSTMENT_IN,
    } and not unit_id:
        if movement.quantity != int(movement.quantity):
            raise InventoryApiError("Unit articles require integer quantities")
        count = int(movement.quantity)
        create_tracked_units(
            article,
            count,
            user,
            location=target_location,
            notes=movement.reason_text or movement.notes,
        )
        movement.target_location = target_location
        return

    tracked_unit = resolve_instance(TrackedUnit, unit_id, "tracked_unit")
    if tracked_unit.article_id != article.id:
        raise InventoryApiError("The tracked unit does not belong to the article")
    movement.tracked_unit = tracked_unit

    if movement.movement_type == StockMovement.MovementType.TRANSFER:
        tracked_unit.current_location = target_location
        tracked_unit.current_sector = article.sector_responsible
        tracked_unit.current_holder_person = None
        tracked_unit.status = TrackedUnit.UnitStatus.AVAILABLE
        update_audit(tracked_unit, user)
        save_validated(tracked_unit)
        movement.source_location = source_location
        movement.target_location = target_location
        return

    if movement.movement_type == StockMovement.MovementType.RETURN_IN:
        tracked_unit.current_location = target_location
        tracked_unit.current_sector = article.sector_responsible
        tracked_unit.current_holder_person = None
        tracked_unit.status = TrackedUnit.UnitStatus.AVAILABLE
        update_audit(tracked_unit, user)
        save_validated(tracked_unit)
        movement.target_location = target_location
        return

    if movement.movement_type == StockMovement.MovementType.DAMAGE_OUT:
        tracked_unit.status = TrackedUnit.UnitStatus.OUT_OF_SERVICE
    elif movement.movement_type == StockMovement.MovementType.DISPOSAL_OUT:
        tracked_unit.status = TrackedUnit.UnitStatus.RETIRED
        tracked_unit.retired_at = timezone.now().date()
    else:
        raise InventoryApiError("Unsupported unit movement type")

    tracked_unit.current_holder_person = None
    tracked_unit.current_location = source_location
    tracked_unit.current_sector = article.sector_responsible
    update_audit(tracked_unit, user)
    save_validated(tracked_unit)
    movement.source_location = source_location


def create_movement(user, payload, allow_initial_load=False, bypass_role_check=False):
    """Crea movement."""
    profile = get_profile(user) if bypass_role_check else require_role(user, MOVEMENT_ROLES)

    movement_type = payload.get("movement_type")
    article = resolve_instance(Article, payload.get("article_id"), "article")
    quantity = parse_decimal(payload.get("quantity"), "quantity")
    if quantity <= 0:
        raise InventoryApiError("quantity must be greater than zero")

    reason_text = (payload.get("reason_text") or "").strip()
    sensitive_types = {
        StockMovement.MovementType.ADJUSTMENT_IN,
        StockMovement.MovementType.DAMAGE_OUT,
        StockMovement.MovementType.EXPIRED_OUT,
        StockMovement.MovementType.COUNT_ADJUST,
        StockMovement.MovementType.DISPOSAL_OUT,
    }
    can_bypass_sensitive = (
        allow_initial_load and movement_type == StockMovement.MovementType.ADJUSTMENT_IN
    )
    if (
        movement_type in sensitive_types
        and profile.role not in APPROVER_ROLES
        and not user.is_superuser
        and not can_bypass_sensitive
    ):
        raise InventoryApiError("Only supervisor roles can register sensitive movements", status=403)

    with transaction.atomic():
        previous_stock = (
            current_stock_for_article(article) if article.minimum_stock is not None else None
        )
        movement = StockMovement(
            movement_type=movement_type,
            article=article,
            quantity=quantity,
            recorded_by=user,
            source_location=resolve_instance(
                Location,
                payload.get("source_location_id"),
                "source_location",
                required=False,
            ),
            target_location=resolve_instance(
                Location,
                payload.get("target_location_id"),
                "target_location",
                required=False,
            ),
            person=resolve_instance(Person, payload.get("person_id"), "person", required=False),
            sector=resolve_instance(Sector, payload.get("sector_id"), "sector", required=False),
            reason_text=reason_text,
            document_ref=payload.get("document_ref") or "",
            authorized_by=user if movement_type in sensitive_types and (user.is_superuser or profile.role in APPROVER_ROLES or can_bypass_sensitive) else None,
            notes=payload.get("notes") or "",
        )
        update_audit(movement, user, is_new=True)

        if article.tracking_mode == Article.TrackingMode.QUANTITY:
            _resolve_quantity_article_movement(article, movement, payload, user)
        else:
            _resolve_unit_article_movement(article, movement, payload, user)

        save_validated(movement)
        current_stock = (
            current_stock_for_article(article) if article.minimum_stock is not None else None
        )
        evaluate_safety_stock_alert(article)
        maybe_create_purchase_request_for_minimum_stock(
            article,
            previous_stock=previous_stock,
            current_stock=current_stock,
            triggered_by_user=user,
            requester_person=movement.person,
            requester_sector=movement.sector,
            source_label=f"movement:{movement.movement_type}",
        )
        evaluate_purchasing_minimum_stock_alarm(article)
        return movement


def create_checkout(user, payload):
    """Crea checkout."""
    require_role(user, CHECKOUT_ROLES)

    with transaction.atomic():
        tracked_unit = resolve_instance(TrackedUnit, payload.get("tracked_unit_id"), "tracked_unit")
        article = tracked_unit.article
        previous_stock = (
            current_stock_for_article(article) if article.minimum_stock is not None else None
        )
        if not article.loanable:
            raise InventoryApiError("This article is not configured as loanable")
        if tracked_unit.status == TrackedUnit.UnitStatus.CHECKED_OUT:
            raise InventoryApiError("The unit is already checked out")

        receiver_person = resolve_instance(
            Person,
            payload.get("receiver_person_id"),
            "receiver_person",
            required=False,
        )
        receiver_sector = resolve_instance(
            Sector,
            payload.get("receiver_sector_id"),
            "receiver_sector",
            required=False,
        )
        if article.requires_assignee and not receiver_person:
            raise InventoryApiError("This article requires an assigned person")

        checkout = AssetCheckout(
            tracked_unit=tracked_unit,
            receiver_person=receiver_person,
            receiver_sector=receiver_sector,
            checkout_kind=payload.get("checkout_kind") or AssetCheckout.CheckoutKind.LOAN,
            recorded_by=user,
            expected_return_at=payload.get("expected_return_at") or None,
            condition_out=payload.get("condition_out") or "",
            notes=payload.get("notes") or "",
        )
        update_audit(checkout, user, is_new=True)
        save_validated(checkout)

        tracked_unit.status = TrackedUnit.UnitStatus.CHECKED_OUT
        tracked_unit.current_holder_person = receiver_person
        tracked_unit.current_sector = receiver_sector or (
            receiver_person.sector if receiver_person and receiver_person.sector else tracked_unit.current_sector
        )
        tracked_unit.current_location = None
        update_audit(tracked_unit, user)
        save_validated(tracked_unit)

        movement = StockMovement(
            movement_type=StockMovement.MovementType.LOAN_OUT,
            article=article,
            quantity=Decimal("1"),
            recorded_by=user,
            tracked_unit=tracked_unit,
            source_location=article.primary_location or get_default_location(),
            person=receiver_person,
            sector=receiver_sector,
            reason_text="Prestamo interno",
            notes=payload.get("notes") or "",
        )
        update_audit(movement, user, is_new=True)
        save_validated(movement)
        current_stock = (
            current_stock_for_article(article) if article.minimum_stock is not None else None
        )
        evaluate_safety_stock_alert(article)
        maybe_create_purchase_request_for_minimum_stock(
            article,
            previous_stock=previous_stock,
            current_stock=current_stock,
            triggered_by_user=user,
            requester_person=receiver_person,
            requester_sector=receiver_sector,
            source_label="checkout:loan_out",
        )
        evaluate_purchasing_minimum_stock_alarm(article)
        return checkout


def return_checkout(user, checkout_id, payload):
    """Maneja return checkout."""
    require_role(user, CHECKOUT_ROLES)

    with transaction.atomic():
        checkout = get_object_or_404(
            AssetCheckout.objects.select_related("tracked_unit__article"),
            pk=checkout_id,
        )
        article = checkout.tracked_unit.article
        previous_stock = (
            current_stock_for_article(article) if article.minimum_stock is not None else None
        )
        if checkout.status != AssetCheckout.CheckoutStatus.OPEN:
            raise InventoryApiError("The checkout is already closed")

        target_location = resolve_instance(
            Location,
            payload.get("target_location_id"),
            "target_location",
            required=False,
        ) or checkout.tracked_unit.article.primary_location or get_default_location()
        if not target_location:
            raise InventoryApiError("A return location is required")

        checkout.status = AssetCheckout.CheckoutStatus.RETURNED
        checkout.returned_at = timezone.now()
        checkout.condition_in = payload.get("condition_in") or ""
        checkout.notes = payload.get("notes") or checkout.notes
        update_audit(checkout, user)
        save_validated(checkout)

        tracked_unit = checkout.tracked_unit
        tracked_unit.status = TrackedUnit.UnitStatus.AVAILABLE
        tracked_unit.current_holder_person = None
        tracked_unit.current_sector = tracked_unit.article.sector_responsible
        tracked_unit.current_location = target_location
        update_audit(tracked_unit, user)
        save_validated(tracked_unit)

        movement = StockMovement(
            movement_type=StockMovement.MovementType.RETURN_IN,
            article=tracked_unit.article,
            quantity=Decimal("1"),
            recorded_by=user,
            tracked_unit=tracked_unit,
            target_location=target_location,
            reason_text="Devolucion interna",
            notes=payload.get("notes") or "",
        )
        update_audit(movement, user, is_new=True)
        save_validated(movement)
        current_stock = (
            current_stock_for_article(article) if article.minimum_stock is not None else None
        )
        evaluate_safety_stock_alert(article)
        maybe_create_purchase_request_for_minimum_stock(
            article,
            previous_stock=previous_stock,
            current_stock=current_stock,
            triggered_by_user=user,
            requester_person=None,
            requester_sector=None,
            source_label="checkout:return_in",
        )
        evaluate_purchasing_minimum_stock_alarm(article)
        return checkout


def create_count_session(user, payload):
    """Crea count session."""
    require_role(user, COUNT_ROLES)

    session = PhysicalCountSession(
        count_type=payload.get("count_type") or PhysicalCountSession.CountType.PARTIAL,
        scope=(payload.get("scope") or "").strip() or "Conteo operativo",
        scheduled_for=payload.get("scheduled_for") or timezone.now(),
        status=PhysicalCountSession.CountStatus.OPEN,
        created_by=user,
        updated_by=user,
        notes=payload.get("notes") or "",
    )
    save_validated(session)
    return session


def system_quantity_for(article, location):
    """Maneja system quantity for."""
    if article.tracking_mode == Article.TrackingMode.UNIT:
        return Decimal(
            TrackedUnit.objects.filter(
                article=article,
                current_location=location,
            )
            .exclude(status__in=[TrackedUnit.UnitStatus.RETIRED, TrackedUnit.UnitStatus.LOST])
            .count()
        )

    return (
        InventoryBalance.objects.filter(article=article, location=location).aggregate(total=Sum("on_hand"))[
            "total"
        ]
        or Decimal("0")
    )


def add_count_line(user, session_id, payload):
    """Maneja add count line."""
    require_role(user, COUNT_ROLES)

    from .models import PhysicalCountLine

    with transaction.atomic():
        session = get_object_or_404(PhysicalCountSession, pk=session_id)
        article = resolve_instance(Article, payload.get("article_id"), "article")
        location = resolve_instance(Location, payload.get("location_id"), "location")
        counter_person = resolve_instance(Person, payload.get("counter_person_id"), "counter_person")
        counted_qty = parse_decimal(payload.get("counted_qty"), "counted_qty")
        system_qty = system_quantity_for(article, location)

        line = PhysicalCountLine(
            session=session,
            article=article,
            location=location,
            system_qty=system_qty,
            counted_qty=counted_qty,
            counter_person=counter_person,
            notes=payload.get("notes") or "",
            possible_cause=payload.get("possible_cause") or "",
        )
        update_audit(line, user, is_new=True)
        save_validated(line)

        if counted_qty != system_qty:
            discrepancy = StockDiscrepancy(
                article=article,
                location=location,
                count_line=line,
                difference_qty=abs(counted_qty - system_qty),
                difference_type=StockDiscrepancy.DifferenceType.POSITIVE
                if counted_qty > system_qty
                else StockDiscrepancy.DifferenceType.NEGATIVE,
                detected_by=user,
                possible_cause=payload.get("possible_cause") or "",
                comment=payload.get("notes") or "",
            )
            update_audit(discrepancy, user, is_new=True)
            save_validated(discrepancy)
            session.status = PhysicalCountSession.CountStatus.REVIEW
            update_audit(session, user)
            save_validated(session)

        return line


def create_discrepancy(user, payload):
    """Crea discrepancy."""
    require_role(user, COUNT_ROLES)

    discrepancy = StockDiscrepancy(
        article=resolve_instance(Article, payload.get("article_id"), "article"),
        location=resolve_instance(Location, payload.get("location_id"), "location", required=False),
        difference_qty=parse_decimal(payload.get("difference_qty"), "difference_qty"),
        difference_type=payload.get("difference_type"),
        detected_by=user,
        possible_cause=payload.get("possible_cause") or "",
        comment=payload.get("comment") or "",
        evidence=payload.get("evidence") or "",
    )
    update_audit(discrepancy, user, is_new=True)
    save_validated(discrepancy)
    return discrepancy


def resolve_discrepancy(user, discrepancy_id, payload):
    """Maneja resolve discrepancy."""
    require_role(user, APPROVER_ROLES)

    with transaction.atomic():
        discrepancy = get_object_or_404(
            StockDiscrepancy.objects.select_related("article", "location"),
            pk=discrepancy_id,
        )
        if discrepancy.status != StockDiscrepancy.DiscrepancyStatus.OPEN:
            raise InventoryApiError("The discrepancy is already resolved")

        reason_text = (payload.get("reason_text") or "").strip() or "Ajuste por diferencia de inventario"
        article = discrepancy.article
        previous_stock = (
            current_stock_for_article(article) if article.minimum_stock is not None else None
        )
        location = discrepancy.location or article.primary_location or get_default_location()
        if not location:
            raise InventoryApiError("A location is required to resolve the discrepancy")

        if article.tracking_mode == Article.TrackingMode.QUANTITY:
            movement = StockMovement(
                movement_type=StockMovement.MovementType.COUNT_ADJUST,
                article=article,
                quantity=discrepancy.difference_qty,
                recorded_by=user,
                source_location=location
                if discrepancy.difference_type == StockDiscrepancy.DifferenceType.NEGATIVE
                else None,
                target_location=location
                if discrepancy.difference_type == StockDiscrepancy.DifferenceType.POSITIVE
                else None,
                reason_text=reason_text,
                authorized_by=user,
                notes=payload.get("notes") or "",
            )
            update_audit(movement, user, is_new=True)
            save_validated(movement)

            if discrepancy.difference_type == StockDiscrepancy.DifferenceType.POSITIVE:
                apply_balance_delta(article, location, discrepancy.difference_qty, user)
            else:
                apply_balance_delta(article, location, discrepancy.difference_qty * Decimal("-1"), user)
        else:
            unit_count = int(discrepancy.difference_qty)
            if discrepancy.difference_qty != unit_count:
                raise InventoryApiError("Unit discrepancies must use integer quantities")

            movement = StockMovement(
                movement_type=StockMovement.MovementType.COUNT_ADJUST,
                article=article,
                quantity=discrepancy.difference_qty,
                recorded_by=user,
                target_location=location,
                reason_text=reason_text,
                authorized_by=user,
                notes=payload.get("notes") or "",
            )
            update_audit(movement, user, is_new=True)
            save_validated(movement)

            if discrepancy.difference_type == StockDiscrepancy.DifferenceType.POSITIVE:
                create_tracked_units(
                    article,
                    unit_count,
                    user,
                    location=location,
                    notes=reason_text,
                )
            else:
                units = list(
                    TrackedUnit.objects.filter(article=article, current_location=location)
                    .exclude(status__in=[TrackedUnit.UnitStatus.RETIRED, TrackedUnit.UnitStatus.LOST])[:unit_count]
                )
                if len(units) < unit_count:
                    raise InventoryApiError("There are not enough units to mark as missing")
                for unit in units:
                    unit.status = TrackedUnit.UnitStatus.LOST
                    unit.current_location = None
                    unit.current_holder_person = None
                    update_audit(unit, user)
                    save_validated(unit)

        discrepancy.status = StockDiscrepancy.DiscrepancyStatus.RESOLVED
        discrepancy.approved_by = user
        discrepancy.action_taken = payload.get("action_taken") or "Ajuste aplicado"
        discrepancy.comment = payload.get("notes") or discrepancy.comment
        discrepancy.movement = movement
        current_stock = (
            current_stock_for_article(article) if article.minimum_stock is not None else None
        )
        evaluate_safety_stock_alert(article)
        maybe_create_purchase_request_for_minimum_stock(
            article,
            previous_stock=previous_stock,
            current_stock=current_stock,
            triggered_by_user=user,
            requester_person=None,
            requester_sector=None,
            source_label="discrepancy:count_adjust",
        )
        evaluate_purchasing_minimum_stock_alarm(article)
        discrepancy.resolved_at = timezone.now()
        update_audit(discrepancy, user)
        save_validated(discrepancy)
        return discrepancy


EXCEL_IMPORT_ALIASES = {
    "internal_code": {"codigo", "codigo_interno", "codigo interno", "internal_code", "sku"},
    "name": {"nombre", "articulo", "producto", "name"},
    "article_type": {"tipo", "tipo_articulo", "tipo de articulo", "article_type"},
    "unit": {"unidad", "unidad_medida", "unidad de medida", "uom", "unit"},
    "sector": {"sector", "sector_responsable", "sector responsable"},
    "location": {"ubicacion", "ubicacion_principal", "location", "deposito"},
    "category": {"categoria", "category"},
    "subcategory": {"subcategoria", "subcategory"},
    "description": {"descripcion", "description"},
    "observations": {"observaciones", "notes", "observations"},
    "supplier": {"proveedor", "supplier"},
    "minimum_stock": {"stock_minimo", "stock minimo", "minimum_stock", "minimo"},
    "initial_quantity": {"stock_inicial", "stock inicial", "initial_quantity", "cantidad_inicial"},
    "status": {"estado", "status"},
    "tracking_mode": {"tracking", "tracking_mode", "modo_tracking"},
    "loanable": {"prestable", "loanable"},
    "is_critical": {"critico", "critical", "is_critical"},
    "requires_lot": {"usa_lote", "requiere_lote", "requires_lot"},
    "requires_expiry": {"usa_vencimiento", "requiere_vencimiento", "requires_expiry"},
}


EXCEL_IMPORT_REQUIRED_COLUMNS = {"name", "article_type", "unit", "sector"}
EXCEL_IMPORT_MAX_COLUMNS = 12
EXCEL_IMPORT_HEADER_SCAN_ROWS = 12
EXCEL_IMPORT_SIMPLE_HEADERS = {"nombre", "articulo", "producto", "descripcion"}


def _import_text(value):
    """Maneja import text."""
    if value in (None, ""):
        return ""
    return str(value).strip()


def _import_is_blank(value):
    """Maneja import is blank."""
    return value is None or (isinstance(value, str) and not value.strip())


def _normalize_import_key(value):
    """Maneja normalize import key."""
    collapsed = re.sub(r"\s+", " ", _import_text(value))
    normalized = unicodedata.normalize("NFKD", collapsed)
    ascii_only = normalized.encode("ascii", "ignore").decode("ascii")
    return ascii_only.casefold()


def _import_error_detail(exc):
    """Maneja import error detail."""
    detail = exc.detail if isinstance(exc, InventoryApiError) else str(exc)
    if isinstance(detail, dict):
        return "; ".join(
            f"{key}: {', '.join(str(item) for item in value) if isinstance(value, list) else value}"
            for key, value in detail.items()
        )
    if isinstance(detail, list):
        return "; ".join(str(item) for item in detail)
    return str(detail)


def _map_excel_columns(headers, required_columns=None):
    """Maneja map excel columns."""
    column_map = {}
    for index, header in enumerate(headers):
        normalized = normalize_excel_header(header)
        if not normalized:
            continue

        for target_field, aliases in EXCEL_IMPORT_ALIASES.items():
            normalized_aliases = {normalize_excel_header(alias) for alias in aliases}
            if normalized in normalized_aliases and target_field not in column_map:
                column_map[target_field] = index
                break

    if required_columns:
        missing = sorted(set(required_columns) - set(column_map))
        if missing:
            return None
    return column_map


def _excel_cell(row, column_map, key):
    """Maneja excel cell."""
    index = column_map.get(key)
    if index is None or index >= len(row):
        return ""
    return row[index]


def _sheet_row_values(worksheet, row_number, max_columns=EXCEL_IMPORT_MAX_COLUMNS):
    """Maneja sheet row values."""
    values = [worksheet.cell(row=row_number, column=column).value for column in range(1, max_columns + 1)]
    while values and _import_is_blank(values[-1]):
        values.pop()
    return values


def _iter_sheet_rows(worksheet, start_row=1, max_columns=EXCEL_IMPORT_MAX_COLUMNS):
    """Maneja iter sheet rows."""
    for row_number in range(start_row, worksheet.max_row + 1):
        values = _sheet_row_values(worksheet, row_number, max_columns=max_columns)
        if any(not _import_is_blank(value) for value in values):
            yield row_number, values


def _build_import_context():
    """Maneja build import context."""
    units = list(UnitOfMeasure.objects.all())
    sectors = list(Sector.objects.all())
    locations = list(Location.objects.all())
    suppliers = list(Supplier.objects.all())

    default_unit = (
        resolve_optional_catalog_by_name_or_code(units, "UN")
        or resolve_optional_catalog_by_name_or_code(units, "Unidad")
        or (units[0] if units else None)
    )
    default_sector = (
        resolve_optional_catalog_by_name_or_code(sectors, "DEP")
        or resolve_optional_catalog_by_name_or_code(sectors, "Deposito")
        or (sectors[0] if sectors else None)
    )
    default_location = get_default_location()

    return {
        "units": units,
        "sectors": sectors,
        "locations": locations,
        "suppliers": suppliers,
        "default_unit": default_unit,
        "default_sector": default_sector,
        "default_location": default_location,
    }


def _structured_header_match(worksheet):
    """Maneja structured header match."""
    for row_number in range(1, min(worksheet.max_row, EXCEL_IMPORT_HEADER_SCAN_ROWS) + 1):
        values = _sheet_row_values(worksheet, row_number)
        if not values:
            continue
        column_map = _map_excel_columns(values, required_columns=EXCEL_IMPORT_REQUIRED_COLUMNS)
        if column_map:
            return row_number, column_map
    return None


def _simple_list_header_row(worksheet):
    """Maneja simple list header row."""
    for row_number in range(1, min(worksheet.max_row, EXCEL_IMPORT_HEADER_SCAN_ROWS) + 1):
        values = _sheet_row_values(worksheet, row_number)
        if not values:
            continue
        if len(values) != 1:
            continue
        if normalize_excel_header(values[0]) in EXCEL_IMPORT_SIMPLE_HEADERS:
            return row_number
    return None


def _default_sector_for_simple_row(context, sheet_name, category_name):
    """Maneja default sector for simple row."""
    text = _normalize_import_key(f"{sheet_name} {category_name}")
    if "mantenimiento" in text or "motor" in text:
        return (
            resolve_optional_catalog_by_name_or_code(context["sectors"], "Mantenimiento")
            or context["default_sector"]
        )
    if "produccion" in text:
        return (
            resolve_optional_catalog_by_name_or_code(context["sectors"], "Produccion")
            or context["default_sector"]
        )
    return context["default_sector"]


def _default_unit_for_simple_row(context, name):
    """Maneja default unit for simple row."""
    text = _normalize_import_key(name)
    if any(token in text for token in {" par", "(par)", "pares", " par)"}):
        return resolve_optional_catalog_by_name_or_code(context["units"], "Par") or context["default_unit"]
    if any(token in text for token in {"kilo", "kg"}):
        return (
            resolve_optional_catalog_by_name_or_code(context["units"], "Kilogramo")
            or context["default_unit"]
        )
    if any(token in text for token in {"litro", "litros", "lts", " lt", " cc"}):
        return resolve_optional_catalog_by_name_or_code(context["units"], "Litro") or context["default_unit"]
    if any(token in text for token in {"metro", "metros"}):
        return resolve_optional_catalog_by_name_or_code(context["units"], "Metro") or context["default_unit"]
    return context["default_unit"]


def _infer_simple_article_type(sheet_name, category_name):
    """Maneja infer simple article type."""
    text = _normalize_import_key(f"{sheet_name} {category_name}")
    if "proteccion personal" in text or normalize_excel_header(sheet_name) == "epp":
        return Article.ArticleType.PPE
    if "produccion" in text:
        return Article.ArticleType.INPUT
    if "mantenimiento" in text or "motor" in text or "repuesto" in text:
        return Article.ArticleType.SPARE_PART
    return Article.ArticleType.CONSUMABLE


def _build_import_candidate(
    *,
    row_number,
    sheet_name,
    payload,
    unit,
    sector,
    location,
    category_name="",
    subcategory_name="",
    source_mode,
):
    """Maneja build import candidate."""
    article_type = payload["article_type"]
    return {
        "row": row_number,
        "sheet_name": sheet_name,
        "source_mode": source_mode,
        "name": payload["name"],
        "internal_code": payload.get("internal_code") or "",
        "article_type": article_type,
        "article_type_label": dict(Article.ArticleType.choices).get(article_type, article_type),
        "tracking_mode": payload["tracking_mode"],
        "tracking_mode_label": dict(Article.TrackingMode.choices).get(
            payload["tracking_mode"],
            payload["tracking_mode"],
        ),
        "unit_name": unit.name,
        "sector_name": sector.name,
        "location_name": location.name if location else "",
        "category_name": category_name,
        "subcategory_name": subcategory_name,
        "minimum_stock": serialize_decimal(parse_optional_decimal(payload.get("minimum_stock"))),
        "initial_quantity": serialize_decimal(parse_optional_decimal(payload.get("initial_quantity"))),
        "payload": payload,
        "decision": "ready",
        "detail": "",
    }


def _build_structured_import_candidate(row_number, sheet_name, row, column_map, context):
    """Maneja build structured import candidate."""
    article_type = resolve_article_type(_excel_cell(row, column_map, "article_type"))
    unit = resolve_catalog_by_name_or_code(context["units"], _excel_cell(row, column_map, "unit"), "unit")
    sector = resolve_catalog_by_name_or_code(
        context["sectors"],
        _excel_cell(row, column_map, "sector"),
        "sector",
    )
    location = resolve_optional_catalog_by_name_or_code(
        context["locations"],
        _excel_cell(row, column_map, "location"),
    )
    supplier = resolve_optional_catalog_by_name_or_code(
        context["suppliers"],
        _excel_cell(row, column_map, "supplier"),
    )
    category_name = _import_text(_excel_cell(row, column_map, "category"))
    subcategory_name = _import_text(_excel_cell(row, column_map, "subcategory"))
    is_critical = parse_boolean(_excel_cell(row, column_map, "is_critical"))
    minimum_stock = _excel_cell(row, column_map, "minimum_stock")
    initial_quantity = _excel_cell(row, column_map, "initial_quantity")

    parse_optional_decimal(minimum_stock)
    parse_optional_decimal(initial_quantity)
    if _import_is_blank(minimum_stock) and should_require_minimum(article_type, is_critical):
        raise InventoryApiError("minimum_stock is required for this article")

    payload = {
        "internal_code": _import_text(_excel_cell(row, column_map, "internal_code")),
        "name": _import_text(_excel_cell(row, column_map, "name")),
        "article_type": article_type,
        "unit_of_measure_id": unit.id,
        "sector_responsible_id": sector.id,
        "primary_location_id": location.id if location else "",
        "category_name": category_name,
        "subcategory_name": subcategory_name,
        "description": _import_text(_excel_cell(row, column_map, "description")),
        "observations": _import_text(_excel_cell(row, column_map, "observations")),
        "minimum_stock": minimum_stock,
        "initial_quantity": initial_quantity,
        "status": resolve_choice_value(
            _excel_cell(row, column_map, "status"),
            Article.ArticleStatus.choices,
            "status",
            default=Article.ArticleStatus.ACTIVE,
        ),
        "tracking_mode": resolve_choice_value(
            _excel_cell(row, column_map, "tracking_mode"),
            Article.TrackingMode.choices,
            "tracking mode",
            default=choose_tracking_mode(article_type, None),
        ),
        "loanable": parse_boolean(_excel_cell(row, column_map, "loanable")),
        "is_critical": is_critical,
        "requires_lot": parse_boolean(_excel_cell(row, column_map, "requires_lot")),
        "requires_expiry": parse_boolean(_excel_cell(row, column_map, "requires_expiry")),
        "supplier_id": supplier.id if supplier else "",
    }
    if location:
        payload["initial_location_id"] = location.id

    return _build_import_candidate(
        row_number=row_number,
        sheet_name=sheet_name,
        payload=payload,
        unit=unit,
        sector=sector,
        location=location,
        category_name=category_name,
        subcategory_name=subcategory_name,
        source_mode="structured",
    )


def _build_simple_import_candidate(row_number, sheet_name, name, category_name, context):
    """Maneja build simple import candidate."""
    article_type = _infer_simple_article_type(sheet_name, category_name)
    unit = _default_unit_for_simple_row(context, name)
    sector = _default_sector_for_simple_row(context, sheet_name, category_name)
    location = context["default_location"]
    if not unit:
        raise InventoryApiError("No default unit is available for Excel imports")
    if not sector:
        raise InventoryApiError("No default sector is available for Excel imports")

    minimum_stock = "0" if should_require_minimum(article_type, False) else ""
    payload = {
        "internal_code": "",
        "name": _import_text(name),
        "article_type": article_type,
        "unit_of_measure_id": unit.id,
        "sector_responsible_id": sector.id,
        "primary_location_id": location.id if location else "",
        "category_name": category_name,
        "subcategory_name": "",
        "description": "",
        "observations": "",
        "minimum_stock": minimum_stock,
        "initial_quantity": "",
        "status": Article.ArticleStatus.ACTIVE,
        "tracking_mode": choose_tracking_mode(article_type, None),
        "loanable": article_type == Article.ArticleType.TOOL,
        "is_critical": False,
        "requires_lot": False,
        "requires_expiry": False,
        "supplier_id": "",
    }
    if location:
        payload["initial_location_id"] = location.id

    return _build_import_candidate(
        row_number=row_number,
        sheet_name=sheet_name,
        payload=payload,
        unit=unit,
        sector=sector,
        location=location,
        category_name=category_name,
        source_mode="simple_list",
    )


def _parse_structured_sheet(worksheet, context, header_row, column_map):
    """Maneja parse structured sheet."""
    candidates = []
    for row_number in range(header_row + 1, worksheet.max_row + 1):
        row = _sheet_row_values(worksheet, row_number)
        if not row:
            continue
        try:
            candidates.append(
                _build_structured_import_candidate(
                    row_number,
                    worksheet.title,
                    row,
                    column_map,
                    context,
                )
            )
        except Exception as exc:  # noqa: BLE001
            candidates.append(
                {
                    "row": row_number,
                    "sheet_name": worksheet.title,
                    "source_mode": "structured",
                    "name": _import_text(_excel_cell(row, column_map, "name")),
                    "internal_code": _import_text(_excel_cell(row, column_map, "internal_code")),
                    "article_type": "",
                    "article_type_label": "",
                    "tracking_mode": "",
                    "tracking_mode_label": "",
                    "unit_name": "",
                    "sector_name": "",
                    "location_name": "",
                    "category_name": "",
                    "subcategory_name": "",
                    "minimum_stock": None,
                    "initial_quantity": None,
                    "payload": None,
                    "decision": "error",
                    "detail": _import_error_detail(exc),
                }
            )
    return candidates


def _parse_simple_sheet(worksheet, context, header_row):
    """Maneja parse simple sheet."""
    candidates = []
    current_category = ""
    for row_number in range(header_row + 1, worksheet.max_row + 1):
        cell = worksheet.cell(row=row_number, column=1)
        raw_value = _import_text(cell.value)
        if not raw_value:
            continue

        if cell.font and cell.font.bold:
            current_category = raw_value
            continue

        try:
            candidates.append(
                _build_simple_import_candidate(
                    row_number,
                    worksheet.title,
                    raw_value,
                    current_category,
                    context,
                )
            )
        except Exception as exc:  # noqa: BLE001
            candidates.append(
                {
                    "row": row_number,
                    "sheet_name": worksheet.title,
                    "source_mode": "simple_list",
                    "name": raw_value,
                    "internal_code": "",
                    "article_type": "",
                    "article_type_label": "",
                    "tracking_mode": "",
                    "tracking_mode_label": "",
                    "unit_name": "",
                    "sector_name": "",
                    "location_name": "",
                    "category_name": current_category,
                    "subcategory_name": "",
                    "minimum_stock": None,
                    "initial_quantity": None,
                    "payload": None,
                    "decision": "error",
                    "detail": _import_error_detail(exc),
                }
            )
    return candidates


def _collect_excel_import_candidates(workbook):
    """Maneja collect excel import candidates."""
    context = _build_import_context()
    candidates = []
    sheet_summaries = []

    for worksheet in workbook.worksheets:
        header_match = _structured_header_match(worksheet)
        if header_match:
            header_row, column_map = header_match
            sheet_candidates = _parse_structured_sheet(worksheet, context, header_row, column_map)
            candidates.extend(sheet_candidates)
            sheet_summaries.append(
                {
                    "sheet_name": worksheet.title,
                    "mode": "structured",
                    "candidate_count": len(sheet_candidates),
                }
            )
            continue

        simple_header_row = _simple_list_header_row(worksheet)
        if simple_header_row:
            sheet_candidates = _parse_simple_sheet(worksheet, context, simple_header_row)
            candidates.extend(sheet_candidates)
            sheet_summaries.append(
                {
                    "sheet_name": worksheet.title,
                    "mode": "simple_list",
                    "candidate_count": len(sheet_candidates),
                }
            )

    if not candidates:
        raise InventoryApiError(
            "No se detectaron productos importables en el Excel. Usa una tabla estructurada o una lista simple con la primera columna 'nombre'."
        )

    return candidates, sheet_summaries


def _classify_import_candidates(candidates):
    """Maneja classify import candidates."""
    existing_name_keys = {
        _normalize_import_key(name): True for name in Article.objects.values_list("name", flat=True)
    }
    existing_code_keys = {
        _normalize_import_key(code): True
        for code in Article.objects.exclude(internal_code="").values_list("internal_code", flat=True)
    }
    seen_name_keys = set()
    seen_code_keys = set()

    for candidate in candidates:
        if candidate["decision"] != "ready" or not candidate["payload"]:
            continue

        name_key = _normalize_import_key(candidate["payload"]["name"])
        code_key = _normalize_import_key(candidate["payload"].get("internal_code"))

        if not name_key:
            candidate["decision"] = "error"
            candidate["detail"] = "name is required"
            continue

        if code_key and code_key in existing_code_keys:
            candidate["decision"] = "skip"
            candidate["detail"] = "Ya existe un articulo con ese codigo."
            continue

        if name_key in existing_name_keys:
            candidate["decision"] = "skip"
            candidate["detail"] = "Ya existe un articulo con ese nombre."
            continue

        if code_key and code_key in seen_code_keys:
            candidate["decision"] = "skip"
            candidate["detail"] = "Codigo repetido dentro del mismo Excel."
            continue

        if name_key in seen_name_keys:
            candidate["decision"] = "skip"
            candidate["detail"] = "Nombre repetido dentro del mismo Excel."
            continue

        seen_name_keys.add(name_key)
        if code_key:
            seen_code_keys.add(code_key)


def _candidate_category_id(category_name, user):
    """Maneja candidate category id."""
    category = get_or_create_category_by_name(category_name, user)
    return category.id if category else ""


def _create_article_from_import_candidate(user, candidate):
    """Maneja create article from import candidate."""
    payload = dict(candidate["payload"])
    payload["category_id"] = _candidate_category_id(payload.pop("category_name", ""), user)
    payload["subcategory_id"] = _candidate_category_id(payload.pop("subcategory_name", ""), user)
    return create_article(user, payload)


def _serialize_import_candidate(candidate):
    """Maneja serialize import candidate."""
    return {
        "row": candidate["row"],
        "sheet_name": candidate["sheet_name"],
        "source_mode": candidate["source_mode"],
        "name": candidate["name"],
        "internal_code": candidate["internal_code"],
        "article_type": candidate["article_type"],
        "article_type_label": candidate["article_type_label"],
        "tracking_mode": candidate["tracking_mode"],
        "tracking_mode_label": candidate["tracking_mode_label"],
        "unit_name": candidate["unit_name"],
        "sector_name": candidate["sector_name"],
        "location_name": candidate["location_name"],
        "category_name": candidate["category_name"],
        "subcategory_name": candidate["subcategory_name"],
        "minimum_stock": candidate["minimum_stock"],
        "initial_quantity": candidate["initial_quantity"],
        "decision": candidate["decision"],
        "detail": candidate["detail"],
    }


def _summarize_import_candidates(candidates, sheet_summaries, mode):
    """Maneja summarize import candidates."""
    items = [_serialize_import_candidate(candidate) for candidate in candidates]
    return {
        "mode": mode,
        "ready_count": sum(1 for item in items if item["decision"] == "ready"),
        "skip_count": sum(1 for item in items if item["decision"] == "skip"),
        "error_count": sum(1 for item in items if item["decision"] == "error"),
        "items": items,
        "sheet_summaries": sheet_summaries,
    }


def import_articles_from_excel(user, excel_file, mode="preview"):
    """Maneja import articles from excel."""
    require_role(user, MASTER_ROLES)

    if not excel_file:
        raise InventoryApiError("Excel file is required")

    try:
        workbook = load_workbook(excel_file, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise InventoryApiError("Could not read the Excel file") from exc

    candidates, sheet_summaries = _collect_excel_import_candidates(workbook)
    _classify_import_candidates(candidates)
    summary = _summarize_import_candidates(candidates, sheet_summaries, mode="preview")

    if mode != "confirm":
        return summary

    ready_candidates = [candidate for candidate in candidates if candidate["decision"] == "ready"]
    if not ready_candidates:
        raise InventoryApiError("No hay productos listos para agregar en este Excel.")

    created = []
    errors = []

    for candidate in ready_candidates:
        try:
            article = _create_article_from_import_candidate(user, candidate)
            created.append(
                {
                    "row": candidate["row"],
                    "sheet_name": candidate["sheet_name"],
                    "id": article.id,
                    "internal_code": article.internal_code,
                    "name": article.name,
                }
            )
        except Exception as exc:  # noqa: BLE001
            errors.append(
                {
                    "row": candidate["row"],
                    "sheet_name": candidate["sheet_name"],
                    "detail": _import_error_detail(exc),
                }
            )

    base_errors = [
        {
            "row": item["row"],
            "sheet_name": item["sheet_name"],
            "detail": item["detail"],
        }
        for item in summary["items"]
        if item["decision"] == "error"
    ]

    return {
        **_summarize_import_candidates(candidates, sheet_summaries, mode="confirm"),
        "error_count": len(base_errors) + len(errors),
        "created_count": len(created),
        "created": created,
        "errors": base_errors + errors,
    }


PERSONAL_REPORT_IMPORT_ALIASES = {
    "fecha": "report_date",
    "date": "report_date",
    "report_date": "report_date",
    "fecha_reporte": "report_date",
    "dia": "day_label",
    "day": "day_label",
    "weekday": "day_label",
    "dia_semana": "day_label",
    "actividades": "activities",
    "actividad": "activities",
    "actividades_del_dia": "activities",
    "actividades_dia": "activities",
    "tareas": "activities",
    "detalle": "activities",
}

PERSONAL_REPORT_EXPORT_COLUMNS = (
    "Fecha",
    "Dia",
    "Actividades del dia",
)

SPANISH_WEEKDAY_LABELS = (
    "Lunes",
    "Martes",
    "Miércoles",
    "Jueves",
    "Viernes",
    "Sábado",
    "Domingo",
)


def _personal_report_weekday_label(report_date):
    """Maneja personal report weekday label."""
    return SPANISH_WEEKDAY_LABELS[report_date.weekday()]


def _parse_personal_report_date(value):
    """Maneja parse personal report date."""
    if value in (None, ""):
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, (int, float)):
        try:
            parsed = from_excel(value)
            if isinstance(parsed, datetime):
                return parsed.date()
            if isinstance(parsed, date):
                return parsed
        except Exception:  # noqa: BLE001
            return None

    text = str(value).strip()
    if not text:
        return None

    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    raise InventoryApiError("Fecha invalida. Usa formato YYYY-MM-DD o DD/MM/AAAA.")


def _map_personal_report_columns(headers):
    """Maneja map personal report columns."""
    column_map = {}
    for index, header in enumerate(headers):
        normalized = normalize_excel_header(header)
        if not normalized:
            continue

        target = PERSONAL_REPORT_IMPORT_ALIASES.get(normalized)
        if target and target not in column_map:
            column_map[target] = index

    required = {"report_date", "day_label", "activities"}
    if required.issubset(set(column_map)):
        return column_map
    return None


def serialize_personal_daily_report(report):
    """Maneja serialize personal daily report."""
    return {
        "id": report.id,
        "report_date": report.report_date.isoformat(),
        "day_label": report.day_label,
        "activities": report.activities,
    }


def list_personal_daily_reports(user, limit=120):
    """Lista personal daily reports."""
    items = PersonalDailyReport.objects.filter(user=user).order_by("-report_date")[:limit]
    return [serialize_personal_daily_report(item) for item in items]


def _get_personal_daily_report(user, report_id):
    """Maneja get personal daily report."""
    report = PersonalDailyReport.objects.filter(id=report_id, user=user).first()
    if not report:
        raise InventoryApiError("Informe no encontrado.", status=404)
    return report


def create_personal_daily_report(user, payload):
    """Crea personal daily report."""
    report_date = _parse_personal_report_date(payload.get("report_date") or payload.get("fecha"))
    if not report_date:
        raise InventoryApiError("La fecha es requerida.")

    if PersonalDailyReport.objects.filter(user=user, report_date=report_date).exists():
        raise InventoryApiError("Ya existe un informe para esa fecha.")

    day_label = clean_string(payload.get("day_label") or payload.get("dia") or "")
    if not day_label:
        day_label = _personal_report_weekday_label(report_date)

    activities = clean_string(payload.get("activities") or payload.get("actividades") or "")
    if not activities:
        raise InventoryApiError("Las actividades del dia son requeridas.")

    report = PersonalDailyReport(
        user=user,
        report_date=report_date,
        day_label=day_label,
        activities=activities,
    )
    update_audit(report, user, is_new=True)
    save_validated(report)
    return report


def update_personal_daily_report(user, report_id, payload):
    """Actualiza personal daily report."""
    report = _get_personal_daily_report(user, report_id)

    next_report_date = payload.get("report_date") or payload.get("fecha")
    if next_report_date not in (None, ""):
        parsed = _parse_personal_report_date(next_report_date)
        if not parsed:
            raise InventoryApiError("La fecha es requerida.")

        if parsed != report.report_date and PersonalDailyReport.objects.filter(
            user=user, report_date=parsed
        ).exclude(id=report.id).exists():
            raise InventoryApiError("Ya existe un informe para esa fecha.")

        report.report_date = parsed

    day_label = payload.get("day_label") or payload.get("dia")
    if day_label is not None:
        resolved = clean_string(day_label)
        report.day_label = resolved or _personal_report_weekday_label(report.report_date)

    activities = payload.get("activities") or payload.get("actividades")
    if activities is not None:
        resolved = clean_string(activities)
        if not resolved:
            raise InventoryApiError("Las actividades del dia son requeridas.")
        report.activities = resolved

    update_audit(report, user)
    save_validated(report)
    return report


def delete_personal_daily_report(user, report_id):
    """Elimina personal daily report."""
    report = _get_personal_daily_report(user, report_id)
    report.delete()
    return True


def bulk_delete_personal_daily_reports(user, report_ids=None, delete_all=False):
    """Maneja bulk delete personal daily reports."""
    queryset = PersonalDailyReport.objects.filter(user=user)

    if delete_all:
        deleted_count, _ = queryset.delete()
        return {"deleted_count": deleted_count, "missing_count": 0}

    if report_ids in (None, ""):
        raise InventoryApiError("ids are required")

    if not isinstance(report_ids, list):
        raise InventoryApiError("ids must be a list")

    if not report_ids:
        raise InventoryApiError("ids are required")

    try:
        unique_ids = sorted({int(item) for item in report_ids})
    except (TypeError, ValueError) as exc:
        raise InventoryApiError("Invalid ids") from exc

    matched_queryset = queryset.filter(id__in=unique_ids)
    matched_count = matched_queryset.count()
    deleted_count, _ = matched_queryset.delete()

    return {
        "deleted_count": deleted_count,
        "missing_count": max(0, len(unique_ids) - matched_count),
    }


def import_personal_daily_reports_from_excel(user, excel_file):
    """Maneja import personal daily reports from excel."""
    if not excel_file:
        raise InventoryApiError("Excel file is required")

    try:
        workbook = load_workbook(excel_file, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise InventoryApiError("Could not read the Excel file") from exc

    sheet_matches = []

    for sheet_index, sheet in enumerate(workbook.worksheets):
        for row_number in range(1, min(sheet.max_row, 20) + 1):
            headers = [cell.value for cell in sheet[row_number]]
            mapped = _map_personal_report_columns(headers)
            if not mapped:
                continue

            normalized_title = normalize_excel_header(sheet.title)
            score = 0
            if "informe" in normalized_title or "report" in normalized_title:
                score = 10
            elif "personal" in normalized_title:
                score = 5

            sheet_matches.append((score, sheet_index, sheet, row_number, mapped))
            break

    if not sheet_matches:
        raise InventoryApiError(
            "No se encontro una hoja con columnas: Fecha, Dia y Actividades del dia."
        )

    sheet_matches.sort(key=lambda item: (-item[0], item[1]))
    _, _, resolved_sheet, header_row, column_map = sheet_matches[0]

    candidates = []
    created_count = 0
    updated_count = 0

    current_report = None
    current_activities = []

    def resolve_value(row_values, key):
        """Maneja resolve value."""
        index = column_map.get(key)
        if index is None or index >= len(row_values):
            return None
        return row_values[index]

    def finalize_current_report():
        """Maneja finalize current report."""
        nonlocal current_report, current_activities

        if not current_report:
            return

        candidate = {
            "row": current_report["row"],
            "sheet_name": current_report["sheet_name"],
            "decision": "ready",
            "detail": "",
            "id": None,
            "report_date": current_report["report_date"],
            "day_label": current_report["day_label"],
            "activities": "",
        }

        activities = "\n".join(line for line in current_activities if line)
        if not activities.strip():
            candidate["decision"] = "error"
            candidate["detail"] = "Las actividades del dia son requeridas."
        else:
            candidate["activities"] = activities.strip()

        candidates.append(candidate)
        current_report = None
        current_activities = []

    for row_number in range(header_row + 1, resolved_sheet.max_row + 1):
        row_values = [cell.value for cell in resolved_sheet[row_number]]

        report_date_value = resolve_value(row_values, "report_date")
        day_label_value = resolve_value(row_values, "day_label")
        activities_value = resolve_value(row_values, "activities")

        if all(_import_is_blank(value) for value in (report_date_value, day_label_value, activities_value)):
            continue

        has_date = not _import_is_blank(report_date_value)
        has_activities = not _import_is_blank(activities_value)

        if has_date:
            finalize_current_report()

            try:
                report_date = _parse_personal_report_date(report_date_value)
                if not report_date:
                    raise InventoryApiError("La fecha es requerida.")

                day_label = _import_text(day_label_value)
                if not day_label:
                    day_label = _personal_report_weekday_label(report_date)

                current_report = {
                    "row": row_number,
                    "sheet_name": resolved_sheet.title,
                    "report_date": report_date,
                    "day_label": day_label,
                }

                if has_activities:
                    current_activities.append(_import_text(activities_value))
            except Exception as exc:  # noqa: BLE001
                candidates.append(
                    {
                        "row": row_number,
                        "sheet_name": resolved_sheet.title,
                        "decision": "error",
                        "detail": _import_error_detail(exc),
                        "id": None,
                        "report_date": None,
                        "day_label": "",
                        "activities": "",
                    }
                )
                current_report = None
                current_activities = []

            continue

        if has_activities:
            if current_report:
                current_activities.append(_import_text(activities_value))
            else:
                candidates.append(
                    {
                        "row": row_number,
                        "sheet_name": resolved_sheet.title,
                        "decision": "error",
                        "detail": "Actividad sin fecha.",
                        "id": None,
                        "report_date": None,
                        "day_label": "",
                        "activities": _import_text(activities_value),
                    }
                )

    finalize_current_report()

    ready_candidates = [item for item in candidates if item["decision"] == "ready"]
    if not ready_candidates:
        raise InventoryApiError("No hay filas listas para importar en este Excel.")

    with transaction.atomic():
        for candidate in ready_candidates:
            try:
                report_date = candidate["report_date"]
                report = PersonalDailyReport.objects.filter(
                    user=user,
                    report_date=report_date,
                ).first()

                if report:
                    report.day_label = candidate["day_label"]
                    report.activities = candidate["activities"]
                    update_audit(report, user)
                    save_validated(report)
                    updated_count += 1
                    candidate["decision"] = "updated"
                else:
                    report = PersonalDailyReport(
                        user=user,
                        report_date=report_date,
                        day_label=candidate["day_label"],
                        activities=candidate["activities"],
                    )
                    update_audit(report, user, is_new=True)
                    save_validated(report)
                    created_count += 1
                    candidate["decision"] = "created"

                candidate["id"] = report.id
            except Exception as exc:  # noqa: BLE001
                candidate["decision"] = "error"
                candidate["detail"] = _import_error_detail(exc)

    serialized_items = []
    error_count = 0
    for item in candidates:
        report_date_value = item.get("report_date")
        if isinstance(report_date_value, date):
            report_date_value = report_date_value.isoformat()

        serialized_items.append(
            {
                "row": item["row"],
                "sheet_name": item["sheet_name"],
                "id": item.get("id"),
                "report_date": report_date_value,
                "day_label": item.get("day_label") or "",
                "activities": item.get("activities") or "",
                "decision": item.get("decision"),
                "detail": item.get("detail") or "",
            }
        )
        if item.get("decision") == "error":
            error_count += 1

    return {
        "filename": getattr(excel_file, "name", "") or "",
        "created_count": created_count,
        "updated_count": updated_count,
        "error_count": error_count,
        "items": serialized_items,
    }


def _excel_safe_text(value):
    """Maneja excel safe text."""
    text = "" if value in (None, "") else str(value)
    if text and text[0] in ("=", "+", "-", "@"):
        return f"'{text}"
    return text


def _parse_personal_report_export_ids(filters):
    """Maneja parse personal report export ids."""
    if not filters:
        return None

    raw_values = []

    if hasattr(filters, "getlist"):
        raw_values = filters.getlist("ids") or filters.getlist("report_ids") or []
        if not raw_values:
            maybe = filters.get("ids") or filters.get("report_ids") or ""
            raw_values = [maybe] if maybe else []
    elif isinstance(filters, dict):
        raw_values = filters.get("ids") or filters.get("report_ids") or []
        if isinstance(raw_values, str):
            raw_values = [raw_values]
    else:
        return None

    resolved = []
    for raw in raw_values:
        if raw in (None, ""):
            continue
        if isinstance(raw, int):
            resolved.append(raw)
            continue
        if isinstance(raw, str) and "," in raw:
            resolved.extend([chunk.strip() for chunk in raw.split(",") if chunk.strip()])
            continue
        resolved.append(raw)

    if not resolved:
        return None

    try:
        unique_ids = sorted({int(item) for item in resolved})
    except (TypeError, ValueError) as exc:
        raise InventoryApiError("Invalid ids") from exc

    if not unique_ids:
        return None

    return unique_ids


def build_personal_daily_reports_export_excel(user, filters=None):
    """Construye personal daily reports export excel."""
    report_ids = _parse_personal_report_export_ids(filters)

    queryset = PersonalDailyReport.objects.filter(user=user)
    if report_ids:
        queryset = queryset.filter(id__in=report_ids)

    reports = list(queryset.order_by("-report_date", "-id"))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Informes personal"
    sheet.append(PERSONAL_REPORT_EXPORT_COLUMNS)

    for report in reports:
        raw_lines = (report.activities or "").splitlines() or [""]
        lines = [line.rstrip() for line in raw_lines]
        first_line = lines[0] if lines else ""

        sheet.append(
            [
                report.report_date,
                _excel_safe_text(report.day_label or ""),
                _excel_safe_text(first_line),
            ]
        )

        for extra_line in lines[1:]:
            sheet.append(["", "", _excel_safe_text(extra_line)])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(PERSONAL_REPORT_EXPORT_COLUMNS))}{sheet.max_row}"

    column_widths = [14, 16, 80]
    for index, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    output = BytesIO()
    workbook.save(output)

    suffix = "seleccion" if report_ids else "completo"
    return f"informes-personal-{suffix}-{timezone.localdate().isoformat()}.xlsx", output.getvalue()


def list_articles():
    """Lista articles."""
    quantity_map, available_quantity_map, unit_total_map, unit_available_map = current_stock_maps()
    articles = Article.objects.select_related(
        "unit_of_measure",
        "sector_responsible",
        "primary_location",
        "category",
        "subcategory",
        "supplier",
    ).order_by("name")
    return [
        serialize_article(
            article,
            quantity_map,
            available_quantity_map,
            unit_total_map,
            unit_available_map,
        )
        for article in articles
    ]


def article_matches_stock_query(article, query):
    """Maneja article matches stock query."""
    target = build_search_target(
        [
            article.get("name"),
            article.get("internal_code"),
            article.get("article_type_label"),
            article.get("status_label"),
            article.get("category"),
            article.get("subcategory"),
            article.get("primary_location"),
            article.get("sector_responsible"),
        ]
    )
    return matches_normalized_query(target, query)


def article_matches_stock_alert(article, alert_filter):
    """Maneja article matches stock alert."""
    normalized_filter = clean_casefold(alert_filter or "all")
    current_stock = article.get("current_stock") or 0

    if normalized_filter == "low":
        return article.get("low_stock", False)
    if normalized_filter == "healthy":
        return not article.get("low_stock", False) and current_stock > 0
    if normalized_filter == "out":
        return current_stock <= 0
    return True


def filter_articles_for_stock_view(articles, filters=None):
    """Maneja filter articles for stock view."""
    filters = filters or {}
    global_query = filters.get("global_query", "")
    stock_query = filters.get("stock_query", "")
    article_type_filter = clean_casefold(filters.get("article_type") or "all")
    status_filter = clean_casefold(filters.get("status") or "all")
    alert_filter = filters.get("alert") or "all"

    return [
        article
        for article in articles
        if article_matches_stock_query(article, global_query)
        and article_matches_stock_query(article, stock_query)
        and (article_type_filter == "all" or clean_casefold(article.get("article_type")) == article_type_filter)
        and (status_filter == "all" or clean_casefold(article.get("status")) == status_filter)
        and article_matches_stock_alert(article, alert_filter)
    ]


def get_article_stock_label(article):
    """Devuelve article stock label."""
    if article.get("minimum_stock") is None:
        return "Sin minimo"
    if (article.get("current_stock") or 0) <= 0:
        return "Sin stock"
    if article.get("low_stock"):
        return "Bajo minimo"
    return "En nivel"


STOCK_EXPORT_COLUMNS = (
    "Codigo interno",
    "Articulo",
    "Tipo",
    "Unidad",
    "Stock actual",
    "Stock disponible",
    "Stock minimo",
    "Ubicacion principal",
    "Sector responsable",
    "Estado stock",
    "Estado articulo",
)


def build_stock_export_excel(filters=None):
    """Construye stock export excel."""
    articles = filter_articles_for_stock_view(list_articles(), filters=filters)
    return build_stock_export_excel_from_articles(articles)


def build_stock_export_excel_from_articles(articles):
    """Construye stock export excel from articles."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Stock"
    sheet.append(STOCK_EXPORT_COLUMNS)

    for article in articles:
        sheet.append(
            [
                article.get("internal_code") or "",
                article.get("name") or "",
                article.get("article_type_label") or "",
                article.get("unit_of_measure", {}).get("code")
                or article.get("unit_of_measure", {}).get("name")
                or "",
                article.get("current_stock"),
                article.get("available_stock"),
                article.get("minimum_stock"),
                article.get("primary_location") or "",
                article.get("sector_responsible") or "",
                get_article_stock_label(article),
                article.get("status_label") or "",
            ]
        )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(STOCK_EXPORT_COLUMNS))}{sheet.max_row}"

    column_widths = [18, 34, 24, 12, 14, 16, 14, 22, 22, 18, 18]
    for index, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=5, max_col=7):
        for cell in row:
            cell.number_format = "0.###"

    output = BytesIO()
    workbook.save(output)
    return f"stock-{timezone.localdate().isoformat()}.xlsx", output.getvalue()
