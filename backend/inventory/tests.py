import json
import threading
from datetime import datetime, timedelta
from decimal import Decimal
from io import BytesIO
from tempfile import TemporaryDirectory
from unittest import mock

from django.contrib.auth import get_user_model
from django.core import mail
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import close_old_connections
from django.test import TestCase, TransactionTestCase, override_settings
from django.utils import timezone
from openpyxl.styles import Font, PatternFill
from openpyxl import Workbook, load_workbook

from accounts.models import UserProfile
from inventory.automation import (
    InventoryAutomationRunner,
    TASK_KEY_FULL_STOCK_REPORT,
    TASK_KEY_MINIMUM_STOCK_DIGEST,
    TASK_KEY_MINIMUM_STOCK_RECONCILE,
    TASK_KEY_SCHEDULER,
    claim_full_stock_report_period,
    claim_minimum_stock_digest_period,
    ensure_automation_task_state,
    get_full_stock_report_due_context,
    get_minimum_stock_digest_due_context,
    maybe_start_inventory_automation,
    renew_lease,
    reset_inventory_automation_runner_for_tests,
    should_bootstrap_inventory_automation,
    try_acquire_lease,
)
from inventory.models import (
    Article,
    AssetCheckout,
    FullStockReportConfig,
    InventoryBalance,
    InventoryAutomationTaskState,
    Location,
    MinimumStockDigestConfig,
    Person,
    SafetyStockAlertRule,
    Sector,
    StockDiscrepancy,
    TrackedUnit,
    UnitOfMeasure,
)
from inventory.services import dispatch_full_stock_report, dispatch_minimum_stock_digest


class InventoryApiTests(TestCase):
    def build_excel_upload(self, workbook, name="articulos.xlsx"):
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return SimpleUploadedFile(
            name,
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def read_excel_response(self, response):
        return load_workbook(BytesIO(response.content))

    def setUp(self):
        self.media_dir = TemporaryDirectory()
        self.override_media = override_settings(MEDIA_ROOT=self.media_dir.name)
        self.override_media.enable()

        user_model = get_user_model()
        self.storekeeper = user_model.objects.create_user(
            username="storekeeper",
            password="StrongPass123",
            email="storekeeper@inventary.local",
        )
        self.storekeeper.profile.role = UserProfile.Role.STOREKEEPER
        self.storekeeper.profile.save(update_fields=["role"])

        self.supervisor = user_model.objects.create_user(
            username="supervisor",
            password="StrongPass123",
            email="supervisor@inventary.local",
        )
        self.supervisor.profile.role = UserProfile.Role.SUPERVISOR
        self.supervisor.profile.save(update_fields=["role"])

        self.location = Location.objects.get(code="DEP-PRINCIPAL")
        self.person = Person.objects.first()

    def tearDown(self):
        self.override_media.disable()
        self.media_dir.cleanup()
        super().tearDown()

    def test_inventory_endpoint_requires_authentication(self):
        response = self.client.get("/api/inventory/overview/")
        self.assertEqual(response.status_code, 401)

    def test_inventory_overview_returns_new_workspace_shape(self):
        self.client.force_login(self.storekeeper)

        response = self.client.get("/api/inventory/overview/")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertIn("articles", payload)
        self.assertIn("catalogs", payload)
        self.assertIn("minimum_stock_digest", payload)
        self.assertIn("full_stock_report", payload)
        self.assertIn("automation_status", payload)
        self.assertGreaterEqual(len(payload["articles"]), 1)
        self.assertIsNone(payload["minimum_stock_digest"]["id"])
        self.assertIsNone(payload["full_stock_report"]["id"])
        self.assertIn("scheduler", payload["automation_status"])
        self.assertIn("full_stock_report", payload["automation_status"])

    def test_consumable_article_requires_minimum_stock(self):
        self.client.force_login(self.storekeeper)
        unit = UnitOfMeasure.objects.get(code="UN")
        sector = Sector.objects.get(code="DEP")

        response = self.client.post(
            "/api/articles/",
            data=json.dumps(
                {
                    "internal_code": "TEST-001",
                    "name": "Articulo sin minimo",
                    "article_type": "consumable",
                    "unit_of_measure_id": unit.id,
                    "sector_responsible_id": sector.id,
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 400)

    def test_article_code_is_auto_generated_when_missing(self):
        self.client.force_login(self.storekeeper)
        unit = UnitOfMeasure.objects.get(code="UN")
        sector = Sector.objects.get(code="DEP")

        response = self.client.post(
            "/api/articles/",
            data=json.dumps(
                {
                    "name": "Guante nitrilo importado",
                    "article_type": "consumable",
                    "unit_of_measure_id": unit.id,
                    "sector_responsible_id": sector.id,
                    "minimum_stock": "10",
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 201)
        payload = response.json()["item"]
        self.assertTrue(payload["internal_code"].startswith("CON-"))
        self.assertTrue(
            Article.objects.filter(
                internal_code=payload["internal_code"],
                name="Guante nitrilo importado",
            ).exists()
        )

    def test_article_detail_can_update_fields_and_image(self):
        self.client.force_login(self.storekeeper)
        article = Article.objects.filter(tracking_mode=Article.TrackingMode.QUANTITY).first()
        upload = SimpleUploadedFile(
            "producto.png",
            b"fake-image-content",
            content_type="image/png",
        )

        response = self.client.post(
            f"/api/articles/{article.id}/",
            data={
                "name": "Articulo con ficha",
                "description": "Descripcion actualizada",
                "image": upload,
            },
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()["item"]["article"]
        self.assertEqual(payload["name"], "Articulo con ficha")
        self.assertEqual(payload["description"], "Descripcion actualizada")
        self.assertIsNotNone(payload["image_url"])

        clear_response = self.client.post(
            f"/api/articles/{article.id}/",
            data={"clear_image": "true"},
        )
        self.assertEqual(clear_response.status_code, 200)
        self.assertIsNone(clear_response.json()["item"]["article"]["image_url"])

    def test_excel_import_requires_preview_then_confirm(self):
        self.client.force_login(self.storekeeper)

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["nombre", "tipo", "unidad", "sector", "stock minimo", "stock inicial"])
        sheet.append(["Lubricante cadena", "consumible", "UN", "DEP", 5, 12])

        response = self.client.post(
            "/api/articles/import-excel/",
            data={
                "mode": "preview",
                "file": self.build_excel_upload(workbook),
            },
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()["item"]
        self.assertEqual(payload["mode"], "preview")
        self.assertEqual(payload["ready_count"], 1)
        self.assertEqual(payload["error_count"], 0)
        self.assertFalse(Article.objects.filter(name="Lubricante cadena").exists())

        confirm_response = self.client.post(
            "/api/articles/import-excel/",
            data={
                "mode": "confirm",
                "file": self.build_excel_upload(workbook),
            },
        )

        self.assertEqual(confirm_response.status_code, 201)
        payload = confirm_response.json()["item"]
        self.assertEqual(payload["mode"], "confirm")
        self.assertEqual(payload["created_count"], 1)
        self.assertTrue(Article.objects.filter(name="Lubricante cadena").exists())

    def test_excel_import_accepts_simple_inventory_list(self):
        self.client.force_login(self.storekeeper)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Inventario pañol"
        sheet["A1"] = "nombre"
        sheet["A2"] = "Elementos de proteccion personal"
        sheet["A2"].font = Font(bold=True)
        sheet["A2"].fill = PatternFill(fill_type="solid")
        sheet["A3"] = "Guantes moteados"
        sheet["A4"] = "Cofias"
        sheet["A5"] = "Insumos mantenimiento"
        sheet["A5"].font = Font(bold=True)
        sheet["A5"].fill = PatternFill(fill_type="solid")
        sheet["A6"] = "Rodamiento 6206"

        preview_response = self.client.post(
            "/api/articles/import-excel/",
            data={
                "mode": "preview",
                "file": self.build_excel_upload(workbook, name="panol.xlsx"),
            },
        )

        self.assertEqual(preview_response.status_code, 200)
        preview_payload = preview_response.json()["item"]
        self.assertEqual(preview_payload["ready_count"], 3)
        self.assertEqual(preview_payload["error_count"], 0)
        self.assertEqual(preview_payload["sheet_summaries"][0]["mode"], "simple_list")

        confirm_response = self.client.post(
            "/api/articles/import-excel/",
            data={
                "mode": "confirm",
                "file": self.build_excel_upload(workbook, name="panol.xlsx"),
            },
        )

        self.assertEqual(confirm_response.status_code, 201)
        self.assertTrue(Article.objects.filter(name="Guantes moteados").exists())
        self.assertTrue(Article.objects.filter(name="Cofias").exists())
        self.assertTrue(Article.objects.filter(name="Rodamiento 6206").exists())
        self.assertEqual(
            Article.objects.get(name="Guantes moteados").article_type,
            Article.ArticleType.PPE,
        )
        self.assertEqual(
            Article.objects.get(name="Rodamiento 6206").article_type,
            Article.ArticleType.SPARE_PART,
        )

    def test_stock_excel_export_requires_authentication(self):
        response = self.client.get("/api/articles/export-excel/")
        self.assertEqual(response.status_code, 401)

    def test_stock_excel_export_returns_expected_workbook(self):
        self.client.force_login(self.storekeeper)
        article = Article.objects.filter(tracking_mode=Article.TrackingMode.QUANTITY).first()
        balance = InventoryBalance.objects.get(article=article, location=self.location)

        article.internal_code = "EXP-UNIQUE-001"
        article.name = "Exportacion Stock Visible QA"
        article.article_type = Article.ArticleType.CONSUMABLE
        article.status = Article.ArticleStatus.ACTIVE
        article.primary_location = self.location
        article.minimum_stock = Decimal("2")
        article.save(
            update_fields=[
                "internal_code",
                "name",
                "article_type",
                "status",
                "primary_location",
                "minimum_stock",
            ]
        )

        balance.on_hand = Decimal("12")
        balance.reserved = Decimal("3")
        balance.save(update_fields=["on_hand", "reserved"])

        response = self.client.get(
            "/api/articles/export-excel/",
            data={"stock_query": "EXP-UNIQUE-001"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn(
            f'filename="stock-{timezone.localdate().isoformat()}.xlsx"',
            response["Content-Disposition"],
        )

        workbook = self.read_excel_response(response)
        sheet = workbook["Stock"]
        headers = [cell.value for cell in sheet[1]]
        self.assertEqual(
            headers,
            [
                "Codigo interno",
                "Articulo",
                "Tipo",
                "Unidad",
                "Stock actual",
                "Stock disponible",
                "Stock minimo",
                "Ubicacion principal",
                "Sector responsable",
                "Estado stock",
                "Estado articulo",
            ],
        )

        row = [cell.value for cell in sheet[2]]
        self.assertEqual(
            row,
            [
                "EXP-UNIQUE-001",
                "Exportacion Stock Visible QA",
                article.get_article_type_display(),
                article.unit_of_measure.code,
                12,
                9,
                2,
                self.location.name,
                article.sector_responsible.name,
                "En nivel",
                article.get_status_display(),
            ],
        )

    def test_stock_excel_export_applies_filters_and_normalizes_queries(self):
        self.client.force_login(self.storekeeper)
        article = Article.objects.filter(tracking_mode=Article.TrackingMode.QUANTITY).first()
        balance = InventoryBalance.objects.get(article=article, location=self.location)

        article.internal_code = "EXP-FILTER-001"
        article.name = "Cafe Arbol Exportable"
        article.article_type = Article.ArticleType.CONSUMABLE
        article.status = Article.ArticleStatus.ACTIVE
        article.primary_location = self.location
        article.minimum_stock = Decimal("5")
        article.save(
            update_fields=[
                "internal_code",
                "name",
                "article_type",
                "status",
                "primary_location",
                "minimum_stock",
            ]
        )

        balance.on_hand = Decimal("0")
        balance.reserved = Decimal("0")
        balance.save(update_fields=["on_hand", "reserved"])

        response = self.client.get(
            "/api/articles/export-excel/",
            data={
                "global_query": "CAFÉ",
                "stock_query": "árbol",
                "article_type": "consumable",
                "status": "active",
                "alert": "out",
            },
        )

        self.assertEqual(response.status_code, 200)
        workbook = self.read_excel_response(response)
        sheet = workbook["Stock"]

        self.assertEqual(sheet.max_row, 2)
        self.assertEqual(sheet["A2"].value, "EXP-FILTER-001")
        self.assertEqual(sheet["B2"].value, "Cafe Arbol Exportable")
        self.assertEqual(sheet["J2"].value, "Sin stock")

    def test_stock_excel_export_returns_headers_when_no_articles_match(self):
        self.client.force_login(self.storekeeper)

        response = self.client.get(
            "/api/articles/export-excel/",
            data={"stock_query": "NO-EXISTE-EXPORT-XYZ"},
        )

        self.assertEqual(response.status_code, 200)
        workbook = self.read_excel_response(response)
        sheet = workbook["Stock"]

        self.assertEqual(sheet.max_row, 1)
        self.assertEqual(sheet.max_column, 11)

    def test_minimum_stock_digest_config_can_be_saved(self):
        self.client.force_login(self.storekeeper)

        response = self.client.post(
            "/api/inventory/minimum-stock-digest/",
            data=json.dumps(
                {
                    "is_enabled": True,
                    "frequency": "weekly",
                    "run_at": "07:30",
                    "run_weekday": 3,
                    "recipient_user_ids": [self.supervisor.id],
                    "notes": "Resumen semanal de articulos criticos.",
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 201)
        payload = response.json()["item"]
        self.assertEqual(payload["frequency"], "weekly")
        self.assertEqual(payload["frequency_label"], "Semanal")
        self.assertEqual(payload["run_at"], "07:30")
        self.assertEqual(payload["run_weekday"], 3)
        self.assertEqual(len(payload["recipients"]), 1)
        self.assertEqual(payload["recipients"][0]["id"], self.supervisor.id)

        config = MinimumStockDigestConfig.objects.get(key="default")
        self.assertTrue(config.is_enabled)
        self.assertEqual(config.frequency, MinimumStockDigestConfig.Frequency.WEEKLY)
        self.assertEqual(config.run_at.strftime("%H:%M"), "07:30")
        self.assertEqual(config.run_weekday, 3)
        self.assertEqual(config.notes, "Resumen semanal de articulos criticos.")

    def test_full_stock_report_config_can_be_saved(self):
        self.client.force_login(self.storekeeper)

        response = self.client.post(
            "/api/inventory/full-stock-report/",
            data=json.dumps(
                {
                    "is_enabled": True,
                    "frequency": "daily",
                    "run_at": "06:15",
                    "recipient_user_ids": [self.supervisor.id],
                    "notes": "Reporte diario de stock completo.",
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 201)
        payload = response.json()["item"]
        self.assertEqual(payload["frequency"], "daily")
        self.assertEqual(payload["frequency_label"], "Diario")
        self.assertEqual(payload["run_at"], "06:15")
        self.assertEqual(payload["run_weekday"], 0)
        self.assertEqual(len(payload["recipients"]), 1)
        self.assertEqual(payload["recipients"][0]["id"], self.supervisor.id)

        config = FullStockReportConfig.objects.get(key="default")
        self.assertTrue(config.is_enabled)
        self.assertEqual(config.frequency, FullStockReportConfig.Frequency.DAILY)
        self.assertEqual(config.run_at.strftime("%H:%M"), "06:15")
        self.assertEqual(config.run_weekday, 0)
        self.assertEqual(config.notes, "Reporte diario de stock completo.")

    def test_quantity_movement_updates_balance(self):
        self.client.force_login(self.storekeeper)
        article = Article.objects.filter(tracking_mode=Article.TrackingMode.QUANTITY).first()
        balance = InventoryBalance.objects.filter(article=article, location=self.location).first()
        previous_stock = balance.on_hand

        response = self.client.post(
            "/api/movements/",
            data=json.dumps(
                {
                    "article_id": article.id,
                    "movement_type": "purchase_in",
                    "quantity": "5",
                    "target_location_id": self.location.id,
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 201)
        balance.refresh_from_db()
        self.assertEqual(balance.on_hand, previous_stock + Decimal("5"))

    def test_checkout_and_return_updates_tracked_unit(self):
        self.client.force_login(self.storekeeper)
        unit = TrackedUnit.objects.filter(status=TrackedUnit.UnitStatus.AVAILABLE).first()

        checkout_response = self.client.post(
            "/api/checkouts/",
            data=json.dumps(
                {
                    "tracked_unit_id": unit.id,
                    "receiver_person_id": self.person.id,
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(checkout_response.status_code, 201)
        unit.refresh_from_db()
        self.assertEqual(unit.status, TrackedUnit.UnitStatus.CHECKED_OUT)
        self.assertTrue(
            AssetCheckout.objects.filter(
                tracked_unit=unit,
                status=AssetCheckout.CheckoutStatus.OPEN,
            ).exists()
        )

        checkout_id = checkout_response.json()["item"]["id"]
        return_response = self.client.post(
            f"/api/checkouts/{checkout_id}/return/",
            data=json.dumps({}),
            content_type="application/json",
        )

        self.assertEqual(return_response.status_code, 200)
        unit.refresh_from_db()
        self.assertEqual(unit.status, TrackedUnit.UnitStatus.AVAILABLE)
        self.assertEqual(unit.current_location, self.location)

    def test_count_line_creates_discrepancy_and_supervisor_can_resolve_it(self):
        article = Article.objects.filter(tracking_mode=Article.TrackingMode.QUANTITY).first()
        initial_balance = InventoryBalance.objects.get(article=article, location=self.location)

        self.client.force_login(self.storekeeper)
        session_response = self.client.post(
            "/api/counts/",
            data=json.dumps({"count_type": "partial", "scope": "Conteo test"}),
            content_type="application/json",
        )
        self.assertEqual(session_response.status_code, 201)

        line_response = self.client.post(
            f"/api/counts/{session_response.json()['item']['id']}/lines/",
            data=json.dumps(
                {
                    "article_id": article.id,
                    "location_id": self.location.id,
                    "counter_person_id": self.person.id,
                    "counted_qty": "0",
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(line_response.status_code, 201)

        discrepancy = StockDiscrepancy.objects.filter(article=article, status="open").latest("id")
        self.assertEqual(discrepancy.difference_type, StockDiscrepancy.DifferenceType.NEGATIVE)
        self.assertEqual(discrepancy.possible_cause, "")

        self.client.force_login(self.supervisor)
        resolve_response = self.client.post(
            f"/api/discrepancies/{discrepancy.id}/resolve/",
            data=json.dumps({"reason_text": "Conteo validado"}),
            content_type="application/json",
        )

        self.assertEqual(resolve_response.status_code, 200)
        discrepancy.refresh_from_db()
        initial_balance.refresh_from_db()
        self.assertEqual(discrepancy.status, StockDiscrepancy.DiscrepancyStatus.RESOLVED)
        self.assertEqual(initial_balance.on_hand, Decimal("0"))

    @override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend")
    def test_safety_alert_rule_sends_mail_when_article_is_already_below_minimum_stock(self):
        self.client.force_login(self.storekeeper)
        article = Article.objects.filter(tracking_mode=Article.TrackingMode.QUANTITY).first()
        balance = InventoryBalance.objects.get(article=article, location=self.location)
        article.minimum_stock = balance.on_hand + Decimal("1")
        article.save(update_fields=["minimum_stock"])

        response = self.client.post(
            "/api/inventory/safety-alerts/",
            data=json.dumps(
                {
                    "article_id": article.id,
                    "is_enabled": True,
                    "recipient_user_ids": [self.supervisor.id],
                    "notes": "Reponer urgente",
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 201)
        rule = SafetyStockAlertRule.objects.get(article=article)
        self.assertEqual(rule.status, SafetyStockAlertRule.AlertStatus.TRIGGERED)
        self.assertEqual(len(mail.outbox), 1)
        self.assertIn(self.supervisor.email, mail.outbox[0].to)
        self.assertIn(article.internal_code, mail.outbox[0].subject)

    @override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend")
    def test_safety_alert_mail_is_not_resent_until_stock_recovers(self):
        self.client.force_login(self.storekeeper)
        article = Article.objects.filter(tracking_mode=Article.TrackingMode.QUANTITY).first()
        balance = InventoryBalance.objects.get(article=article, location=self.location)
        previous_stock = balance.on_hand
        article.minimum_stock = previous_stock - Decimal("1")
        article.save(update_fields=["minimum_stock"])

        create_rule_response = self.client.post(
            "/api/inventory/safety-alerts/",
            data=json.dumps(
                {
                    "article_id": article.id,
                    "is_enabled": True,
                    "recipient_user_ids": [self.supervisor.id],
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(create_rule_response.status_code, 201)
        self.assertEqual(len(mail.outbox), 0)

        first_drop_qty = (previous_stock - article.minimum_stock) + Decimal("1")
        first_drop_response = self.client.post(
            "/api/movements/",
            data=json.dumps(
                {
                    "article_id": article.id,
                    "movement_type": "consumption_out",
                    "quantity": str(first_drop_qty),
                    "source_location_id": self.location.id,
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(first_drop_response.status_code, 201)
        self.assertEqual(len(mail.outbox), 1)

        second_drop_response = self.client.post(
            "/api/movements/",
            data=json.dumps(
                {
                    "article_id": article.id,
                    "movement_type": "consumption_out",
                    "quantity": "1",
                    "source_location_id": self.location.id,
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(second_drop_response.status_code, 201)
        self.assertEqual(len(mail.outbox), 1)

        replenish_qty = article.minimum_stock + Decimal("3")
        replenish_response = self.client.post(
            "/api/movements/",
            data=json.dumps(
                {
                    "article_id": article.id,
                    "movement_type": "purchase_in",
                    "quantity": str(replenish_qty),
                    "target_location_id": self.location.id,
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(replenish_response.status_code, 201)
        rule = SafetyStockAlertRule.objects.get(article=article)
        self.assertEqual(rule.status, SafetyStockAlertRule.AlertStatus.MONITORING)

        drop_again_response = self.client.post(
            "/api/movements/",
            data=json.dumps(
                {
                    "article_id": article.id,
                    "movement_type": "consumption_out",
                    "quantity": str(replenish_qty),
                    "source_location_id": self.location.id,
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(drop_again_response.status_code, 201)
        self.assertEqual(len(mail.outbox), 2)


@override_settings(INVENTORY_AUTOMATION_ENABLED=False)
class InventoryAutomationLeaseTests(TransactionTestCase):
    def setUp(self):
        reset_inventory_automation_runner_for_tests()
        ensure_automation_task_state(TASK_KEY_SCHEDULER)

    def tearDown(self):
        reset_inventory_automation_runner_for_tests()
        super().tearDown()

    def test_try_acquire_lease_is_atomic_between_threads(self):
        barrier = threading.Barrier(2)
        results = []
        errors = []

        def worker(index):
            close_old_connections()
            try:
                barrier.wait(timeout=5)
                lease = try_acquire_lease(
                    TASK_KEY_SCHEDULER,
                    f"owner-{index}",
                    f"owner-{index}",
                    90,
                    now=timezone.now(),
                )
                results.append(lease.acquired)
            except Exception as exc:  # noqa: BLE001
                errors.append(str(exc))
            finally:
                close_old_connections()

        threads = [threading.Thread(target=worker, args=(index,)) for index in (1, 2)]
        for thread in threads:
            thread.start()
        for thread in threads:
            thread.join()

        self.assertEqual(errors, [])
        self.assertEqual(sum(1 for item in results if item), 1)

    def test_expired_lease_can_be_taken_over(self):
        now = timezone.now()
        first_lease = try_acquire_lease(TASK_KEY_SCHEDULER, "owner-a", "owner-a", 90, now=now)
        self.assertTrue(first_lease.acquired)
        InventoryAutomationTaskState.objects.filter(key=TASK_KEY_SCHEDULER).update(
            lease_expires_at=now - timedelta(seconds=1),
            heartbeat_at=now - timedelta(seconds=5),
        )

        second_lease = try_acquire_lease(
            TASK_KEY_SCHEDULER,
            "owner-b",
            "owner-b",
            90,
            now=now,
        )

        self.assertTrue(second_lease.acquired)
        self.assertTrue(second_lease.takeover)
        task_state = InventoryAutomationTaskState.objects.get(key=TASK_KEY_SCHEDULER)
        self.assertEqual(task_state.owner_token, "owner-b")

    def test_renew_lease_updates_heartbeat(self):
        now = timezone.now()
        acquired = try_acquire_lease(TASK_KEY_SCHEDULER, "owner-a", "owner-a", 90, now=now)
        self.assertTrue(acquired.acquired)
        task_state = InventoryAutomationTaskState.objects.get(key=TASK_KEY_SCHEDULER)
        old_heartbeat = task_state.heartbeat_at

        renewed = renew_lease(
            TASK_KEY_SCHEDULER,
            "owner-a",
            90,
            now=now + timedelta(seconds=5),
        )

        self.assertTrue(renewed)
        task_state.refresh_from_db()
        self.assertGreater(task_state.heartbeat_at, old_heartbeat)

    def test_renew_lease_returns_false_after_takeover(self):
        now = timezone.now()
        acquired = try_acquire_lease(TASK_KEY_SCHEDULER, "owner-a", "owner-a", 90, now=now)
        self.assertTrue(acquired.acquired)
        InventoryAutomationTaskState.objects.filter(key=TASK_KEY_SCHEDULER).update(
            lease_expires_at=now - timedelta(seconds=1),
        )
        takeover = try_acquire_lease(
            TASK_KEY_SCHEDULER,
            "owner-b",
            "owner-b",
            90,
            now=now,
        )
        self.assertTrue(takeover.acquired)

        renewed = renew_lease(
            TASK_KEY_SCHEDULER,
            "owner-a",
            90,
            now=now + timedelta(seconds=1),
        )
        self.assertFalse(renewed)


@override_settings(INVENTORY_AUTOMATION_ENABLED=False)
class InventoryAutomationBehaviorTests(TestCase):
    def setUp(self):
        reset_inventory_automation_runner_for_tests()
        user_model = get_user_model()
        self.storekeeper = user_model.objects.create_user(
            username="auto_storekeeper",
            password="StrongPass123",
            email="auto_storekeeper@inventary.local",
        )
        self.storekeeper.profile.role = UserProfile.Role.STOREKEEPER
        self.storekeeper.profile.save(update_fields=["role"])

        self.supervisor = user_model.objects.create_user(
            username="auto_supervisor",
            password="StrongPass123",
            email="auto_supervisor@inventary.local",
        )
        self.supervisor.profile.role = UserProfile.Role.SUPERVISOR
        self.supervisor.profile.save(update_fields=["role"])

        self.location = Location.objects.get(code="DEP-PRINCIPAL")
        self.sector = Sector.objects.get(code="DEP")
        self.unit = UnitOfMeasure.objects.get(code="UN")

    def tearDown(self):
        reset_inventory_automation_runner_for_tests()
        super().tearDown()

    def create_quantity_article(self, suffix, on_hand="4", minimum_stock="2"):
        article = Article.objects.create(
            internal_code=f"AUTO-{suffix}",
            name=f"Articulo auto {suffix}",
            article_type=Article.ArticleType.CONSUMABLE,
            unit_of_measure=self.unit,
            sector_responsible=self.sector,
            tracking_mode=Article.TrackingMode.QUANTITY,
            minimum_stock=Decimal(str(minimum_stock)),
            primary_location=self.location,
        )
        InventoryBalance.objects.create(
            article=article,
            location=self.location,
            on_hand=Decimal(str(on_hand)),
            reserved=Decimal("0"),
        )
        return article

    def create_digest_config(self, *, recipient_user=None, additional_emails="", **overrides):
        config = MinimumStockDigestConfig.objects.create(
            key="default",
            is_enabled=overrides.get("is_enabled", True),
            frequency=overrides.get("frequency", MinimumStockDigestConfig.Frequency.DAILY),
            run_at=overrides.get("run_at", datetime.strptime("08:00", "%H:%M").time()),
            run_weekday=overrides.get("run_weekday", 0),
            additional_emails=additional_emails,
            notes=overrides.get("notes", ""),
        )
        if recipient_user:
            config.recipients.add(recipient_user)
        return config

    def create_full_stock_report_config(self, *, recipient_user=None, additional_emails="", **overrides):
        config = FullStockReportConfig.objects.create(
            key="default",
            is_enabled=overrides.get("is_enabled", True),
            frequency=overrides.get("frequency", FullStockReportConfig.Frequency.DAILY),
            run_at=overrides.get("run_at", datetime.strptime("08:00", "%H:%M").time()),
            run_weekday=overrides.get("run_weekday", 0),
            additional_emails=additional_emails,
            notes=overrides.get("notes", ""),
        )
        if recipient_user:
            config.recipients.add(recipient_user)
        return config

    def test_should_bootstrap_inventory_automation_skips_unwanted_commands(self):
        self.assertFalse(should_bootstrap_inventory_automation(argv=["manage.py", "test"]))
        self.assertFalse(
            should_bootstrap_inventory_automation(
                argv=["manage.py", "runserver"],
                environ={},
            )
        )
        self.assertTrue(
            should_bootstrap_inventory_automation(
                argv=["manage.py", "runserver"],
                environ={"RUN_MAIN": "true"},
            )
        )

    def test_maybe_start_inventory_automation_only_starts_once_per_process(self):
        class DummyRunner:
            def __init__(self):
                self.started = False

            def start(self):
                self.started = True

            def is_alive(self):
                return self.started

            def stop(self):
                self.started = False

        with mock.patch("inventory.automation.should_bootstrap_inventory_automation", return_value=True):
            with mock.patch("inventory.automation.InventoryAutomationRunner", DummyRunner):
                first = maybe_start_inventory_automation()
                second = maybe_start_inventory_automation()

        self.assertIs(first, second)

    def test_runner_cycle_skips_jobs_without_scheduler_lease(self):
        runner = InventoryAutomationRunner()
        with mock.patch.object(runner, "_ensure_scheduler_lease", return_value=False):
            with mock.patch.object(runner, "run_reconcile_job") as reconcile_mock:
                with mock.patch.object(runner, "run_digest_job_if_due") as digest_mock:
                    with mock.patch.object(runner, "run_full_stock_report_job_if_due") as report_mock:
                        runner.run_cycle()

        reconcile_mock.assert_not_called()
        digest_mock.assert_not_called()
        report_mock.assert_not_called()

    def test_digest_due_context_and_next_run_are_computed(self):
        config = self.create_digest_config(
            recipient_user=self.supervisor,
            frequency=MinimumStockDigestConfig.Frequency.WEEKLY,
            run_at=datetime.strptime("07:30", "%H:%M").time(),
            run_weekday=1,
        )
        now = timezone.make_aware(datetime(2026, 4, 7, 9, 0))

        due_context = get_minimum_stock_digest_due_context(config, now=now)

        self.assertTrue(due_context["due"])
        self.assertEqual(due_context["due_key"], "weekly:2026-04-07")
        self.assertEqual(timezone.localtime(due_context["next_run_at"]).date().isoformat(), "2026-04-14")

    def test_full_stock_report_due_context_and_next_run_are_computed(self):
        config = self.create_full_stock_report_config(
            recipient_user=self.supervisor,
            frequency=FullStockReportConfig.Frequency.WEEKLY,
            run_at=datetime.strptime("07:30", "%H:%M").time(),
            run_weekday=1,
        )
        now = timezone.make_aware(datetime(2026, 4, 7, 9, 0))

        due_context = get_full_stock_report_due_context(config, now=now)

        self.assertTrue(due_context["due"])
        self.assertEqual(due_context["due_key"], "weekly:2026-04-07")
        self.assertEqual(timezone.localtime(due_context["next_run_at"]).date().isoformat(), "2026-04-14")

    def test_claim_minimum_stock_digest_period_supports_stale_takeover(self):
        config = self.create_digest_config(recipient_user=self.supervisor)
        due_key = "daily:2026-04-07"
        now = timezone.now()

        first_claim, first_takeover = claim_minimum_stock_digest_period(config, due_key, now=now)
        self.assertTrue(first_claim)
        self.assertFalse(first_takeover)

        second_claim, second_takeover = claim_minimum_stock_digest_period(
            config,
            due_key,
            now=now + timedelta(minutes=1),
        )
        self.assertFalse(second_claim)
        self.assertFalse(second_takeover)

        MinimumStockDigestConfig.objects.filter(pk=config.pk).update(
            inflight_period_key=due_key,
            inflight_started_at=now - timedelta(hours=2),
            last_delivery_status=MinimumStockDigestConfig.DeliveryStatus.NEVER,
        )
        stale_claim, stale_takeover = claim_minimum_stock_digest_period(
            config,
            due_key,
            now=now,
        )
        self.assertTrue(stale_claim)
        self.assertTrue(stale_takeover)

    def test_claim_full_stock_report_period_supports_stale_takeover(self):
        config = self.create_full_stock_report_config(recipient_user=self.supervisor)
        due_key = "daily:2026-04-07"
        now = timezone.now()

        first_claim, first_takeover = claim_full_stock_report_period(config, due_key, now=now)
        self.assertTrue(first_claim)
        self.assertFalse(first_takeover)

        second_claim, second_takeover = claim_full_stock_report_period(
            config,
            due_key,
            now=now + timedelta(minutes=1),
        )
        self.assertFalse(second_claim)
        self.assertFalse(second_takeover)

        FullStockReportConfig.objects.filter(pk=config.pk).update(
            inflight_period_key=due_key,
            inflight_started_at=now - timedelta(hours=2),
            last_delivery_status=FullStockReportConfig.DeliveryStatus.NEVER,
        )
        stale_claim, stale_takeover = claim_full_stock_report_period(
            config,
            due_key,
            now=now,
        )
        self.assertTrue(stale_claim)
        self.assertTrue(stale_takeover)

    def test_dispatch_digest_records_warning_when_all_recipients_are_discarded(self):
        self.create_quantity_article("WARN", on_hand="1", minimum_stock="2")
        config = self.create_digest_config(recipient_user=self.supervisor)
        self.supervisor.email = ""
        self.supervisor.save(update_fields=["email"])

        with self.assertLogs("inventory.automation.digest", level="WARNING") as captured_logs:
            result = dispatch_minimum_stock_digest(config.pk, "daily:2026-04-07")

        config.refresh_from_db()
        self.assertEqual(
            result["delivery_status"],
            MinimumStockDigestConfig.DeliveryStatus.WARNING,
        )
        self.assertEqual(config.last_delivery_status, MinimumStockDigestConfig.DeliveryStatus.WARNING)
        self.assertIn("sin email", config.last_recipient_warning.lower())
        self.assertEqual(config.last_email_error, "")
        self.assertEqual(config.last_period_key, "daily:2026-04-07")
        self.assertTrue(any("digest_recipients_warning" in line for line in captured_logs.output))

    def test_dispatch_digest_records_real_send_error_separately(self):
        self.create_quantity_article("ERR", on_hand="1", minimum_stock="2")
        config = self.create_digest_config(recipient_user=self.supervisor)

        with mock.patch("inventory.services.send_mail", side_effect=RuntimeError("smtp down")):
            with self.assertLogs("inventory.automation.digest", level="INFO") as info_logs:
                result = dispatch_minimum_stock_digest(config.pk, "daily:2026-04-07")

        config.refresh_from_db()
        self.assertEqual(
            result["delivery_status"],
            MinimumStockDigestConfig.DeliveryStatus.ERROR,
        )
        self.assertEqual(config.last_delivery_status, MinimumStockDigestConfig.DeliveryStatus.ERROR)
        self.assertIn("smtp down", config.last_email_error)
        self.assertEqual(config.last_recipient_warning, "")
        self.assertEqual(config.inflight_period_key, "daily:2026-04-07")
        self.assertEqual(config.last_period_key, "")
        self.assertTrue(any("digest_send_start" in line for line in info_logs.output))

    def test_dispatch_full_stock_report_records_warning_when_all_recipients_are_discarded(self):
        self.create_quantity_article("WARN-REPORT", on_hand="1", minimum_stock="2")
        config = self.create_full_stock_report_config(recipient_user=self.supervisor)
        self.supervisor.email = ""
        self.supervisor.save(update_fields=["email"])

        with self.assertLogs("inventory.automation.full_stock_report", level="WARNING") as captured_logs:
            result = dispatch_full_stock_report(config.pk, "daily:2026-04-07")

        config.refresh_from_db()
        self.assertEqual(
            result["delivery_status"],
            FullStockReportConfig.DeliveryStatus.WARNING,
        )
        self.assertEqual(config.last_delivery_status, FullStockReportConfig.DeliveryStatus.WARNING)
        self.assertIn("sin email", config.last_recipient_warning.lower())
        self.assertEqual(config.last_email_error, "")
        self.assertEqual(config.last_period_key, "daily:2026-04-07")
        self.assertTrue(
            any("stock_report_recipients_warning" in line for line in captured_logs.output)
        )

    def test_dispatch_full_stock_report_sends_email_with_attachment(self):
        self.create_quantity_article("OK-REPORT", on_hand="5", minimum_stock="2")
        config = self.create_full_stock_report_config(recipient_user=self.supervisor)

        result = dispatch_full_stock_report(config.pk, "daily:2026-04-07")

        config.refresh_from_db()
        self.assertEqual(
            result["delivery_status"],
            FullStockReportConfig.DeliveryStatus.SUCCESS,
        )
        self.assertEqual(config.last_delivery_status, FullStockReportConfig.DeliveryStatus.SUCCESS)
        self.assertEqual(config.last_period_key, "daily:2026-04-07")
        self.assertIsNotNone(config.last_notified_at)

        self.assertEqual(len(mail.outbox), 1)
        message = mail.outbox[0]
        self.assertIn("Reporte de stock completo", message.subject)
        self.assertIn(self.supervisor.email, message.to)
        self.assertEqual(len(message.attachments), 1)
        attachment_name, attachment_payload, attachment_type = message.attachments[0]
        self.assertTrue(attachment_name.startswith("stock-"))
        self.assertIn("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", attachment_type)
        self.assertGreater(len(attachment_payload), 200)

    def test_dispatch_full_stock_report_records_real_send_error_separately(self):
        self.create_quantity_article("ERR-REPORT", on_hand="1", minimum_stock="2")
        config = self.create_full_stock_report_config(recipient_user=self.supervisor)

        with mock.patch("inventory.services.EmailMessage.send", side_effect=RuntimeError("smtp down")):
            with self.assertLogs("inventory.automation.full_stock_report", level="INFO") as info_logs:
                result = dispatch_full_stock_report(config.pk, "daily:2026-04-07")

        config.refresh_from_db()
        self.assertEqual(
            result["delivery_status"],
            FullStockReportConfig.DeliveryStatus.ERROR,
        )
        self.assertEqual(config.last_delivery_status, FullStockReportConfig.DeliveryStatus.ERROR)
        self.assertIn("smtp down", config.last_email_error)
        self.assertEqual(config.last_recipient_warning, "")
        self.assertEqual(config.inflight_period_key, "daily:2026-04-07")
        self.assertEqual(config.last_period_key, "")
        self.assertTrue(any("stock_report_send_start" in line for line in info_logs.output))

    @override_settings(INVENTORY_AUTOMATION_BATCH_SIZE=2)
    def test_reconcile_job_marks_warning_when_an_item_fails_mid_batch(self):
        articles = [
            self.create_quantity_article("REC-1"),
            self.create_quantity_article("REC-2"),
            self.create_quantity_article("REC-3"),
        ]
        for article in articles:
            SafetyStockAlertRule.objects.create(article=article, is_enabled=True)

        runner = InventoryAutomationRunner()
        scheduler_lease = try_acquire_lease(
            TASK_KEY_SCHEDULER,
            runner.owner_token,
            runner.owner_label,
            90,
            now=timezone.now(),
        )
        self.assertTrue(scheduler_lease.acquired)
        runner._scheduler_has_lease = True

        processed_ids = []

        def fake_evaluate(article):
            processed_ids.append(article.id)
            if article.internal_code == "AUTO-REC-2":
                raise RuntimeError("boom")
            return None

        with mock.patch("inventory.services.evaluate_safety_stock_alert", side_effect=fake_evaluate):
            with self.assertLogs("inventory.automation.reconcile", level="INFO") as captured_logs:
                runner.run_reconcile_job()

        task_state = InventoryAutomationTaskState.objects.get(key=TASK_KEY_MINIMUM_STOCK_RECONCILE)
        self.assertEqual(
            task_state.last_run_status,
            InventoryAutomationTaskState.LastRunStatus.WARNING,
        )
        self.assertEqual(task_state.last_processed_count, 3)
        self.assertEqual(len(processed_ids), 3)
        self.assertTrue(any("reconcile_batch_progress" in line for line in captured_logs.output))
